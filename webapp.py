import csv
import io
import json
import os
import re
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (
    Response,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import or_, func
from werkzeug.utils import secure_filename

from models import (
    AccountingExport,
    AccountingIntegration,
    AuditLog,
    CustomerCategory,
    CustomerCommunication,
    CustomerCredit,
    CustomerDocument,
    CustomerFeedback,
    CustomerNotification,
    CustomerPortalAccount,
    CustomerTransaction,
    CustomerAnalytics,
    DeliveryType,
    DispatchTrip,
    Driver,
    Expense,
    Loan,
    Location,
    PaymentReceipt,
    RateList,
    SubscriptionPayment,
    SubscriptionPlan,
    Tenant,
    TenantPermission,
    TenantSubscription,
    TransportBill,
    User,
    UserAmendmentRequest,
    Vehicle,
    VehicleDriverAssignment,
    VehicleFuelLog,
    VehicleServiceLog,
    VehicleServiceSchedule,
    Vendor,
    VendorAddress,
    VendorUser,
    VendorFieldPermission,
    app,
    db,
    ensure_tenant_permission_rows,
    generate_bill_name,
    generate_access_token,
    generate_driver_code,
    generate_expense_name,
    generate_loan_name,
    generate_trip_number,
    generate_vendor_code,
)


MODULE_META = {
    "dispatch": {"label": "Dispatch / Trips", "icon": "bi-truck-front"},
    "transport_bills": {"label": "Bilty Generation", "icon": "bi-file-text"},
    "accounting": {"label": "Accounting", "icon": "bi-journal-text"},
    "gps": {"label": "GPS Integration", "icon": "bi-broadcast-pin"},
    "vendors": {"label": "Vendors", "icon": "bi-shop"},
    "vehicles": {"label": "Vehicles", "icon": "bi-car-front"},
    "drivers": {"label": "Drivers", "icon": "bi-person-vcard"},
    "expenses": {"label": "Expenses", "icon": "bi-wallet2"},
    "loans": {"label": "Loans", "icon": "bi-bank"},
    "locations": {"label": "Locations", "icon": "bi-geo-alt"},
    "ratelists": {"label": "Rate List", "icon": "bi-currency-dollar"},
    "delivery_types": {"label": "Delivery Types", "icon": "bi-box-seam"},
    "reports": {"label": "Reports", "icon": "bi-graph-up"},
}

ROLE_LABELS = {
    "superadmin": "Superadmin",
    "tenant_admin": "Tenant Admin",
    "tenant_user": "Tenant User",
    "vendor": "Vendor",
}

# Upload configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
DRIVER_UPLOAD_FOLDER = os.path.join(UPLOAD_FOLDER, 'drivers')
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Create upload directories
os.makedirs(DRIVER_UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_uploaded_file(file, subfolder='', prefix=''):
    """Save uploaded file and return the relative path"""
    if file and file.filename and allowed_file(file.filename):
        # Check file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            flash(f'File too large. Maximum size is 5MB.', 'error')
            return None
        
        filename = secure_filename(file.filename)
        if prefix:
            filename = f"{prefix}_{filename}"
        
        upload_path = os.path.join(DRIVER_UPLOAD_FOLDER, subfolder) if subfolder else DRIVER_UPLOAD_FOLDER
        os.makedirs(upload_path, exist_ok=True)
        
        file_path = os.path.join(upload_path, filename)
        counter = 1
        while os.path.exists(file_path):
            name, ext = os.path.splitext(filename)
            filename = f"{name}_{counter}{ext}"
            file_path = os.path.join(upload_path, filename)
            counter += 1
        
        file.save(file_path)
        # Return relative URL path
        rel_path = os.path.join('uploads', 'drivers', subfolder, filename) if subfolder else os.path.join('uploads', 'drivers', filename)
        return rel_path.replace('\\', '/')
    return None


@app.template_filter('from_json')
def from_json(value):
    """Parse JSON string to Python object"""
    if not value:
        return []
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return []


def export_to_excel(data, headers, filename, dropdowns=None):
    """Export data to Excel file with optional dropdown validation
    
    Args:
        data: List of row data (list of lists)
        headers: List of column headers
        filename: Output filename
        dropdowns: Dict mapping header names to list of allowed values
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Add dropdown validation if specified
    if dropdowns:
        for col_idx, header in enumerate(headers, 1):
            if header in dropdowns and dropdowns[header]:
                col_letter = get_column_letter(col_idx)
                add_dropdown_validation(ws, col_letter, dropdowns[header])
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def csv_response(rows, headers, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def parse_excel_to_json(file_storage):
    """Parse uploaded Excel file to JSON"""
    try:
        wb = load_workbook(file_storage)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Get data from remaining rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers):
                    header = headers[idx]
                    # Convert datetime to string
                    if isinstance(value, datetime):
                        row_dict[header] = value.strftime('%Y-%m-%d')
                    else:
                        row_dict[header] = value
            data.append(row_dict)
        
        return {'success': True, 'headers': headers, 'data': data}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def get_current_tenant_id():
    """Get the current tenant ID based on user session"""
    if is_superadmin():
        # For superadmin, use the selected tenant from session or first available
        return session.get("selected_tenant_id") or (
            Tenant.query.first().id if Tenant.query.first() else None
        )
    else:
        # For tenant users, use their assigned tenant
        user = getattr(g, "current_user", None)
        return user.tenant_id if user else None


def parse_int(value):
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def parse_float(value):
    try:
        return float(value) if value not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def parse_date(value, default_value=None):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date() if value else default_value
    except (TypeError, ValueError):
        return default_value


def parse_datetime_local(value, default_value=None):
    if not value:
        return default_value
    value = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return default_value


def parse_bool(value):
    return value in ("on", "true", "1", "yes", True)


def safe_str(value, default=None):
    """Safely convert Excel value to string, handling integers/floats/dates"""
    if value is None or value == "":
        return default
    # Convert to string and strip whitespace
    result = str(value).strip()
    return result if result else default


def safe_float_excel(value):
    """Safely parse float from Excel value handling currency symbols"""
    if value is None or value == "":
        return 0
    try:
        if isinstance(value, str):
            # Remove currency symbols, commas, and whitespace
            value = value.replace('₹', '').replace('$', '').replace(',', '').replace(' ', '').strip()
        return float(value)
    except (TypeError, ValueError):
        return 0


def record_audit(action, entity, entity_id=None, summary=None, details=None, tenant_id=None):
    try:
        log = AuditLog(
            tenant_id=tenant_id or getattr(entity, "tenant_id", None),
            user_id=g.current_user.id if getattr(g, "current_user", None) else None,
            action=action,
            entity_type=entity.__class__.__name__ if entity is not None else "System",
            entity_id=entity_id or getattr(entity, "id", None),
            summary=summary,
            details=json.dumps(details, default=str) if isinstance(details, (dict, list)) else details,
            ip_address=request.remote_addr,
        )
        db.session.add(log)
    except Exception:
        pass


def ensure_trip_tokens(trip):
    changed = False
    if not trip.customer_tracking_token:
        trip.customer_tracking_token = generate_access_token()
        changed = True
    if not trip.driver_access_token:
        trip.driver_access_token = generate_access_token()
        changed = True
    return changed


def login_limited(key, max_attempts=8, window_seconds=900):
    now = datetime.utcnow().timestamp()
    attempts = [ts for ts in session.get(key, []) if now - ts < window_seconds]
    session[key] = attempts
    return len(attempts) >= max_attempts


def record_login_failure(key):
    attempts = session.get(key, [])
    attempts.append(datetime.utcnow().timestamp())
    session[key] = attempts


def add_dropdown_validation(ws, column_letter, allowed_values, start_row=2, end_row=1000):
    """Add Excel data validation dropdown to a column"""
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Create validation with comma-separated values
    values_str = ",".join([str(v) for v in allowed_values])
    dv = DataValidation(type="list", formula1=f'"{values_str}"', allow_blank=True)
    dv.error = 'Please select from the dropdown list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Select a value from the list'
    dv.promptTitle = 'Dropdown Selection'
    
    ws.add_data_validation(dv)
    dv.add(f'{column_letter}{start_row}:{column_letter}{end_row}')
    return dv


def get_column_letter(col_idx):
    """Convert column index (1-based) to Excel column letter"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_user_by_id(user_id):
    return db.session.get(User, user_id) if user_id else None


def is_superadmin():
    return bool(getattr(g, "current_user", None) and g.current_user.role == "superadmin")


def can_manage_tenants():
    return is_superadmin()


def can_manage_users():
    return bool(
        getattr(g, "current_user", None)
        and g.current_user.role in ("superadmin", "tenant_admin")
    )


def has_permission(module_name, action="view"):
    user = getattr(g, "current_user", None)
    if not user:
        return False
    if user.role == "superadmin":
        return True
    
    # Vendor users have special permission handling
    if user.role == "vendor":
        # Vendor users can access reports module if they have field permissions
        if module_name == "reports":
            # Check if vendor user has any field permissions
            vendor_user = VendorUser.query.filter_by(user_id=user.id).first()
            if vendor_user and vendor_user.field_permissions.count() > 0:
                return True
        return False

    permission = getattr(g, "permission_map", {}).get(module_name)
    if not permission:
        return False
    return bool(getattr(permission, f"can_{action}", False))


def permission_required(module_name, action="view"):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not has_permission(module_name, action):
                flash(
                    f"You do not have permission to {action} {MODULE_META[module_name]['label'].lower()}.",
                    "error",
                )
                return redirect(url_for("dashboard"))
            return func(*args, **kwargs)

        return wrapper

    return decorator


def superadmin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not is_superadmin():
            flash("Only the superadmin can access that page.", "error")
            return redirect(url_for("dashboard"))
        return func(*args, **kwargs)

    return wrapper


def user_admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not can_manage_users():
            flash("You do not have permission to manage users.", "error")
            return redirect(url_for("dashboard"))
        return func(*args, **kwargs)

    return wrapper


def get_tenant_options():
    return Tenant.query.order_by(Tenant.name).all()


def get_default_selected_tenant_id(record=None):
    if record is not None:
        return record.tenant_id
    if is_superadmin():
        requested_tenant_id = parse_int(
            request.form.get("tenant_id") or request.args.get("tenant_id")
        )
        if requested_tenant_id:
            return requested_tenant_id
        first_tenant = Tenant.query.order_by(Tenant.name).first()
        return first_tenant.id if first_tenant else None
    return g.current_user.tenant_id if getattr(g, "current_user", None) else None


def resolve_target_tenant_id(record=None, required=True):
    if is_superadmin():
        tenant_id = parse_int(request.form.get("tenant_id"))
        if tenant_id is None and record is not None:
            tenant_id = record.tenant_id
        if tenant_id is None and required:
            raise ValueError("Tenant is required.")
        tenant = db.session.get(Tenant, tenant_id) if tenant_id else None
        if required and tenant is None:
            raise ValueError("Selected tenant does not exist.")
        return tenant_id

    if not getattr(g, "current_user", None) or not g.current_user.tenant_id:
        raise ValueError("A tenant context is required.")
    return g.current_user.tenant_id


def get_form_tenant_context(record=None):
    return {
        "show_tenant_field": is_superadmin(),
        "tenant_options": get_tenant_options() if is_superadmin() else [],
        "selected_tenant_id": get_default_selected_tenant_id(record),
    }


def scoped_query(model):
    query = model.query
    if is_superadmin():
        tenant_filter = parse_int(request.args.get("tenant_id"))
        if tenant_filter:
            query = query.filter_by(tenant_id=tenant_filter)
    elif g.current_user.role == "vendor":
        # Vendor users can only see their own vendor's data
        vendor_user = VendorUser.query.filter_by(user_id=g.current_user.id).first()
        if vendor_user:
            # For Vendor model, filter by vendor_id
            if model == Vendor:
                query = query.filter(Vendor.id == vendor_user.vendor_id)
            # For other models, filter by tenant_id
            else:
                query = query.filter_by(tenant_id=vendor_user.tenant_id)
        else:
            # If no vendor user association, return empty query
            query = query.filter(model.id == None)
    else:
        query = query.filter_by(tenant_id=g.current_user.tenant_id)
    return query


def get_scoped_record(model, record_id):
    return scoped_query(model).filter_by(id=record_id).first_or_404()


def get_related_record(model, raw_id, tenant_id, label, required=False):
    if raw_id in (None, "", 0):
        if required:
            raise ValueError(f"{label} is required.")
        return None

    record = model.query.filter_by(id=raw_id, tenant_id=tenant_id).first()
    if not record:
        raise ValueError(f"Selected {label.lower()} is not available for that tenant.")
    return record


def get_tenant_filtered_records(model, order_column, tenant_id):
    if tenant_id is None:
        return []
    return model.query.filter_by(tenant_id=tenant_id).order_by(order_column).all()


def is_valid_slug(slug):
    return bool(re.fullmatch(r"[a-z0-9-]+", slug or ""))


def get_manageable_user(record_id):
    query = User.query
    if not is_superadmin():
        query = query.filter_by(tenant_id=g.current_user.tenant_id)
    user = query.filter_by(id=record_id).first_or_404()
    if not is_superadmin() and user.role == "superadmin":
        flash("That user cannot be managed from this account.", "error")
        return None
    return user


PUBLIC_ENDPOINTS = {
    "login",
    "customer_login",
    "customer_logout",
    "customer_dashboard",
    "public_track_trip",
    "driver_trip_update",
    "gps_webhook",
}


@app.before_request
def load_security_context():
    g.original_user = get_user_by_id(session.get("impersonator_id"))
    g.current_user = get_user_by_id(session.get("user_id"))
    g.current_tenant = g.current_user.tenant if g.current_user else None
    g.is_impersonating = bool(g.original_user and g.current_user)
    g.permission_map = {}

    # Load permissions for non-superadmin users
    if g.current_user and g.current_user.role != "superadmin" and g.current_user.role != "vendor" and g.current_user.tenant_id:
        permissions = TenantPermission.query.filter_by(
            tenant_id=g.current_user.tenant_id
        ).all()
        g.permission_map = {permission.module_name: permission for permission in permissions}

    if request.endpoint is None:
        return None
    if request.endpoint == "static" or request.endpoint.startswith("static"):
        return None
    if request.endpoint in PUBLIC_ENDPOINTS:
        return None

    if g.current_user is None:
        return redirect(url_for("login"))

    if not g.current_user.is_active:
        session.clear()
        flash("Your account is inactive. Please contact an administrator.", "error")
        return redirect(url_for("login"))

    # Vendor users don't need tenant permission check - they have their own field-level permissions
    if g.current_user.role != "superadmin" and g.current_user.role != "vendor":
        if not g.current_tenant or not g.current_tenant.is_active:
            session.clear()
            flash("Your tenant is inactive. Please contact the superadmin.", "error")
            return redirect(url_for("login"))

    return None


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(self), geolocation=(self), microphone=()",
    )
    if os.environ.get("ENABLE_HSTS", "false").lower() == "true":
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@app.context_processor
def inject_template_context():
    return {
        "current_user": getattr(g, "current_user", None),
        "current_tenant": getattr(g, "current_tenant", None),
        "original_user": getattr(g, "original_user", None),
        "is_impersonating": getattr(g, "is_impersonating", False),
        "module_meta": MODULE_META,
        "role_labels": ROLE_LABELS,
        "has_permission": has_permission,
        "can_manage_users": can_manage_users,
        "can_manage_tenants": can_manage_tenants,
        "is_superadmin": is_superadmin,
        "getattr": getattr,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if getattr(g, "current_user", None):
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        if login_limited("login_failures"):
            flash("Too many failed login attempts. Please try again later.", "error")
            return render_template("login.html")
        identifier = request.form.get("identifier", "").strip()
        password = request.form.get("password", "")

        user = User.query.filter(
            or_(User.username == identifier, User.email == identifier)
        ).first()

        if not user or not user.check_password(password):
            record_login_failure("login_failures")
            flash("Invalid username/email or password.", "error")
            return render_template("login.html")

        if not user.is_active:
            flash("Your account is inactive. Please contact an administrator.", "error")
            return render_template("login.html")

        if user.role != "superadmin" and (not user.tenant or not user.tenant.is_active):
            flash("Your tenant is inactive. Please contact the superadmin.", "error")
            return render_template("login.html")

        session.clear()
        session["user_id"] = user.id
        session.permanent = True
        flash(f"Welcome back, {user.full_name}!", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out successfully.", "success")
    return redirect(url_for("login"))


@app.route("/impersonate/<int:user_id>")
@superadmin_required
def start_impersonation(user_id):
    target_user = User.query.get_or_404(user_id)
    if target_user.role == "superadmin":
        flash("You cannot impersonate another superadmin.", "error")
        return redirect(url_for("users"))
    if not target_user.is_active:
        flash("That user is inactive.", "error")
        return redirect(url_for("users"))
    if not target_user.tenant_id:
        flash("That user is not assigned to a tenant.", "error")
        return redirect(url_for("users"))

    session["impersonator_id"] = g.current_user.id
    session["user_id"] = target_user.id
    flash(f"You are now impersonating {target_user.full_name}.", "success")
    return redirect(url_for("dashboard"))


@app.route("/impersonate/exit")
def exit_impersonation():
    impersonator_id = session.get("impersonator_id")
    if not impersonator_id:
        flash("You are not impersonating any user.", "error")
        return redirect(url_for("dashboard"))

    session["user_id"] = impersonator_id
    session.pop("impersonator_id", None)
    flash("Returned to the superadmin session.", "success")
    return redirect(url_for("dashboard"))


@app.route("/")
def dashboard():
    # Basic stats
    stats = {
        "total_bills": scoped_query(TransportBill).count()
        if has_permission("transport_bills", "view")
        else 0,
        "active_trips": scoped_query(DispatchTrip)
        .filter(DispatchTrip.status.in_(["planned", "dispatched", "in_transit", "delayed"]))
        .count()
        if has_permission("dispatch", "view")
        else 0,
        "total_vehicles": scoped_query(Vehicle).count()
        if has_permission("vehicles", "view")
        else 0,
        "total_delivery_types": scoped_query(DeliveryType).count()
        if has_permission("delivery_types", "view")
        else 0,
        "total_tenants": Tenant.query.count() if is_superadmin() else 1,
    }

    # Financial stats
    bills = scoped_query(TransportBill).all() if has_permission("transport_bills", "view") else []
    total_freight = sum(b.rate or 0 for b in bills)
    
    # Calculate paid amounts from payment receipts
    payment_receipts = scoped_query(PaymentReceipt).all() if has_permission("transport_bills", "view") else []
    total_paid = sum(p.amount_received or 0 for p in payment_receipts)
    
    # Calculate balance for each bill
    bill_balances = {}
    for receipt in payment_receipts:
        if receipt.bilty_id not in bill_balances:
            bill_balances[receipt.bilty_id] = 0
        bill_balances[receipt.bilty_id] += (receipt.amount_received or 0)
    
    unpaid_bills = []
    total_balance = 0
    for bill in bills:
        paid = bill_balances.get(bill.id, 0)
        balance = (bill.rate or 0) - paid
        if balance > 0:
            unpaid_bills.append(bill)
            total_balance += balance
    
    stats.update({
        "total_freight": total_freight,
        "total_paid": total_paid,
        "total_balance": total_balance,
        "unpaid_bills_count": len(unpaid_bills),
    })

    # Loan stats
    loans = scoped_query(Loan).filter(Loan.status == "Active").all() if has_permission("loans", "view") else []
    total_loan_amount = sum(l.principal_amount or 0 for l in loans)
    total_loan_paid = sum(l.amount_paid or 0 for l in loans)
    total_loan_balance = sum(l.balance_amount or 0 for l in loans)
    upcoming_loan_payments = [
        {
            "date": l.next_emi_due_date,
            "bank_name": l.lender_name,
            "amount": l.emi_amount,
            "loan_name": l.name
        }
        for l in loans
        if l.next_emi_due_date and l.next_emi_due_date >= date.today()
    ]
    upcoming_loan_payments.sort(key=lambda x: x["date"])

    stats.update({
        "total_loans": len(loans),
        "total_loan_amount": total_loan_amount,
        "total_loan_paid": total_loan_paid,
        "total_loan_balance": total_loan_balance,
    })

    # Top 5 vendors by amount
    vendor_totals = {}
    for bill in bills:
        if bill.party_information:
            vendor_totals[bill.party_information] = vendor_totals.get(bill.party_information, 0) + (bill.rate or 0)
    top_vendors = sorted(vendor_totals.items(), key=lambda x: x[1], reverse=True)[:5]

    # Pending deliveries (bills with status not 'delivered')
    pending_deliveries = [b for b in bills if b.status != 'delivered']
    active_trips = (
        scoped_query(DispatchTrip)
        .filter(DispatchTrip.status.in_(["planned", "dispatched", "in_transit", "delayed"]))
        .order_by(DispatchTrip.created_at.desc())
        .limit(5)
        .all()
        if has_permission("dispatch", "view")
        else []
    )

    recent_bills = (
        scoped_query(TransportBill)
        .order_by(TransportBill.created_at.desc())
        .limit(5)
        .all()
        if has_permission("transport_bills", "view")
        else []
    )

    return render_template(
        "dashboard.html",
        stats=stats,
        recent_bills=recent_bills,
        upcoming_loan_payments=upcoming_loan_payments[:5],
        top_vendors=top_vendors,
        unpaid_bills=unpaid_bills[:5],
        pending_deliveries=pending_deliveries[:5],
        active_trips=active_trips,
    )


@app.route("/tenants")
@superadmin_required
def tenants():
    tenant_list = Tenant.query.order_by(Tenant.created_at.desc()).all()
    return render_template("tenants/list.html", tenants=tenant_list)


@app.route("/tenants/create", methods=["GET", "POST"])
@superadmin_required
def create_tenant():
    if request.method == "POST":
        from datetime import datetime, timedelta
        import os
        
        name = request.form.get("name", "").strip()
        slug = request.form.get("slug", "").strip().lower()
        contact_email = request.form.get("contact_email", "").strip() or None
        admin_full_name = request.form.get("admin_full_name", "").strip()
        admin_username = request.form.get("admin_username", "").strip()
        admin_email = request.form.get("admin_email", "").strip()
        admin_password = request.form.get("admin_password", "")
        is_active = request.form.get("is_active") == "on"

        # Company Information
        cin_number = request.form.get("cin_number", "").strip() or None
        pan_number = request.form.get("pan_number", "").strip()
        gstin = request.form.get("gstin", "").strip()
        business_type = request.form.get("business_type", "").strip()
        establishment_date_str = request.form.get("establishment_date", "").strip()
        website_url = request.form.get("website_url", "").strip() or None

        # Contact Information
        primary_phone = request.form.get("primary_phone", "").strip()
        secondary_contact_person = request.form.get("secondary_contact_person", "").strip() or None
        secondary_contact_phone = request.form.get("secondary_contact_phone", "").strip() or None
        emergency_contact_name = request.form.get("emergency_contact_name", "").strip() or None
        emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip() or None

        # Registered Address
        reg_address_line1 = request.form.get("reg_address_line1", "").strip()
        reg_address_line2 = request.form.get("reg_address_line2", "").strip() or None
        reg_city = request.form.get("reg_city", "").strip()
        reg_state = request.form.get("reg_state", "").strip()
        reg_pincode = request.form.get("reg_pincode", "").strip()
        reg_country = request.form.get("reg_country", "").strip() or "India"

        # Billing Address
        billing_address_line1 = request.form.get("billing_address_line1", "").strip() or None
        billing_address_line2 = request.form.get("billing_address_line2", "").strip() or None
        billing_city = request.form.get("billing_city", "").strip() or None
        billing_state = request.form.get("billing_state", "").strip() or None
        billing_pincode = request.form.get("billing_pincode", "").strip() or None
        billing_country = request.form.get("billing_country", "").strip() or "India"
        same_as_registered = request.form.get("same_as_registered") == "on"

        # Bank Details
        bank_name = request.form.get("bank_name", "").strip()
        bank_branch = request.form.get("bank_branch", "").strip()
        account_number = request.form.get("account_number", "").strip()
        account_type = request.form.get("account_type", "Current")
        ifsc_code = request.form.get("ifsc_code", "").strip()
        micr_code = request.form.get("micr_code", "").strip() or None
        upi_id = request.form.get("upi_id", "").strip() or None

        # Financial Details
        credit_limit_str = request.form.get("credit_limit", "0").strip()
        credit_limit = float(credit_limit_str) if credit_limit_str else 0
        payment_terms = request.form.get("payment_terms", "").strip() or None

        # Subscription Details
        subscription_plan = request.form.get("subscription_plan", "Basic")
        max_users = int(request.form.get("max_users", "5"))
        plan_expiry_date_str = request.form.get("plan_expiry_date", "").strip()

        # Validate required fields
        required_fields = [name, slug, admin_full_name, admin_username, admin_email, admin_password,
                          pan_number, gstin, business_type, primary_phone, reg_address_line1,
                          reg_city, reg_state, reg_pincode, bank_name, bank_branch, account_number, ifsc_code]
        
        if not all(required_fields):
            flash("Please complete all required fields.", "error")
            return render_template("tenants/form.html")

        if not is_valid_slug(slug):
            flash("Slug must contain only lowercase letters, numbers, and hyphens.", "error")
            return render_template("tenants/form.html")

        if Tenant.query.filter_by(name=name).first():
            flash("A tenant with that name already exists.", "error")
            return render_template("tenants/form.html")
        if Tenant.query.filter_by(slug=slug).first():
            flash("A tenant with that slug already exists.", "error")
            return render_template("tenants/form.html")
        if User.query.filter_by(username=admin_username).first():
            flash("That admin username is already in use.", "error")
            return render_template("tenants/form.html")
        if User.query.filter_by(email=admin_email).first():
            flash("That admin email is already in use.", "error")
            return render_template("tenants/form.html")

        # Parse dates
        establishment_date = None
        if establishment_date_str:
            try:
                establishment_date = datetime.strptime(establishment_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        plan_expiry_date = None
        if plan_expiry_date_str:
            try:
                plan_expiry_date = datetime.strptime(plan_expiry_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Handle file uploads
        upload_dir = os.path.join('uploads', 'tenants')
        os.makedirs(upload_dir, exist_ok=True)
        
        def save_file(file_field, filename_prefix):
            if file_field and file_field.filename:
                filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_field.filename}"
                filepath = os.path.join(upload_dir, filename)
                file_field.save(filepath)
                return filepath
            return None

        gst_certificate_path = save_file(request.files.get('gst_certificate'), 'gst')
        pan_card_path = save_file(request.files.get('pan_card'), 'pan')
        cin_certificate_path = save_file(request.files.get('cin_certificate'), 'cin')
        cancelled_cheque_path = save_file(request.files.get('cancelled_cheque'), 'cheque')

        tenant = Tenant(
            name=name,
            slug=slug,
            contact_email=contact_email,
            is_active=is_active,
            # Company Information
            cin_number=cin_number,
            pan_number=pan_number,
            gstin=gstin,
            business_type=business_type,
            establishment_date=establishment_date,
            website_url=website_url,
            # Contact Information
            primary_phone=primary_phone,
            secondary_contact_person=secondary_contact_person,
            secondary_contact_phone=secondary_contact_phone,
            emergency_contact_name=emergency_contact_name,
            emergency_contact_phone=emergency_contact_phone,
            # Registered Address
            reg_address_line1=reg_address_line1,
            reg_address_line2=reg_address_line2,
            reg_city=reg_city,
            reg_state=reg_state,
            reg_pincode=reg_pincode,
            reg_country=reg_country,
            # Billing Address
            billing_address_line1=billing_address_line1,
            billing_address_line2=billing_address_line2,
            billing_city=billing_city,
            billing_state=billing_state,
            billing_pincode=billing_pincode,
            billing_country=billing_country,
            same_as_registered=same_as_registered,
            # Bank Details
            bank_name=bank_name,
            bank_branch=bank_branch,
            account_number=account_number,
            account_type=account_type,
            ifsc_code=ifsc_code,
            micr_code=micr_code,
            upi_id=upi_id,
            # Financial Details
            credit_limit=credit_limit,
            payment_terms=payment_terms,
            # Document Uploads
            gst_certificate_path=gst_certificate_path,
            pan_card_path=pan_card_path,
            cin_certificate_path=cin_certificate_path,
            cancelled_cheque_path=cancelled_cheque_path,
            # Subscription Details
            subscription_plan=subscription_plan,
            plan_expiry_date=plan_expiry_date,
            max_users=max_users,
        )
        db.session.add(tenant)
        db.session.flush()

        # Create registered address as primary address
        from models import TenantAddress
        registered_address = TenantAddress(
            tenant_id=tenant.id,
            address_type='Registered',
            address_line1=reg_address_line1,
            address_line2=reg_address_line2,
            city=reg_city,
            state=reg_state,
            pincode=reg_pincode,
            country=reg_country,
            is_primary=True,
            is_active=True
        )
        db.session.add(registered_address)

        # Create billing address if different from registered
        if not same_as_registered and billing_address_line1:
            billing_address = TenantAddress(
                tenant_id=tenant.id,
                address_type='Billing',
                address_line1=billing_address_line1,
                address_line2=billing_address_line2,
                city=billing_city,
                state=billing_state,
                pincode=billing_pincode,
                country=billing_country,
                is_primary=False,
                is_active=True
            )
            db.session.add(billing_address)

        tenant_admin = User(
            username=admin_username,
            email=admin_email,
            full_name=admin_full_name,
            role="tenant_admin",
            tenant_id=tenant.id,
            is_active=True,
        )
        tenant_admin.set_password(admin_password)
        db.session.add(tenant_admin)
        db.session.commit()
        ensure_tenant_permission_rows()

        flash("Tenant and initial tenant admin created successfully.", "success")
        return redirect(url_for("tenant_permissions", id=tenant.id))

    return render_template("tenants/form.html")


@app.route("/tenants/<int:id>/edit", methods=["GET", "POST"])
@superadmin_required
def edit_tenant(id):
    from datetime import datetime
    import os
    from models import TenantAddress
    
    tenant = Tenant.query.get_or_404(id)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        slug = request.form.get("slug", "").strip().lower()
        contact_email = request.form.get("contact_email", "").strip() or None
        is_active = request.form.get("is_active") == "on"

        # Company Information
        cin_number = request.form.get("cin_number", "").strip() or None
        pan_number = request.form.get("pan_number", "").strip()
        gstin = request.form.get("gstin", "").strip()
        business_type = request.form.get("business_type", "").strip()
        establishment_date_str = request.form.get("establishment_date", "").strip()
        website_url = request.form.get("website_url", "").strip() or None

        # Contact Information
        primary_phone = request.form.get("primary_phone", "").strip()
        secondary_contact_person = request.form.get("secondary_contact_person", "").strip() or None
        secondary_contact_phone = request.form.get("secondary_contact_phone", "").strip() or None
        emergency_contact_name = request.form.get("emergency_contact_name", "").strip() or None
        emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip() or None

        # Registered Address
        reg_address_line1 = request.form.get("reg_address_line1", "").strip()
        reg_address_line2 = request.form.get("reg_address_line2", "").strip() or None
        reg_city = request.form.get("reg_city", "").strip()
        reg_state = request.form.get("reg_state", "").strip()
        reg_pincode = request.form.get("reg_pincode", "").strip()
        reg_country = request.form.get("reg_country", "").strip() or "India"

        # Billing Address
        billing_address_line1 = request.form.get("billing_address_line1", "").strip() or None
        billing_address_line2 = request.form.get("billing_address_line2", "").strip() or None
        billing_city = request.form.get("billing_city", "").strip() or None
        billing_state = request.form.get("billing_state", "").strip() or None
        billing_pincode = request.form.get("billing_pincode", "").strip() or None
        billing_country = request.form.get("billing_country", "").strip() or "India"
        same_as_registered = request.form.get("same_as_registered") == "on"

        # Bank Details
        bank_name = request.form.get("bank_name", "").strip()
        bank_branch = request.form.get("bank_branch", "").strip()
        account_number = request.form.get("account_number", "").strip()
        account_type = request.form.get("account_type", "Current")
        ifsc_code = request.form.get("ifsc_code", "").strip()
        micr_code = request.form.get("micr_code", "").strip() or None
        upi_id = request.form.get("upi_id", "").strip() or None

        # Financial Details
        credit_limit_str = request.form.get("credit_limit", "0").strip()
        credit_limit = float(credit_limit_str) if credit_limit_str else 0
        payment_terms = request.form.get("payment_terms", "").strip() or None

        # Subscription Details
        subscription_plan = request.form.get("subscription_plan", "Basic")
        max_users = int(request.form.get("max_users", "5"))
        plan_expiry_date_str = request.form.get("plan_expiry_date", "").strip()

        if not name or not slug:
            flash("Tenant name and slug are required.", "error")
            return render_template("tenants/form.html", tenant=tenant, edit=True)

        if not is_valid_slug(slug):
            flash("Slug must contain only lowercase letters, numbers, and hyphens.", "error")
            return render_template("tenants/form.html", tenant=tenant, edit=True)

        existing_name = Tenant.query.filter_by(name=name).first()
        if existing_name and existing_name.id != tenant.id:
            flash("A tenant with that name already exists.", "error")
            return render_template("tenants/form.html", tenant=tenant, edit=True)

        existing_slug = Tenant.query.filter_by(slug=slug).first()
        if existing_slug and existing_slug.id != tenant.id:
            flash("A tenant with that slug already exists.", "error")
            return render_template("tenants/form.html", tenant=tenant, edit=True)

        # Parse dates
        establishment_date = None
        if establishment_date_str:
            try:
                establishment_date = datetime.strptime(establishment_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        plan_expiry_date = None
        if plan_expiry_date_str:
            try:
                plan_expiry_date = datetime.strptime(plan_expiry_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Handle file uploads
        upload_dir = os.path.join('uploads', 'tenants')
        os.makedirs(upload_dir, exist_ok=True)
        
        def save_file(file_field, filename_prefix, existing_path):
            if file_field and file_field.filename:
                filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_field.filename}"
                filepath = os.path.join(upload_dir, filename)
                file_field.save(filepath)
                return filepath
            return existing_path

        gst_certificate_path = save_file(request.files.get('gst_certificate'), 'gst', tenant.gst_certificate_path)
        pan_card_path = save_file(request.files.get('pan_card'), 'pan', tenant.pan_card_path)
        cin_certificate_path = save_file(request.files.get('cin_certificate'), 'cin', tenant.cin_certificate_path)
        cancelled_cheque_path = save_file(request.files.get('cancelled_cheque'), 'cheque', tenant.cancelled_cheque_path)

        # Update tenant fields
        tenant.name = name
        tenant.slug = slug
        tenant.contact_email = contact_email
        tenant.is_active = is_active
        # Company Information
        tenant.cin_number = cin_number
        tenant.pan_number = pan_number
        tenant.gstin = gstin
        tenant.business_type = business_type
        tenant.establishment_date = establishment_date
        tenant.website_url = website_url
        # Contact Information
        tenant.primary_phone = primary_phone
        tenant.secondary_contact_person = secondary_contact_person
        tenant.secondary_contact_phone = secondary_contact_phone
        tenant.emergency_contact_name = emergency_contact_name
        tenant.emergency_contact_phone = emergency_contact_phone
        # Registered Address
        tenant.reg_address_line1 = reg_address_line1
        tenant.reg_address_line2 = reg_address_line2
        tenant.reg_city = reg_city
        tenant.reg_state = reg_state
        tenant.reg_pincode = reg_pincode
        tenant.reg_country = reg_country
        # Billing Address
        tenant.billing_address_line1 = billing_address_line1
        tenant.billing_address_line2 = billing_address_line2
        tenant.billing_city = billing_city
        tenant.billing_state = billing_state
        tenant.billing_pincode = billing_pincode
        tenant.billing_country = billing_country
        tenant.same_as_registered = same_as_registered
        # Bank Details
        tenant.bank_name = bank_name
        tenant.bank_branch = bank_branch
        tenant.account_number = account_number
        tenant.account_type = account_type
        tenant.ifsc_code = ifsc_code
        tenant.micr_code = micr_code
        tenant.upi_id = upi_id
        # Financial Details
        tenant.credit_limit = credit_limit
        tenant.payment_terms = payment_terms
        # Document Uploads
        tenant.gst_certificate_path = gst_certificate_path
        tenant.pan_card_path = pan_card_path
        tenant.cin_certificate_path = cin_certificate_path
        tenant.cancelled_cheque_path = cancelled_cheque_path
        # Subscription Details
        tenant.subscription_plan = subscription_plan
        tenant.plan_expiry_date = plan_expiry_date
        tenant.max_users = max_users

        # Update or create registered address
        registered_address = TenantAddress.query.filter_by(tenant_id=tenant.id, address_type='Registered').first()
        if registered_address:
            registered_address.address_line1 = reg_address_line1
            registered_address.address_line2 = reg_address_line2
            registered_address.city = reg_city
            registered_address.state = reg_state
            registered_address.pincode = reg_pincode
            registered_address.country = reg_country
        else:
            registered_address = TenantAddress(
                tenant_id=tenant.id,
                address_type='Registered',
                address_line1=reg_address_line1,
                address_line2=reg_address_line2,
                city=reg_city,
                state=reg_state,
                pincode=reg_pincode,
                country=reg_country,
                is_primary=True,
                is_active=True
            )
            db.session.add(registered_address)

        # Update or create billing address
        if not same_as_registered and billing_address_line1:
            billing_address = TenantAddress.query.filter_by(tenant_id=tenant.id, address_type='Billing').first()
            if billing_address:
                billing_address.address_line1 = billing_address_line1
                billing_address.address_line2 = billing_address_line2
                billing_address.city = billing_city
                billing_address.state = billing_state
                billing_address.pincode = billing_pincode
                billing_address.country = billing_country
                billing_address.is_active = True
            else:
                billing_address = TenantAddress(
                    tenant_id=tenant.id,
                    address_type='Billing',
                    address_line1=billing_address_line1,
                    address_line2=billing_address_line2,
                    city=billing_city,
                    state=billing_state,
                    pincode=billing_pincode,
                    country=billing_country,
                    is_primary=False,
                    is_active=True
                )
                db.session.add(billing_address)
        else:
            # Deactivate billing address if same as registered
            billing_address = TenantAddress.query.filter_by(tenant_id=tenant.id, address_type='Billing').first()
            if billing_address:
                billing_address.is_active = False

        db.session.commit()
        ensure_tenant_permission_rows()
        flash("Tenant updated successfully.", "success")
        return redirect(url_for("tenants"))

    return render_template("tenants/form.html", tenant=tenant, edit=True)


@app.route("/tenants/<int:id>/permissions", methods=["GET", "POST"])
@superadmin_required
def tenant_permissions(id):
    tenant = Tenant.query.get_or_404(id)
    ensure_tenant_permission_rows()
    permissions = {
        permission.module_name: permission
        for permission in TenantPermission.query.filter_by(tenant_id=tenant.id).all()
    }

    if request.method == "POST":
        for module_name, _label in MODULE_DEFINITIONS:
            permission = permissions.get(module_name)
            if permission is None:
                permission = TenantPermission(tenant_id=tenant.id, module_name=module_name)
                db.session.add(permission)

            permission.can_view = request.form.get(f"{module_name}__view") == "on"
            permission.can_create = request.form.get(f"{module_name}__create") == "on"
            permission.can_edit = request.form.get(f"{module_name}__edit") == "on"
            permission.can_delete = request.form.get(f"{module_name}__delete") == "on"
            permission.can_export = request.form.get(f"{module_name}__export") == "on"

            if not permission.can_view:
                permission.can_create = False
                permission.can_edit = False
                permission.can_delete = False
                permission.can_export = False

        db.session.commit()
        flash("Tenant permissions updated successfully.", "success")
        return redirect(url_for("tenant_permissions", id=tenant.id))

    return render_template(
        "tenants/permissions.html",
        tenant=tenant,
        module_definitions=MODULE_DEFINITIONS,
        permissions=permissions,
    )


@app.route("/users")
@user_admin_required
def users():
    user_query = User.query.order_by(User.created_at.desc())
    if not is_superadmin():
        user_query = user_query.filter_by(tenant_id=g.current_user.tenant_id)
    user_list = user_query.all()
    return render_template(
        "users/list.html",
        users=user_list,
        show_tenant_column=is_superadmin(),
    )


@app.route("/users/create", methods=["GET", "POST"])
@user_admin_required
def create_user():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role", "tenant_user")
        is_active = request.form.get("is_active") == "on"

        try:
            tenant_id = resolve_target_tenant_id(required=True) if is_superadmin() else g.current_user.tenant_id
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context()
            context.update({"roles": ["tenant_admin", "tenant_user"]})
            return render_template("users/form.html", **context)

        if not all([full_name, username, email, password]):
            flash("Full name, username, email, and password are required.", "error")
            context = get_form_tenant_context()
            context.update({"roles": ["tenant_admin", "tenant_user"]})
            return render_template("users/form.html", **context)

        if role not in ("tenant_admin", "tenant_user"):
            flash("Invalid user role selected.", "error")
            context = get_form_tenant_context()
            context.update({"roles": ["tenant_admin", "tenant_user"]})
            return render_template("users/form.html", **context)

        if User.query.filter_by(username=username).first():
            flash("That username is already in use.", "error")
            context = get_form_tenant_context()
            context.update({"roles": ["tenant_admin", "tenant_user"]})
            return render_template("users/form.html", **context)
        if User.query.filter_by(email=email).first():
            flash("That email is already in use.", "error")
            context = get_form_tenant_context()
            context.update({"roles": ["tenant_admin", "tenant_user"]})
            return render_template("users/form.html", **context)

        user = User(
            full_name=full_name,
            username=username,
            email=email,
            role=role,
            tenant_id=tenant_id,
            is_active=is_active,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        
        # Update subscription usage
        update_subscription_usage(tenant_id)
        
        flash("User created successfully.", "success")
        return redirect(url_for("users"))

    context = get_form_tenant_context()
    context.update({"roles": ["tenant_admin", "tenant_user"]})
    return render_template("users/form.html", **context)


@app.route("/users/<int:id>/edit", methods=["GET", "POST"])
@user_admin_required
def edit_user(id):
    user = get_manageable_user(id)
    if user is None:
        return redirect(url_for("users"))

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        role = request.form.get("role", user.role)
        is_active = request.form.get("is_active") == "on"

        if not all([full_name, username, email]):
            flash("Full name, username, and email are required.", "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        existing_username = User.query.filter_by(username=username).first()
        if existing_username and existing_username.id != user.id:
            flash("That username is already in use.", "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        existing_email = User.query.filter_by(email=email).first()
        if existing_email and existing_email.id != user.id:
            flash("That email is already in use.", "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        if role not in ("tenant_admin", "tenant_user", "superadmin"):
            flash("Invalid user role selected.", "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        if not is_superadmin() and role == "superadmin":
            flash("You cannot assign the superadmin role.", "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        try:
            target_tenant_id = (
                resolve_target_tenant_id(user, required=role != "superadmin")
                if is_superadmin()
                else g.current_user.tenant_id
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(user)
            context.update(
                {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
            )
            return render_template("users/form.html", **context)

        user.full_name = full_name
        user.username = username
        user.email = email
        if user.role != "superadmin":
            user.role = role
            user.tenant_id = target_tenant_id
        user.is_active = is_active
        db.session.commit()
        flash("User updated successfully.", "success")
        return redirect(url_for("users"))

    context = get_form_tenant_context(user)
    context.update(
        {"user_obj": user, "edit": True, "roles": ["tenant_admin", "tenant_user"]}
    )
    return render_template("users/form.html", **context)


@app.route("/users/<int:id>/disable", methods=["POST"])
@user_admin_required
def toggle_user_status(id):
    user = get_manageable_user(id)
    if user is None:
        return redirect(url_for("users"))

    protected_ids = {g.current_user.id}
    if getattr(g, "original_user", None):
        protected_ids.add(g.original_user.id)

    if user.id in protected_ids:
        flash("You cannot disable the current session user.", "error")
        return redirect(url_for("users"))

    user.is_active = not user.is_active
    db.session.commit()
    flash(
        f"User {'enabled' if user.is_active else 'disabled'} successfully.",
        "success",
    )
    return redirect(url_for("users"))


@app.route("/profile")
def profile():
    return render_template("users/profile.html")


@app.route("/audit-logs")
@user_admin_required
def audit_logs():
    query = AuditLog.query.order_by(AuditLog.created_at.desc())
    if not is_superadmin():
        query = query.filter_by(tenant_id=g.current_user.tenant_id)
    logs = query.limit(300).all()
    return render_template("audit_logs/list.html", logs=logs)


@app.route("/accounting", methods=["GET", "POST"])
@permission_required("accounting", "view")
def accounting_integration():
    tenant_id = get_current_tenant_id()
    integration = AccountingIntegration.query.filter_by(tenant_id=tenant_id).first()
    if request.method == "POST":
        if not has_permission("accounting", "edit"):
            flash("You do not have permission to edit accounting settings.", "error")
            return redirect(url_for("accounting_integration"))
        if not integration:
            integration = AccountingIntegration(tenant_id=tenant_id)
            db.session.add(integration)
        integration.provider = request.form.get("provider", "tally")
        integration.export_format = request.form.get("export_format", "csv")
        integration.api_base_url = request.form.get("api_base_url", "").strip() or None
        integration.api_key = request.form.get("api_key", "").strip() or None
        integration.ledger_sales = request.form.get("ledger_sales", "").strip() or "Freight Sales"
        integration.ledger_receivable = request.form.get("ledger_receivable", "").strip() or "Sundry Debtors"
        integration.ledger_tax = request.form.get("ledger_tax", "").strip() or "GST Output"
        integration.ledger_cash_bank = request.form.get("ledger_cash_bank", "").strip() or "Bank"
        integration.is_active = request.form.get("is_active") == "on"
        record_audit("accounting_settings_updated", integration, summary="Accounting integration settings updated")
        db.session.commit()
        flash("Accounting integration settings saved.", "success")
        return redirect(url_for("accounting_integration"))

    exports = AccountingExport.query.filter_by(tenant_id=tenant_id).order_by(AccountingExport.created_at.desc()).limit(20).all()
    return render_template("accounting/index.html", integration=integration, exports=exports)


@app.route("/accounting/export")
@permission_required("accounting", "export")
def accounting_export():
    tenant_id = get_current_tenant_id()
    integration = AccountingIntegration.query.filter_by(tenant_id=tenant_id).first()
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"), date.today())
    export_type = request.args.get("type", "sales")
    bills_query = TransportBill.query.filter_by(tenant_id=tenant_id)
    receipts_query = PaymentReceipt.query.filter_by(tenant_id=tenant_id)
    if from_date:
        bills_query = bills_query.filter(TransportBill.date >= from_date)
        receipts_query = receipts_query.filter(PaymentReceipt.receipt_date >= from_date)
    if to_date:
        bills_query = bills_query.filter(TransportBill.date <= to_date)
        receipts_query = receipts_query.filter(PaymentReceipt.receipt_date <= to_date)

    rows = []
    headers = [
        "Voucher Type",
        "Voucher Date",
        "Voucher Number",
        "Party Ledger",
        "Sales Ledger",
        "Tax Ledger",
        "Amount",
        "Narration",
    ]
    if export_type in ("sales", "all"):
        for bill in bills_query.order_by(TransportBill.date).all():
            amount = float(getattr(bill, "freight_due", None) or bill.rate or 0)
            rows.append(
                [
                    "Sales",
                    bill.date.strftime("%Y-%m-%d") if bill.date else "",
                    bill.name,
                    bill.party_information or bill.consignor_name or "Customer",
                    integration.ledger_sales if integration else "Freight Sales",
                    integration.ledger_tax if integration else "GST Output",
                    amount,
                    f"Freight invoice for bilty {bill.name}",
                ]
            )
    if export_type in ("receipts", "all"):
        for receipt in receipts_query.order_by(PaymentReceipt.receipt_date).all():
            rows.append(
                [
                    "Receipt",
                    receipt.receipt_date.strftime("%Y-%m-%d") if receipt.receipt_date else "",
                    receipt.receipt_number,
                    receipt.bilty.party_information if receipt.bilty else "Customer",
                    integration.ledger_cash_bank if integration else "Bank",
                    "",
                    float(receipt.amount_received or 0),
                    f"Payment against {receipt.bilty.name if receipt.bilty else ''}",
                ]
            )

    file_name = f"accounting_{export_type}_{date.today().strftime('%Y%m%d')}.csv"
    export = AccountingExport(
        tenant_id=tenant_id,
        integration_id=integration.id if integration else None,
        export_type=export_type,
        from_date=from_date,
        to_date=to_date,
        records_count=len(rows),
        file_name=file_name,
        created_by=g.current_user.id if g.current_user else None,
    )
    db.session.add(export)
    record_audit("accounting_export_generated", export, summary=f"Accounting export generated: {file_name}", tenant_id=tenant_id)
    db.session.commit()
    return csv_response(rows, headers, file_name)


@app.route("/gps", methods=["GET", "POST"])
@permission_required("gps", "view")
def gps_integration():
    tenant_id = get_current_tenant_id()
    if request.method == "POST":
        if not has_permission("gps", "edit"):
            flash("You do not have permission to edit GPS settings.", "error")
            return redirect(url_for("gps_integration"))
        device = GPSDevice(
            tenant_id=tenant_id,
            vehicle_id=parse_int(request.form.get("vehicle_id")),
            device_imei=request.form.get("device_imei", "").strip(),
            provider=request.form.get("provider", "generic").strip() or "generic",
            api_key=request.form.get("api_key", "").strip() or generate_access_token(),
            is_active=True,
        )
        if not device.device_imei:
            flash("Device IMEI is required.", "error")
            return redirect(url_for("gps_integration"))
        db.session.add(device)
        db.session.flush()
        record_audit("gps_device_added", device, summary=f"GPS device added: {device.device_imei}")
        db.session.commit()
        flash("GPS device added.", "success")
        return redirect(url_for("gps_integration"))

    devices = GPSDevice.query.filter_by(tenant_id=tenant_id).order_by(GPSDevice.created_at.desc()).all()
    vehicles = get_tenant_filtered_records(Vehicle, Vehicle.registration_number, tenant_id)
    recent_pings = GPSPing.query.filter_by(tenant_id=tenant_id).order_by(GPSPing.recorded_at.desc()).limit(50).all()
    return render_template("gps/index.html", devices=devices, vehicles=vehicles, recent_pings=recent_pings)


@app.route("/gps/webhook/<api_key>", methods=["POST"])
def gps_webhook(api_key):
    payload = request.get_json(silent=True) or request.form.to_dict()
    device_imei = str(payload.get("device_imei") or payload.get("imei") or "").strip()
    lat = parse_float(payload.get("lat") or payload.get("latitude"))
    lng = parse_float(payload.get("lng") or payload.get("longitude"))
    if not device_imei or not lat or not lng:
        return jsonify({"success": False, "error": "device_imei, latitude and longitude are required"}), 400

    device = GPSDevice.query.filter_by(device_imei=device_imei, api_key=api_key, is_active=True).first()
    if not device:
        return jsonify({"success": False, "error": "Invalid device or API key"}), 403

    active_trip = (
        DispatchTrip.query.filter_by(tenant_id=device.tenant_id, vehicle_id=device.vehicle_id)
        .filter(DispatchTrip.status.in_(["planned", "dispatched", "in_transit", "delayed"]))
        .order_by(DispatchTrip.created_at.desc())
        .first()
    )
    recorded_at = parse_datetime_local(payload.get("recorded_at"), datetime.utcnow())
    address = payload.get("address") or f"{lat}, {lng}"
    ping = GPSPing(
        tenant_id=device.tenant_id,
        device_id=device.id,
        vehicle_id=device.vehicle_id,
        trip_id=active_trip.id if active_trip else None,
        latitude=lat,
        longitude=lng,
        speed_kmph=parse_float(payload.get("speed_kmph") or payload.get("speed")),
        heading=parse_float(payload.get("heading")),
        address=address,
        recorded_at=recorded_at,
        raw_payload=json.dumps(payload, default=str),
    )
    db.session.add(ping)
    if active_trip:
        active_trip.status = "in_transit" if active_trip.status in ("planned", "dispatched") else active_trip.status
        active_trip.current_location = address
        active_trip.last_latitude = lat
        active_trip.last_longitude = lng
        active_trip.last_tracking_update_at = recorded_at
        if active_trip.bilty:
            active_trip.bilty.status = "in_transit"
    db.session.commit()
    return jsonify({"success": True, "trip": active_trip.trip_number if active_trip else None})


@app.route("/customer/login", methods=["GET", "POST"])
def customer_login():
    if request.method == "POST":
        if login_limited("customer_login_failures"):
            flash("Too many failed login attempts. Please try again later.", "error")
            return render_template("customer/login.html")
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        account = CustomerPortalAccount.query.filter_by(email=email, is_active=True).first()
        if not account or not account.check_password(password):
            record_login_failure("customer_login_failures")
            flash("Invalid customer email or password.", "error")
            return render_template("customer/login.html")
        session.clear()
        session["customer_id"] = account.id
        session.permanent = True
        account.last_login_at = datetime.utcnow()
        db.session.commit()
        return redirect(url_for("customer_dashboard"))
    return render_template("customer/login.html")


@app.route("/customer/logout")
def customer_logout():
    session.pop("customer_id", None)
    flash("You have been logged out.", "success")
    return redirect(url_for("customer_login"))


@app.route("/customer")
def customer_dashboard():
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return redirect(url_for("customer_login"))
    bill_filters = [TransportBill.consignor_email == account.email]
    if account.vendor:
        bill_filters.extend(
            [
                TransportBill.party_information == account.vendor.vendor_name,
                TransportBill.consignor_name == account.vendor.vendor_name,
            ]
        )
    bills = (
        TransportBill.query.filter_by(tenant_id=account.tenant_id)
        .filter(or_(*bill_filters))
        .order_by(TransportBill.created_at.desc())
        .all()
    )
    trips = (
        DispatchTrip.query.filter_by(tenant_id=account.tenant_id)
        .filter(DispatchTrip.bilty_id.in_([bill.id for bill in bills] or [0]))
        .order_by(DispatchTrip.created_at.desc())
        .all()
    )
    return render_template("customer/dashboard.html", account=account, bills=bills, trips=trips)


@app.route("/customer-accounts")
@user_admin_required
def customer_accounts():
    accounts_query = CustomerPortalAccount.query.order_by(CustomerPortalAccount.created_at.desc())
    vendors_query = Vendor.query.order_by(Vendor.vendor_name)
    if not is_superadmin():
        accounts_query = accounts_query.filter_by(tenant_id=g.current_user.tenant_id)
        vendors_query = vendors_query.filter_by(tenant_id=g.current_user.tenant_id)
    return render_template(
        "customer/accounts.html",
        accounts=accounts_query.all(),
        vendors=vendors_query.all(),
    )


@app.route("/customer-accounts/create", methods=["POST"])
@user_admin_required
def create_customer_account():
    tenant_id = get_current_tenant_id()
    vendor = get_related_record(Vendor, parse_int(request.form.get("vendor_id")), tenant_id, "Vendor")
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")
    full_name = request.form.get("full_name", "").strip() or (vendor.vendor_name if vendor else email)
    if not email or not password:
        flash("Email and password are required.", "error")
        return redirect(url_for("customer_accounts"))
    existing = CustomerPortalAccount.query.filter_by(tenant_id=tenant_id, email=email).first()
    if existing:
        flash("A customer account already exists for that email.", "error")
        return redirect(url_for("customer_accounts"))
    account = CustomerPortalAccount(
        tenant_id=tenant_id,
        vendor_id=vendor.id if vendor else None,
        email=email,
        full_name=full_name,
    )
    account.set_password(password)
    db.session.add(account)
    db.session.flush()
    record_audit("customer_account_created", account, summary=f"Customer portal account created for {email}")
    db.session.commit()
    flash("Customer portal account created.", "success")
    return redirect(url_for("customer_accounts"))


@app.route("/profile/request-amendment", methods=["GET", "POST"])
def request_amendment():
    import json
    
    # Define field classifications with labels
    FIELD_CONFIG = {
        'full_name': {'label': 'Full Name', 'type': 'basic'},
        'email': {'label': 'Email', 'type': 'basic'},
        'username': {'label': 'Username', 'type': 'major'},
        'role': {'label': 'Role', 'type': 'major'},
    }
    
    BASIC_FIELDS = [k for k, v in FIELD_CONFIG.items() if v['type'] == 'basic']
    MAJOR_FIELDS = [k for k, v in FIELD_CONFIG.items() if v['type'] == 'major']
    
    if request.method == "POST":
        field_changes = {}
        change_type = "basic"
        
        # Check which fields are being changed
        for field in BASIC_FIELDS + MAJOR_FIELDS:
            old_value = getattr(g.current_user, field, None)
            new_value = request.form.get(field, "").strip()
            
            if new_value and new_value != old_value:
                field_changes[field] = {"old": old_value, "new": new_value}
        
        if not field_changes:
            flash("No changes detected.", "error")
            return redirect(url_for("request_amendment"))
        
        # Determine change type based on fields
        for field in field_changes.keys():
            if field in MAJOR_FIELDS:
                change_type = "major"
                break
        
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Please provide a reason for the amendment request.", "error")
            return redirect(url_for("request_amendment"))
        
        # Create amendment request
        amendment = UserAmendmentRequest(
            user_id=g.current_user.id,
            requested_by=g.current_user.id,
            tenant_id=g.current_user.tenant_id,
            field_changes=json.dumps(field_changes),
            change_type=change_type,
            reason=reason,
            status="pending"
        )
        
        # Auto-approve basic changes
        if change_type == "basic":
            # Apply changes immediately
            for field, values in field_changes.items():
                setattr(g.current_user, field, values["new"])
            
            amendment.status = "auto_approved"
            amendment.approved_by = g.current_user.id
            amendment.approved_at = datetime.utcnow()
            
            db.session.add(amendment)
            db.session.commit()
            
            flash("Your basic profile changes have been auto-approved and applied.", "success")
            return redirect(url_for("profile"))
        else:
            # Major changes require approval
            db.session.add(amendment)
            db.session.commit()
            
            flash("Your amendment request has been submitted for superadmin approval.", "success")
            return redirect(url_for("profile"))
    
    return render_template("users/amendment_form.html", 
                          field_config=FIELD_CONFIG,
                          basic_fields=BASIC_FIELDS, 
                          major_fields=MAJOR_FIELDS,
                          current_user=g.current_user)


@app.route("/amendment-requests")
@superadmin_required
def amendment_requests():
    requests = UserAmendmentRequest.query.filter_by(status="pending").order_by(
        UserAmendmentRequest.created_at.desc()
    ).all()
    
    return render_template("users/amendment_requests.html", requests=requests)


@app.route("/amendment-requests/<int:id>/approve", methods=["POST"])
@superadmin_required
def approve_amendment(id):
    amendment = UserAmendmentRequest.query.get_or_404(id)
    
    if amendment.status != "pending":
        flash("This request has already been processed.", "error")
        return redirect(url_for("amendment_requests"))
    
    import json
    field_changes = json.loads(amendment.field_changes)
    
    # Apply changes to user
    user = amendment.user
    for field, values in field_changes.items():
        setattr(user, field, values["new"])
    
    amendment.status = "approved"
    amendment.approved_by = g.current_user.id
    amendment.approved_at = datetime.utcnow()
    
    db.session.commit()
    
    flash("Amendment request approved and changes applied.", "success")
    return redirect(url_for("amendment_requests"))


@app.route("/amendment-requests/<int:id>/reject", methods=["POST"])
@superadmin_required
def reject_amendment(id):
    amendment = UserAmendmentRequest.query.get_or_404(id)
    
    if amendment.status != "pending":
        flash("This request has already been processed.", "error")
        return redirect(url_for("amendment_requests"))
    
    rejection_reason = request.form.get("rejection_reason", "").strip()
    if not rejection_reason:
        flash("Please provide a reason for rejection.", "error")
        return redirect(url_for("amendment_requests"))
    
    amendment.status = "rejected"
    amendment.rejection_reason = rejection_reason
    
    db.session.commit()
    
    flash("Amendment request rejected.", "success")
    return redirect(url_for("amendment_requests"))


@app.route("/users/<int:id>/reset-password", methods=["GET", "POST"])
@user_admin_required
def reset_user_password(id):
    user = get_manageable_user(id)
    if user is None:
        return redirect(url_for("users"))

    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not password:
            flash("Password is required.", "error")
            return render_template("users/reset_password.html", user_obj=user)

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("users/reset_password.html", user_obj=user)

        user.set_password(password)
        db.session.commit()
        flash("Password updated successfully.", "success")
        return redirect(url_for("users"))

    return render_template("users/reset_password.html", user_obj=user)


@app.route("/delivery-types")
@permission_required("delivery_types", "view")
def delivery_types():
    types = scoped_query(DeliveryType).order_by(DeliveryType.delivery_type).all()
    return render_template(
        "delivery_types/list.html",
        types=types,
        show_tenant_column=is_superadmin(),
    )


@app.route("/delivery-types/create", methods=["GET", "POST"])
@permission_required("delivery_types", "create")
def create_delivery_type():
    if request.method == "POST":
        name = request.form.get("delivery_type", "").strip()
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template("delivery_types/form.html", **get_form_tenant_context())

        if not name:
            flash("Delivery type name is required.", "error")
            return render_template("delivery_types/form.html", **get_form_tenant_context())

        existing = DeliveryType.query.filter_by(
            tenant_id=tenant_id, delivery_type=name
        ).first()
        if existing:
            flash("Delivery type already exists for that tenant.", "error")
            return render_template("delivery_types/form.html", **get_form_tenant_context())

        db.session.add(DeliveryType(tenant_id=tenant_id, delivery_type=name))
        db.session.commit()
        flash("Delivery type created successfully.", "success")
        return redirect(url_for("delivery_types"))

    return render_template("delivery_types/form.html", **get_form_tenant_context())


@app.route("/delivery-types/edit/<int:id>", methods=["GET", "POST"])
@permission_required("delivery_types", "edit")
def edit_delivery_type(id):
    dt = get_scoped_record(DeliveryType, id)
    if request.method == "POST":
        name = request.form.get("delivery_type", "").strip()
        try:
            tenant_id = resolve_target_tenant_id(dt, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(dt)
            context.update({"dt": dt, "edit": True})
            return render_template("delivery_types/form.html", **context)

        if not name:
            flash("Delivery type name is required.", "error")
            context = get_form_tenant_context(dt)
            context.update({"dt": dt, "edit": True})
            return render_template("delivery_types/form.html", **context)

        existing = DeliveryType.query.filter_by(
            tenant_id=tenant_id, delivery_type=name
        ).first()
        if existing and existing.id != dt.id:
            flash("Delivery type already exists for that tenant.", "error")
            context = get_form_tenant_context(dt)
            context.update({"dt": dt, "edit": True})
            return render_template("delivery_types/form.html", **context)

        dt.delivery_type = name
        dt.tenant_id = tenant_id
        db.session.commit()
        flash("Delivery type updated successfully.", "success")
        return redirect(url_for("delivery_types"))

    context = get_form_tenant_context(dt)
    context.update({"dt": dt, "edit": True})
    return render_template("delivery_types/form.html", **context)


@app.route("/delivery-types/delete/<int:id>", methods=["POST"])
@permission_required("delivery_types", "delete")
def delete_delivery_type(id):
    dt = get_scoped_record(DeliveryType, id)
    try:
        db.session.delete(dt)
        db.session.commit()
        flash("Delivery type deleted successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Cannot delete: delivery type is in use.", "error")
    return redirect(url_for("delivery_types"))


@app.route("/locations")
@permission_required("locations", "view")
def locations():
    location_list = scoped_query(Location).order_by(Location.location).all()
    return render_template(
        "locations/list.html",
        locations=location_list,
        show_tenant_column=is_superadmin(),
    )


@app.route("/locations/create", methods=["GET", "POST"])
@permission_required("locations", "create")
def create_location():
    if request.method == "POST":
        name = request.form.get("location", "").strip()
        rate = parse_int(request.form.get("rate")) or 0
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template("locations/form.html", **get_form_tenant_context())

        if not name:
            flash("Location name is required.", "error")
            return render_template("locations/form.html", **get_form_tenant_context())

        existing = Location.query.filter_by(tenant_id=tenant_id, location=name).first()
        if existing:
            flash("Location already exists for that tenant.", "error")
            return render_template("locations/form.html", **get_form_tenant_context())

        db.session.add(Location(tenant_id=tenant_id, location=name, rate=rate))
        db.session.commit()
        flash("Location created successfully.", "success")
        return redirect(url_for("locations"))

    return render_template("locations/form.html", **get_form_tenant_context())


@app.route("/locations/edit/<int:id>", methods=["GET", "POST"])
@permission_required("locations", "edit")
def edit_location(id):
    location = get_scoped_record(Location, id)
    if request.method == "POST":
        name = request.form.get("location", "").strip()
        rate = parse_int(request.form.get("rate")) or 0
        try:
            tenant_id = resolve_target_tenant_id(location, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(location)
            context.update({"location": location, "edit": True})
            return render_template("locations/form.html", **context)

        if not name:
            flash("Location name is required.", "error")
            context = get_form_tenant_context(location)
            context.update({"location": location, "edit": True})
            return render_template("locations/form.html", **context)

        existing = Location.query.filter_by(tenant_id=tenant_id, location=name).first()
        if existing and existing.id != location.id:
            flash("Location already exists for that tenant.", "error")
            context = get_form_tenant_context(location)
            context.update({"location": location, "edit": True})
            return render_template("locations/form.html", **context)

        location.location = name
        location.rate = rate
        location.tenant_id = tenant_id
        db.session.commit()
        flash("Location updated successfully.", "success")
        return redirect(url_for("locations"))

    context = get_form_tenant_context(location)
    context.update({"location": location, "edit": True})
    return render_template("locations/form.html", **context)


@app.route("/locations/delete/<int:id>", methods=["POST"])
@permission_required("locations", "delete")
def delete_location(id):
    location = get_scoped_record(Location, id)
    try:
        db.session.delete(location)
        db.session.commit()
        flash("Location deleted successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Cannot delete: location is in use.", "error")
    return redirect(url_for("locations"))


@app.route("/api/location-rate/<int:id>")
@permission_required("transport_bills", "view")
def get_location_rate(id):
    location = scoped_query(Location).filter_by(id=id).first()
    if location:
        return jsonify({"rate": location.rate, "location": location.location})
    return jsonify({"rate": 0}), 404


@app.route("/ratelists")
@permission_required("ratelists", "view")
def ratelists():
    ratelist_list = scoped_query(RateList).order_by(RateList.effective_date.desc(), RateList.name).all()
    return render_template(
        "ratelists/list.html",
        ratelists=ratelist_list,
        show_tenant_column=is_superadmin(),
        today=date.today(),
    )


@app.route("/ratelists/create", methods=["GET", "POST"])
@permission_required("ratelists", "create")
def create_ratelist():
    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template("ratelists/form.html", **get_form_tenant_context())

        name = request.form.get("name", "").strip()
        if not name:
            flash("Rate list name is required.", "error")
            return render_template("ratelists/form.html", **get_form_tenant_context())

        party_id = parse_int(request.form.get("party_id")) or None
        origin = request.form.get("origin", "").strip() or None
        destination = request.form.get("destination", "").strip() or None
        vehicle_type = request.form.get("vehicle_type", "").strip() or None
        effective_date = parse_date(request.form.get("effective_date")) or date.today()

        # Vehicle-specific rates
        rate_14ft = parse_float(request.form.get("rate_14ft")) or None
        rate_17ft = parse_float(request.form.get("rate_17ft")) or None
        rate_t5_1109 = parse_float(request.form.get("rate_t5_1109")) or None
        rate_19ft = parse_float(request.form.get("rate_19ft")) or None
        rate_22ft = parse_float(request.form.get("rate_22ft")) or None
        rate_32ft = parse_float(request.form.get("rate_32ft")) or None

        # Other charges
        basic_freight = parse_float(request.form.get("basic_freight")) or None
        loading_charge = parse_float(request.form.get("loading_charge")) or None
        unloading_charge = parse_float(request.form.get("unloading_charge")) or None
        door_pickup_charge = parse_float(request.form.get("door_pickup_charge")) or None
        door_delivery_charge = parse_float(request.form.get("door_delivery_charge")) or None
        hamali_charge = parse_float(request.form.get("hamali_charge")) or None
        detention_charge = parse_float(request.form.get("detention_charge")) or None
        waiting_charge = parse_float(request.form.get("waiting_charge")) or None
        halting_charge = parse_float(request.form.get("halting_charge")) or None
        toll_charge = parse_float(request.form.get("toll_charge")) or None
        border_charge = parse_float(request.form.get("border_charge")) or None
        fuel_surcharge = parse_float(request.form.get("fuel_surcharge")) or None
        packing_charge = parse_float(request.form.get("packing_charge")) or None
        weighment_charge = parse_float(request.form.get("weighment_charge")) or None
        permit_charge = parse_float(request.form.get("permit_charge")) or None
        driver_allowance = parse_float(request.form.get("driver_allowance")) or None
        insurance_charge = parse_float(request.form.get("insurance_charge")) or None
        other_charges = parse_float(request.form.get("other_charges")) or None

        # GST
        igst_rate = parse_float(request.form.get("igst_rate")) or None
        cgst_rate = parse_float(request.form.get("cgst_rate")) or None
        sgst_rate = parse_float(request.form.get("sgst_rate")) or None

        ratelist = RateList(
            tenant_id=tenant_id,
            party_id=party_id,
            name=name,
            description=request.form.get("description", "").strip() or None,
            origin=origin,
            destination=destination,
            vehicle_type=vehicle_type,
            rate_14ft=rate_14ft,
            rate_17ft=rate_17ft,
            rate_t5_1109=rate_t5_1109,
            rate_19ft=rate_19ft,
            rate_22ft=rate_22ft,
            rate_32ft=rate_32ft,
            basic_freight=basic_freight,
            loading_charge=loading_charge,
            unloading_charge=unloading_charge,
            door_pickup_charge=door_pickup_charge,
            door_delivery_charge=door_delivery_charge,
            hamali_charge=hamali_charge,
            detention_charge=detention_charge,
            waiting_charge=waiting_charge,
            halting_charge=halting_charge,
            toll_charge=toll_charge,
            border_charge=border_charge,
            fuel_surcharge=fuel_surcharge,
            packing_charge=packing_charge,
            weighment_charge=weighment_charge,
            permit_charge=permit_charge,
            driver_allowance=driver_allowance,
            insurance_charge=insurance_charge,
            other_charges=other_charges,
            igst_rate=igst_rate,
            cgst_rate=cgst_rate,
            sgst_rate=sgst_rate,
            effective_date=effective_date,
            is_active=True,
        )
        db.session.add(ratelist)
        db.session.commit()
        flash("Rate list created successfully.", "success")
        return redirect(url_for("ratelists"))

    context = get_form_tenant_context()
    context["vendors"] = get_tenant_filtered_records(Vendor, Vendor.vendor_name, get_default_selected_tenant_id())
    context["today"] = date.today()
    return render_template("ratelists/form.html", **context)


@app.route("/ratelists/edit/<int:id>", methods=["GET", "POST"])
@permission_required("ratelists", "edit")
def edit_ratelist(id):
    ratelist = get_scoped_record(RateList, id)
    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(ratelist, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(ratelist)
            context.update({"ratelist": ratelist, "edit": True})
            return render_template("ratelists/form.html", **context)

        ratelist.name = request.form.get("name", "").strip()
        ratelist.party_id = parse_int(request.form.get("party_id")) or None
        ratelist.origin = request.form.get("origin", "").strip() or None
        ratelist.destination = request.form.get("destination", "").strip() or None
        ratelist.vehicle_type = request.form.get("vehicle_type", "").strip() or None
        ratelist.effective_date = parse_date(request.form.get("effective_date")) or date.today()
        ratelist.description = request.form.get("description", "").strip() or None

        # Vehicle-specific rates
        ratelist.rate_14ft = parse_float(request.form.get("rate_14ft")) or None
        ratelist.rate_17ft = parse_float(request.form.get("rate_17ft")) or None
        ratelist.rate_t5_1109 = parse_float(request.form.get("rate_t5_1109")) or None
        ratelist.rate_19ft = parse_float(request.form.get("rate_19ft")) or None
        ratelist.rate_22ft = parse_float(request.form.get("rate_22ft")) or None
        ratelist.rate_32ft = parse_float(request.form.get("rate_32ft")) or None

        # Other charges
        ratelist.basic_freight = parse_float(request.form.get("basic_freight")) or None
        ratelist.loading_charge = parse_float(request.form.get("loading_charge")) or None
        ratelist.unloading_charge = parse_float(request.form.get("unloading_charge")) or None
        ratelist.door_pickup_charge = parse_float(request.form.get("door_pickup_charge")) or None
        ratelist.door_delivery_charge = parse_float(request.form.get("door_delivery_charge")) or None
        ratelist.hamali_charge = parse_float(request.form.get("hamali_charge")) or None
        ratelist.detention_charge = parse_float(request.form.get("detention_charge")) or None
        ratelist.waiting_charge = parse_float(request.form.get("waiting_charge")) or None
        ratelist.halting_charge = parse_float(request.form.get("halting_charge")) or None
        ratelist.toll_charge = parse_float(request.form.get("toll_charge")) or None
        ratelist.border_charge = parse_float(request.form.get("border_charge")) or None
        ratelist.fuel_surcharge = parse_float(request.form.get("fuel_surcharge")) or None
        ratelist.packing_charge = parse_float(request.form.get("packing_charge")) or None
        ratelist.weighment_charge = parse_float(request.form.get("weighment_charge")) or None
        ratelist.permit_charge = parse_float(request.form.get("permit_charge")) or None
        ratelist.driver_allowance = parse_float(request.form.get("driver_allowance")) or None
        ratelist.insurance_charge = parse_float(request.form.get("insurance_charge")) or None
        ratelist.other_charges = parse_float(request.form.get("other_charges")) or None

        # GST
        ratelist.igst_rate = parse_float(request.form.get("igst_rate")) or None
        ratelist.cgst_rate = parse_float(request.form.get("cgst_rate")) or None
        ratelist.sgst_rate = parse_float(request.form.get("sgst_rate")) or None

        ratelist.tenant_id = tenant_id
        db.session.commit()
        flash("Rate list updated successfully.", "success")
        return redirect(url_for("ratelists"))

    context = get_form_tenant_context(ratelist)
    context.update({"ratelist": ratelist, "edit": True})
    context["vendors"] = get_tenant_filtered_records(Vendor, Vendor.vendor_name, get_default_selected_tenant_id())
    context["today"] = date.today()
    return render_template("ratelists/form.html", **context)


@app.route("/ratelists/delete/<int:id>", methods=["POST"])
@permission_required("ratelists", "delete")
def delete_ratelist(id):
    ratelist = get_scoped_record(RateList, id)
    try:
        db.session.delete(ratelist)
        db.session.commit()
        flash("Rate list deleted successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Cannot delete: rate list is in use.", "error")
    return redirect(url_for("ratelists"))


@app.route("/ratelists/export")
@permission_required("ratelists", "view")
def export_ratelists():
    """Export rate lists to Excel with all fields matching the form"""
    ratelists = scoped_query(RateList).all()
    
    # Headers matching the rate list form fields
    headers = [
        'NAME', 'PARTY', 'ORIGIN', 'DESTINATION', 'EFFECTIVE DATE', 'DESCRIPTION',
        'RATE 14FT', 'RATE 17FT', 'RATE T5-1109', 'RATE 19FT', 'RATE 22FT', 'RATE 32FT',
        'BASIC FREIGHT', 'LOADING CHARGE', 'UNLOADING CHARGE', 'DOOR PICKUP', 'DOOR DELIVERY',
        'HAMALI CHARGE', 'DETENTION CHARGE', 'WAITING CHARGE', 'TOLL CHARGE', 'FUEL SURCHARGE',
        'DRIVER ALLOWANCE', 'OTHER CHARGES', 'IGST RATE', 'CGST RATE', 'SGST RATE'
    ]
    
    data = []
    for ratelist in ratelists:
        data.append([
            ratelist.name or '',
            ratelist.party.vendor_name if ratelist.party else '',
            ratelist.origin or '',
            ratelist.destination or '',
            ratelist.effective_date.strftime('%Y-%m-%d') if ratelist.effective_date else '',
            ratelist.description or '',
            float(ratelist.rate_14ft) if ratelist.rate_14ft else '',
            float(ratelist.rate_17ft) if ratelist.rate_17ft else '',
            float(ratelist.rate_t5_1109) if ratelist.rate_t5_1109 else '',
            float(ratelist.rate_19ft) if ratelist.rate_19ft else '',
            float(ratelist.rate_22ft) if ratelist.rate_22ft else '',
            float(ratelist.rate_32ft) if ratelist.rate_32ft else '',
            float(ratelist.basic_freight) if ratelist.basic_freight else '',
            float(ratelist.loading_charge) if ratelist.loading_charge else '',
            float(ratelist.unloading_charge) if ratelist.unloading_charge else '',
            float(ratelist.door_pickup_charge) if ratelist.door_pickup_charge else '',
            float(ratelist.door_delivery_charge) if ratelist.door_delivery_charge else '',
            float(ratelist.hamali_charge) if ratelist.hamali_charge else '',
            float(ratelist.detention_charge) if ratelist.detention_charge else '',
            float(ratelist.waiting_charge) if ratelist.waiting_charge else '',
            float(ratelist.toll_charge) if ratelist.toll_charge else '',
            float(ratelist.fuel_surcharge) if ratelist.fuel_surcharge else '',
            float(ratelist.driver_allowance) if ratelist.driver_allowance else '',
            float(ratelist.other_charges) if ratelist.other_charges else '',
            float(ratelist.igst_rate) if ratelist.igst_rate else '',
            float(ratelist.cgst_rate) if ratelist.cgst_rate else '',
            float(ratelist.sgst_rate) if ratelist.sgst_rate else ''
        ])
    
    return export_to_excel(data, headers, 'rate_lists.xlsx')


@app.route("/ratelists/import", methods=["POST"])
@permission_required("ratelists", "create")
def import_ratelists():
    """Import rate lists from Excel matching the form format"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('ratelists'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('ratelists'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('ratelists'))
    
    def parse_float_safe(value):
        if not value:
            return None
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return None
    
    def parse_date_safe(value):
        if not value:
            return None
        try:
            if isinstance(value, str):
                value = value.strip()
                # Try different date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except:
                        continue
            return None
        except:
            return None
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                name = row.get('NAME', '').strip() or row.get('Name', '').strip()
                if not name:
                    errors.append(f'Row {idx}: NAME is required')
                    continue
                
                party_name = row.get('PARTY', '').strip() or row.get('Party', '').strip()
                
                # Find vendor by name
                party_id = None
                if party_name:
                    vendor = Vendor.query.filter_by(tenant_id=tenant_id, vendor_name=party_name).first()
                    if vendor:
                        party_id = vendor.id
                
                effective_date = parse_date_safe(row.get('EFFECTIVE DATE', '').strip() or row.get('EFFECTIVE_DATE', '').strip())
                
                ratelist = RateList(
                    tenant_id=tenant_id,
                    name=name,
                    description=row.get('DESCRIPTION', '').strip() or row.get('Description', '').strip() or None,
                    origin=row.get('ORIGIN', '').strip() or row.get('Origin', '').strip() or None,
                    destination=row.get('DESTINATION', '').strip() or row.get('Destination', '').strip() or None,
                    party_id=party_id,
                    rate_14ft=parse_float_safe(row.get('RATE 14FT', '').strip() or row.get('RATE_14FT', '').strip()),
                    rate_17ft=parse_float_safe(row.get('RATE 17FT', '').strip() or row.get('RATE_17FT', '').strip()),
                    rate_t5_1109=parse_float_safe(row.get('RATE T5-1109', '').strip() or row.get('RATE_T5_1109', '').strip()),
                    rate_19ft=parse_float_safe(row.get('RATE 19FT', '').strip() or row.get('RATE_19FT', '').strip()),
                    rate_22ft=parse_float_safe(row.get('RATE 22FT', '').strip() or row.get('RATE_22FT', '').strip()),
                    rate_32ft=parse_float_safe(row.get('RATE 32FT', '').strip() or row.get('RATE_32FT', '').strip()),
                    basic_freight=parse_float_safe(row.get('BASIC FREIGHT', '').strip() or row.get('BASIC_FREIGHT', '').strip()),
                    loading_charge=parse_float_safe(row.get('LOADING CHARGE', '').strip() or row.get('LOADING_CHARGE', '').strip()),
                    unloading_charge=parse_float_safe(row.get('UNLOADING CHARGE', '').strip() or row.get('UNLOADING_CHARGE', '').strip()),
                    door_pickup_charge=parse_float_safe(row.get('DOOR PICKUP', '').strip() or row.get('DOOR_PICKUP', '').strip()),
                    door_delivery_charge=parse_float_safe(row.get('DOOR DELIVERY', '').strip() or row.get('DOOR_DELIVERY', '').strip()),
                    hamali_charge=parse_float_safe(row.get('HAMALI CHARGE', '').strip() or row.get('HAMALI_CHARGE', '').strip()),
                    detention_charge=parse_float_safe(row.get('DETENTION CHARGE', '').strip() or row.get('DETENTION_CHARGE', '').strip()),
                    waiting_charge=parse_float_safe(row.get('WAITING CHARGE', '').strip() or row.get('WAITING_CHARGE', '').strip()),
                    toll_charge=parse_float_safe(row.get('TOLL CHARGE', '').strip() or row.get('TOLL_CHARGE', '').strip()),
                    fuel_surcharge=parse_float_safe(row.get('FUEL SURCHARGE', '').strip() or row.get('FUEL_SURCHARGE', '').strip()),
                    driver_allowance=parse_float_safe(row.get('DRIVER ALLOWANCE', '').strip() or row.get('DRIVER_ALLOWANCE', '').strip()),
                    other_charges=parse_float_safe(row.get('OTHER CHARGES', '').strip() or row.get('OTHER_CHARGES', '').strip()),
                    igst_rate=parse_float_safe(row.get('IGST RATE', '').strip() or row.get('IGST_RATE', '').strip()),
                    cgst_rate=parse_float_safe(row.get('CGST RATE', '').strip() or row.get('CGST_RATE', '').strip()),
                    sgst_rate=parse_float_safe(row.get('SGST RATE', '').strip() or row.get('SGST_RATE', '').strip()),
                    effective_date=effective_date or date.today(),
                    is_active=True,
                )
                db.session.add(ratelist)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        
        if errors:
            flash(f'Imported {imported_count} rate lists. Errors: {"; ".join(errors[:5])}', 'warning')
        else:
            flash(f'Successfully imported {imported_count} rate lists.', 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error during import: {str(e)}', 'error')
    
    return redirect(url_for('ratelists'))


@app.route("/vehicles")
@permission_required("vehicles", "view")
def vehicles():
    vehicle_list = scoped_query(Vehicle).order_by(Vehicle.registration_number).all()
    return render_template(
        "vehicles/list.html",
        vehicles=vehicle_list,
        show_tenant_column=is_superadmin(),
    )


@app.route("/vehicles/create", methods=["GET", "POST"])
@permission_required("vehicles", "create")
def create_vehicle():
    if request.method == "POST":
        reg_num = request.form.get("registration_number", "").strip()
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template("vehicles/form.html", **get_form_tenant_context())

        if not reg_num:
            flash("Registration number is required.", "error")
            return render_template("vehicles/form.html", **get_form_tenant_context())

        existing = Vehicle.query.filter_by(
            tenant_id=tenant_id, registration_number=reg_num
        ).first()
        if existing:
            flash("Vehicle with this registration number already exists for that tenant.", "error")
            return render_template("vehicles/form.html", **get_form_tenant_context())

        vehicle = Vehicle(
            tenant_id=tenant_id,
            registration_number=reg_num,
            vehicle_type=request.form.get("vehicle_type", "").strip() or None,
            make=request.form.get("make", "").strip() or None,
            model=request.form.get("model", "").strip() or None,
            year=parse_int(request.form.get("year")),
            color=request.form.get("color", "").strip() or None,
            fuel_type=request.form.get("fuel_type", "").strip() or None,
            engine_number=request.form.get("engine_number", "").strip() or None,
            chassis_number=request.form.get("chassis_number", "").strip() or None,
            seating_capacity=parse_int(request.form.get("seating_capacity")),
            load_capacity_kg=parse_int(request.form.get("load_capacity_kg")),
            owner_name=request.form.get("owner_name", "").strip() or None,
            owner_contact=request.form.get("owner_contact", "").strip() or None,
            purchase_date=parse_date(request.form.get("purchase_date")),
            # Expiry dates
            insurance_expiry=parse_date(request.form.get("insurance_expiry")),
            fitness_expiry=parse_date(request.form.get("fitness_expiry")),
            permit_1_year_expiry=parse_date(request.form.get("permit_1_year_expiry")),
            permit_5_year_expiry=parse_date(request.form.get("permit_5_year_expiry")),
            road_tax_expiry=parse_date(request.form.get("road_tax_expiry")),
            puc_expiry=parse_date(request.form.get("puc_expiry")),
            # Certificate attachments (paths)
            insurance_attachment_path=request.form.get("insurance_attachment_path", "").strip() or None,
            fitness_certificate_path=request.form.get("fitness_certificate_path", "").strip() or None,
            permit_1_year_attachment_path=request.form.get("permit_1_year_attachment_path", "").strip() or None,
            permit_5_year_attachment_path=request.form.get("permit_5_year_attachment_path", "").strip() or None,
            road_tax_attachment_path=request.form.get("road_tax_attachment_path", "").strip() or None,
            puc_attachment_path=request.form.get("puc_attachment_path", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            status=request.form.get("status", "Active"),
        )
        db.session.add(vehicle)
        db.session.commit()
        
        # Update subscription usage
        update_subscription_usage(tenant_id)
        
        flash("Vehicle created successfully.", "success")
        return redirect(url_for("vehicles"))

    return render_template("vehicles/form.html", **get_form_tenant_context())


@app.route("/vehicles/edit/<int:id>", methods=["GET", "POST"])
@permission_required("vehicles", "edit")
def edit_vehicle(id):
    vehicle = get_scoped_record(Vehicle, id)

    if request.method == "POST":
        reg_num = request.form.get("registration_number", "").strip()
        try:
            tenant_id = resolve_target_tenant_id(vehicle, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(vehicle)
            context.update({"vehicle": vehicle, "edit": True})
            return render_template("vehicles/form.html", **context)

        if not reg_num:
            flash("Registration number is required.", "error")
            context = get_form_tenant_context(vehicle)
            context.update({"vehicle": vehicle, "edit": True})
            return render_template("vehicles/form.html", **context)

        existing = Vehicle.query.filter_by(
            tenant_id=tenant_id, registration_number=reg_num
        ).first()
        if existing and existing.id != vehicle.id:
            flash("Vehicle with this registration number already exists for that tenant.", "error")
            context = get_form_tenant_context(vehicle)
            context.update({"vehicle": vehicle, "edit": True})
            return render_template("vehicles/form.html", **context)

        vehicle.tenant_id = tenant_id
        vehicle.registration_number = reg_num
        vehicle.vehicle_type = request.form.get("vehicle_type", "").strip() or None
        vehicle.make = request.form.get("make", "").strip() or None
        vehicle.model = request.form.get("model", "").strip() or None
        vehicle.year = parse_int(request.form.get("year"))
        vehicle.color = request.form.get("color", "").strip() or None
        vehicle.fuel_type = request.form.get("fuel_type", "").strip() or None
        vehicle.engine_number = request.form.get("engine_number", "").strip() or None
        vehicle.chassis_number = request.form.get("chassis_number", "").strip() or None
        vehicle.seating_capacity = parse_int(request.form.get("seating_capacity"))
        vehicle.load_capacity_kg = parse_int(request.form.get("load_capacity_kg"))
        vehicle.owner_name = request.form.get("owner_name", "").strip() or None
        vehicle.owner_contact = request.form.get("owner_contact", "").strip() or None
        vehicle.purchase_date = parse_date(request.form.get("purchase_date"))
        # Expiry dates
        vehicle.insurance_expiry = parse_date(request.form.get("insurance_expiry"))
        vehicle.fitness_expiry = parse_date(request.form.get("fitness_expiry"))
        vehicle.permit_1_year_expiry = parse_date(request.form.get("permit_1_year_expiry"))
        vehicle.permit_5_year_expiry = parse_date(request.form.get("permit_5_year_expiry"))
        vehicle.road_tax_expiry = parse_date(request.form.get("road_tax_expiry"))
        vehicle.puc_expiry = parse_date(request.form.get("puc_expiry"))
        # Certificate attachments (paths)
        vehicle.insurance_attachment_path = request.form.get("insurance_attachment_path", "").strip() or None
        vehicle.fitness_certificate_path = request.form.get("fitness_certificate_path", "").strip() or None
        vehicle.permit_1_year_attachment_path = request.form.get("permit_1_year_attachment_path", "").strip() or None
        vehicle.permit_5_year_attachment_path = request.form.get("permit_5_year_attachment_path", "").strip() or None
        vehicle.road_tax_attachment_path = request.form.get("road_tax_attachment_path", "").strip() or None
        vehicle.puc_attachment_path = request.form.get("puc_attachment_path", "").strip() or None
        vehicle.notes = request.form.get("notes", "").strip() or None
        vehicle.status = request.form.get("status", "Active")
        db.session.commit()
        
        # Update subscription usage
        update_subscription_usage(tenant_id)
        
        flash("Vehicle updated successfully.", "success")
        return redirect(url_for("vehicles"))

    context = get_form_tenant_context(vehicle)
    context.update({"vehicle": vehicle, "edit": True})
    return render_template("vehicles/form.html", **context)


@app.route("/vehicles/delete/<int:id>", methods=["POST"])
@permission_required("vehicles", "delete")
def delete_vehicle(id):
    vehicle = get_scoped_record(Vehicle, id)
    try:
        db.session.delete(vehicle)
        db.session.commit()
        flash("Vehicle deleted successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Cannot delete: vehicle has associated records.", "error")
    return redirect(url_for("vehicles"))


@app.route("/vehicles/<int:id>/dashboard")
@permission_required("vehicles", "view")
def vehicle_dashboard(id):
    """Comprehensive vehicle dashboard showing all information, metrics, and status."""
    from datetime import date, timedelta
    
    vehicle = get_scoped_record(Vehicle, id)
    
    # Get assigned driver
    driver = None
    if vehicle.driver_id:
        driver = Driver.query.get(vehicle.driver_id)
    
    # Trip statistics
    trip_count = DispatchTrip.query.filter_by(vehicle_id=id).count()
    recent_trips = DispatchTrip.query.filter_by(vehicle_id=id).order_by(
        DispatchTrip.trip_date.desc()
    ).limit(5).all()
    
    # Revenue calculation
    total_revenue = db.session.query(func.sum(TransportBill.rate)).join(
        DispatchTrip, DispatchTrip.bilty_id == TransportBill.id
    ).filter(DispatchTrip.vehicle_id == id).scalar() or 0
    
    # Expenses
    total_expenses = db.session.query(func.sum(Expense.amount)).filter(
        Expense.vehicle_id == id
    ).scalar() or 0
    
    # Calculate document expiry status
    def get_expiry_status(expiry_date):
        if not expiry_date:
            return {'status': 'unknown', 'days': None, 'color': 'secondary'}
        today = date.today()
        days = (expiry_date - today).days
        if days < 0:
            return {'status': 'expired', 'days': days, 'color': 'danger'}
        elif days <= 7:
            return {'status': 'critical', 'days': days, 'color': 'danger'}
        elif days <= 14:
            return {'status': 'warning', 'days': days, 'color': 'warning'}
        elif days <= 30:
            return {'status': 'expiring', 'days': days, 'color': 'info'}
        else:
            return {'status': 'valid', 'days': days, 'color': 'success'}
    
    documents = {
        'insurance': {
            'name': 'Insurance',
            'expiry': vehicle.insurance_expiry,
            'attachment': vehicle.insurance_attachment_path,
            **get_expiry_status(vehicle.insurance_expiry)
        },
        'fitness': {
            'name': 'Fitness Certificate',
            'expiry': vehicle.fitness_expiry,
            'attachment': vehicle.fitness_certificate_path,
            **get_expiry_status(vehicle.fitness_expiry)
        },
        'permit_1': {
            'name': 'Permit (1 Year)',
            'expiry': vehicle.permit_1_year_expiry,
            'attachment': vehicle.permit_1_year_attachment_path,
            **get_expiry_status(vehicle.permit_1_year_expiry)
        },
        'permit_5': {
            'name': 'Permit (5 Year)',
            'expiry': vehicle.permit_5_year_expiry,
            'attachment': vehicle.permit_5_year_attachment_path,
            **get_expiry_status(vehicle.permit_5_year_expiry)
        },
        'road_tax': {
            'name': 'Road Tax',
            'expiry': vehicle.road_tax_expiry,
            'attachment': vehicle.road_tax_attachment_path,
            **get_expiry_status(vehicle.road_tax_expiry)
        },
        'puc': {
            'name': 'PUC',
            'expiry': vehicle.puc_expiry,
            'attachment': vehicle.puc_attachment_path,
            **get_expiry_status(vehicle.puc_expiry)
        }
    }
    
    # Calculate vehicle age
    age_text = ""
    if vehicle.purchase_date:
        years = (date.today() - vehicle.purchase_date).days // 365
        if years < 1:
            months = (date.today() - vehicle.purchase_date).days // 30
            age_text = f"{months} months"
        else:
            age_text = f"{years} years"
    
    context = {
        'vehicle': vehicle,
        'driver': driver,
        'trip_count': trip_count,
        'recent_trips': recent_trips,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': total_revenue - total_expenses,
        'documents': documents,
        'age_text': age_text,
        'show_tenant_column': is_superadmin(),
        'now': date.today
    }
    return render_template("vehicles/dashboard.html", **context)


@app.route("/vehicles/<int:id>/trips")
@permission_required("vehicles", "view")
def vehicle_trips(id):
    """Trip history and revenue tracking for a specific vehicle."""
    from datetime import date, timedelta
    
    vehicle = get_scoped_record(Vehicle, id)
    
    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.getlist('status')
    party_search = request.args.get('party', '').strip()
    route_search = request.args.get('route', '').strip()
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    
    # Base query - join DispatchTrip with TransportBill
    query = db.session.query(
        DispatchTrip,
        TransportBill
    ).join(
        TransportBill, DispatchTrip.bilty_id == TransportBill.id
    ).filter(
        DispatchTrip.vehicle_id == id
    )
    
    # Apply filters
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(DispatchTrip.trip_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(DispatchTrip.trip_date <= to_date)
        except ValueError:
            pass
    
    if status_filter:
        query = query.filter(DispatchTrip.status.in_(status_filter))
    
    if party_search:
        query = query.filter(TransportBill.party_information.ilike(f'%{party_search}%'))
    
    if route_search:
        query = query.filter(
            or_(
                DispatchTrip.origin.ilike(f'%{route_search}%'),
                DispatchTrip.destination.ilike(f'%{route_search}%')
            )
        )
    
    if min_amount:
        try:
            query = query.filter(TransportBill.rate >= float(min_amount))
        except ValueError:
            pass
    
    if max_amount:
        try:
            query = query.filter(TransportBill.rate <= float(max_amount))
        except ValueError:
            pass
    
    # Order by date descending
    query = query.order_by(DispatchTrip.trip_date.desc())
    
    # Get all trips for the list
    trips = query.all()
    
    # Calculate summary statistics
    total_trips = len(trips)
    total_revenue = sum(float(t[1].rate or 0) for t in trips)
    avg_revenue = total_revenue / total_trips if total_trips > 0 else 0
    
    # Status counts
    status_counts = {}
    for trip, bill in trips:
        status = trip.status or 'Unknown'
        status_counts[status] = status_counts.get(status, 0) + 1
    
    context = {
        'vehicle': vehicle,
        'trips': trips,
        'total_trips': total_trips,
        'total_revenue': total_revenue,
        'avg_revenue': avg_revenue,
        'status_counts': status_counts,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'status': status_filter,
            'party': party_search,
            'route': route_search,
            'min_amount': min_amount,
            'max_amount': max_amount
        },
        'show_tenant_column': is_superadmin()
    }
    return render_template("vehicles/trips.html", **context)


@app.route("/vehicles/<int:id>/expenses")
@permission_required("vehicles", "view")
def vehicle_expenses(id):
    """Expense tracking and categorization for a specific vehicle."""
    from datetime import date, timedelta
    
    vehicle = get_scoped_record(Vehicle, id)
    
    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    category_filter = request.args.getlist('category')
    vendor_search = request.args.get('vendor', '').strip()
    min_amount = request.args.get('min_amount', '')
    max_amount = request.args.get('max_amount', '')
    
    # Base query
    query = Expense.query.filter_by(vehicle_id=id)
    
    # Apply filters
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Expense.expense_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Expense.expense_date <= to_date)
        except ValueError:
            pass
    
    if category_filter:
        query = query.filter(Expense.category.in_(category_filter))
    
    if vendor_search:
        query = query.filter(Expense.vendor_name.ilike(f'%{vendor_search}%'))
    
    if min_amount:
        try:
            query = query.filter(Expense.amount >= float(min_amount))
        except ValueError:
            pass
    
    if max_amount:
        try:
            query = query.filter(Expense.amount <= float(max_amount))
        except ValueError:
            pass
    
    # Order by date descending
    query = query.order_by(Expense.expense_date.desc())
    
    # Get all expenses
    expenses = query.all()
    
    # Calculate summary statistics
    total_expenses = sum(float(e.amount or 0) for e in expenses)
    
    # Category breakdown
    category_totals = {}
    for expense in expenses:
        cat = expense.category or 'Other'
        category_totals[cat] = category_totals.get(cat, 0) + float(expense.amount or 0)
    
    # Find highest expense category
    highest_category = max(category_totals, key=category_totals.get) if category_totals else None
    
    # Recent 30 days expenses
    thirty_days_ago = date.today() - timedelta(days=30)
    recent_expenses = sum(
        float(e.amount or 0) 
        for e in expenses 
        if e.expense_date and e.expense_date >= thirty_days_ago
    )
    
    # Monthly average (if date range spans multiple months)
    avg_per_month = 0
    if expenses:
        dates = [e.expense_date for e in expenses if e.expense_date]
        if dates:
            min_date = min(dates)
            max_date = max(dates)
            months = max(1, (max_date.year - min_date.year) * 12 + (max_date.month - min_date.month) + 1)
            avg_per_month = total_expenses / months
    
    context = {
        'vehicle': vehicle,
        'expenses': expenses,
        'total_expenses': total_expenses,
        'category_totals': category_totals,
        'highest_category': highest_category,
        'recent_30_days': recent_expenses,
        'avg_per_month': avg_per_month,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'category': category_filter,
            'vendor': vendor_search,
            'min_amount': min_amount,
            'max_amount': max_amount
        },
        'categories': ['Fuel', 'Maintenance', 'Repairs', 'Tires', 'Insurance', 'Tax', 'Accessories', 'Other'],
        'show_tenant_column': is_superadmin()
    }
    return render_template("vehicles/expenses.html", **context)


@app.route("/vehicles/<int:id>/profitability")
@permission_required("vehicles", "view")
def vehicle_profitability(id):
    """Profitability analysis and P&L statement for a specific vehicle."""
    from datetime import date, timedelta
    
    vehicle = get_scoped_record(Vehicle, id)
    
    # Get period selection
    period = request.args.get('period', 'lifetime')  # lifetime, 3m, 6m, 1y, custom
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Calculate date range
    today = date.today()
    if period == '3m':
        start_date = today - timedelta(days=90)
        end_date = today
    elif period == '6m':
        start_date = today - timedelta(days=180)
        end_date = today
    elif period == '1y':
        start_date = today - timedelta(days=365)
        end_date = today
    elif period == 'custom' and date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            start_date = None
            end_date = None
    else:
        start_date = None
        end_date = None
    
    # Base queries with date filters
    trip_query = DispatchTrip.query.filter_by(vehicle_id=id)
    expense_query = Expense.query.filter_by(vehicle_id=id)
    
    if start_date and end_date:
        trip_query = trip_query.filter(DispatchTrip.trip_date.between(start_date, end_date))
        expense_query = expense_query.filter(Expense.expense_date.between(start_date, end_date))
    
    # Get trips with revenue
    trips = trip_query.all()
    trip_ids = [t.bilty_id for t in trips if t.bilty_id]
    
    # Calculate revenue
    total_revenue = db.session.query(func.sum(TransportBill.rate)).filter(
        TransportBill.id.in_(trip_ids) if trip_ids else False
    ).scalar() or 0
    
    # Calculate expenses by category
    expenses = expense_query.all()
    total_expenses = sum(float(e.amount or 0) for e in expenses)
    
    # Category breakdown
    category_totals = {}
    for expense in expenses:
        cat = expense.category or 'Other'
        category_totals[cat] = category_totals.get(cat, 0) + float(expense.amount or 0)
    
    # Net profit/loss
    net_profit = float(total_revenue) - total_expenses
    profit_margin = (net_profit / float(total_revenue) * 100) if total_revenue > 0 else 0
    
    # Monthly breakdown for chart (last 12 months)
    monthly_data = []
    for i in range(11, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Revenue for month
        month_trips = DispatchTrip.query.filter_by(vehicle_id=id).filter(
            DispatchTrip.trip_date.between(month_start, month_end)
        ).all()
        month_trip_ids = [t.bilty_id for t in month_trips if t.bilty_id]
        month_revenue = db.session.query(func.sum(TransportBill.rate)).filter(
            TransportBill.id.in_(month_trip_ids) if month_trip_ids else False
        ).scalar() or 0
        
        # Expenses for month
        month_expenses = db.session.query(func.sum(Expense.amount)).filter(
            Expense.vehicle_id == id,
            Expense.expense_date.between(month_start, month_end)
        ).scalar() or 0
        
        month_profit = float(month_revenue) - float(month_expenses)
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': float(month_revenue),
            'expenses': float(month_expenses),
            'profit': month_profit
        })
    
    # Profitability status
    if profit_margin >= 20:
        profit_status = 'excellent'
        profit_color = 'success'
    elif profit_margin >= 10:
        profit_status = 'good'
        profit_color = 'info'
    elif profit_margin >= 0:
        profit_status = 'break_even'
        profit_color = 'warning'
    else:
        profit_status = 'loss'
        profit_color = 'danger'
    
    context = {
        'vehicle': vehicle,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'profit_margin': profit_margin,
        'profit_status': profit_status,
        'profit_color': profit_color,
        'category_totals': category_totals,
        'monthly_data': monthly_data,
        'trip_count': len(trips),
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'show_tenant_column': is_superadmin()
    }
    return render_template("vehicles/profitability.html", **context)


@app.route("/vehicles/<int:id>/services")
@permission_required("vehicles", "view")
def vehicle_services(id):
    """Service and maintenance log for a specific vehicle."""
    from datetime import date, timedelta
    import json
    
    vehicle = get_scoped_record(Vehicle, id)
    
    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    service_type_filter = request.args.getlist('service_type')
    garage_search = request.args.get('garage', '').strip()
    
    # Base query
    query = VehicleServiceLog.query.filter_by(vehicle_id=id)
    
    # Apply filters
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(VehicleServiceLog.service_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(VehicleServiceLog.service_date <= to_date)
        except ValueError:
            pass
    
    if service_type_filter:
        query = query.filter(VehicleServiceLog.service_type.in_(service_type_filter))
    
    if garage_search:
        query = query.filter(VehicleServiceLog.garage_name.ilike(f'%{garage_search}%'))
    
    # Order by date descending
    query = query.order_by(VehicleServiceLog.service_date.desc())
    
    # Get all services
    services = query.all()
    
    # Calculate summary statistics
    total_services = len(services)
    total_cost = sum(float(s.total_cost or 0) for s in services)
    avg_cost = total_cost / total_services if total_services > 0 else 0
    
    # Service type breakdown
    type_totals = {}
    for service in services:
        stype = service.service_type or 'Other'
        type_totals[stype] = type_totals.get(stype, 0) + float(service.total_cost or 0)
    
    # Last service info
    last_service = services[0] if services else None
    
    # Next service due
    next_service_due = None
    if last_service and last_service.next_service_date:
        days_until = (last_service.next_service_date - date.today()).days
        next_service_due = {
            'date': last_service.next_service_date,
            'days_until': days_until,
            'km': last_service.next_service_km,
            'is_overdue': days_until < 0,
            'is_soon': 0 <= days_until <= 7
        }
    
    # Services this year
    this_year_start = date.today().replace(month=1, day=1)
    services_this_year = sum(1 for s in services if s.service_date and s.service_date >= this_year_start)
    cost_this_year = sum(float(s.total_cost or 0) for s in services if s.service_date and s.service_date >= this_year_start)
    
    context = {
        'vehicle': vehicle,
        'services': services,
        'total_services': total_services,
        'total_cost': total_cost,
        'avg_cost': avg_cost,
        'type_totals': type_totals,
        'last_service': last_service,
        'next_service_due': next_service_due,
        'services_this_year': services_this_year,
        'cost_this_year': cost_this_year,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'service_type': service_type_filter,
            'garage': garage_search
        },
        'service_types': [
            'Oil Change', 'Filter Replacement', 'Tire Service', 'Brake Service',
            'Battery Service', 'AC Service', 'General Service', 'Repair', 'Inspection'
        ],
        'show_tenant_column': is_superadmin()
    }
    return render_template("vehicles/services.html", **context)


@app.route("/vehicles/<int:id>/services/create", methods=["GET", "POST"])
@permission_required("vehicles", "create")
def create_vehicle_service(id):
    """Create a new service log entry for a vehicle."""
    import json
    vehicle = get_scoped_record(Vehicle, id)
    
    if request.method == "POST":
        try:
            service_log = VehicleServiceLog(
                tenant_id=vehicle.tenant_id,
                vehicle_id=id,
                service_date=parse_date(request.form.get("service_date")),
                service_type=request.form.get("service_type", "").strip() or "Other",
                service_description=request.form.get("service_description", "").strip() or None,
                odometer_reading=int(request.form.get("odometer_reading", 0)) if request.form.get("odometer_reading") else None,
                garage_name=request.form.get("garage_name", "").strip() or None,
                garage_contact=request.form.get("garage_contact", "").strip() or None,
                garage_address=request.form.get("garage_address", "").strip() or None,
                labor_cost=float(request.form.get("labor_cost", 0)) if request.form.get("labor_cost") else 0,
                parts_cost=float(request.form.get("parts_cost", 0)) if request.form.get("parts_cost") else 0,
                total_cost=float(request.form.get("total_cost", 0)) if request.form.get("total_cost") else 0,
                invoice_number=request.form.get("invoice_number", "").strip() or None,
                next_service_date=parse_date(request.form.get("next_service_date")),
                next_service_km=int(request.form.get("next_service_km", 0)) if request.form.get("next_service_km") else None,
                notes=request.form.get("notes", "").strip() or None,
                created_by=g.current_user.id if g.current_user else None
            )
            
            # Handle parts replaced as JSON
            parts_list = []
            part_names = request.form.getlist("part_name[]")
            part_costs = request.form.getlist("part_cost[]")
            for i, name in enumerate(part_names):
                if name.strip():
                    parts_list.append({
                        "part": name.strip(),
                        "cost": float(part_costs[i]) if i < len(part_costs) and part_costs[i] else 0
                    })
            if parts_list:
                service_log.parts_replaced = json.dumps(parts_list)
            
            # Handle file upload for invoice
            if "invoice_file" in request.files:
                file = request.files["invoice_file"]
                if file and file.filename:
                    service_log.invoice_path = save_uploaded_file(file, "service_invoices")
            
            db.session.add(service_log)
            db.session.commit()
            
            flash("Service log entry created successfully.", "success")
            return redirect(url_for("vehicle_services", id=id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error creating service log: {str(e)}", "error")
    
    context = {
        'vehicle': vehicle,
        'service_types': [
            'Oil Change', 'Filter Replacement', 'Tire Service', 'Brake Service',
            'Battery Service', 'AC Service', 'General Service', 'Repair', 'Inspection'
        ],
        'today': date.today().isoformat()
    }
    return render_template("vehicles/service_form.html", **context)


@app.route("/vehicles/<int:vehicle_id>/services/<int:service_id>/edit", methods=["GET", "POST"])
@permission_required("vehicles", "edit")
def edit_vehicle_service(vehicle_id, service_id):
    """Edit a service log entry."""
    import json
    vehicle = get_scoped_record(Vehicle, vehicle_id)
    service_log = VehicleServiceLog.query.filter_by(id=service_id, vehicle_id=vehicle_id).first_or_404()
    
    if request.method == "POST":
        try:
            service_log.service_date = parse_date(request.form.get("service_date"))
            service_log.service_type = request.form.get("service_type", "").strip() or "Other"
            service_log.service_description = request.form.get("service_description", "").strip() or None
            service_log.odometer_reading = int(request.form.get("odometer_reading", 0)) if request.form.get("odometer_reading") else None
            service_log.garage_name = request.form.get("garage_name", "").strip() or None
            service_log.garage_contact = request.form.get("garage_contact", "").strip() or None
            service_log.garage_address = request.form.get("garage_address", "").strip() or None
            service_log.labor_cost = float(request.form.get("labor_cost", 0)) if request.form.get("labor_cost") else 0
            service_log.parts_cost = float(request.form.get("parts_cost", 0)) if request.form.get("parts_cost") else 0
            service_log.total_cost = float(request.form.get("total_cost", 0)) if request.form.get("total_cost") else 0
            service_log.invoice_number = request.form.get("invoice_number", "").strip() or None
            service_log.next_service_date = parse_date(request.form.get("next_service_date"))
            service_log.next_service_km = int(request.form.get("next_service_km", 0)) if request.form.get("next_service_km") else None
            service_log.notes = request.form.get("notes", "").strip() or None
            
            # Handle parts replaced as JSON
            parts_list = []
            part_names = request.form.getlist("part_name[]")
            part_costs = request.form.getlist("part_cost[]")
            for i, name in enumerate(part_names):
                if name.strip():
                    parts_list.append({
                        "part": name.strip(),
                        "cost": float(part_costs[i]) if i < len(part_costs) and part_costs[i] else 0
                    })
            service_log.parts_replaced = json.dumps(parts_list) if parts_list else None
            
            # Handle file upload for invoice
            if "invoice_file" in request.files:
                file = request.files["invoice_file"]
                if file and file.filename:
                    service_log.invoice_path = save_uploaded_file(file, "service_invoices")
            
            db.session.commit()
            
            flash("Service log entry updated successfully.", "success")
            return redirect(url_for("vehicle_services", id=vehicle_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating service log: {str(e)}", "error")
    
    # Parse parts for display in form
    parts_list = []
    if service_log.parts_replaced:
        try:
            parts_list = json.loads(service_log.parts_replaced)
        except:
            pass
    
    context = {
        'vehicle': vehicle,
        'service': service_log,
        'parts_list': parts_list,
        'service_types': [
            'Oil Change', 'Filter Replacement', 'Tire Service', 'Brake Service',
            'Battery Service', 'AC Service', 'General Service', 'Repair', 'Inspection'
        ],
        'edit': True
    }
    return render_template("vehicles/service_form.html", **context)


@app.route("/vehicles/<int:vehicle_id>/services/<int:service_id>/delete", methods=["POST"])
@permission_required("vehicles", "delete")
def delete_vehicle_service(vehicle_id, service_id):
    """Delete a service log entry."""
    service_log = VehicleServiceLog.query.filter_by(id=service_id, vehicle_id=vehicle_id).first_or_404()
    
    try:
        db.session.delete(service_log)
        db.session.commit()
        flash("Service log entry deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting service log: {str(e)}", "error")
    
    return redirect(url_for("vehicle_services", id=vehicle_id))


@app.route("/vehicles/<int:id>/fuel-log")
@permission_required("vehicles", "view")
def vehicle_fuel_log(id):
    """Display fuel log with efficiency calculations for a vehicle."""
    vehicle = Vehicle.query.get_or_404(id)
    
    # Filters
    from_date = request.args.get("from_date")
    to_date = request.args.get("to_date")
    fuel_type = request.args.get("fuel_type", "")
    fuel_station = request.args.get("fuel_station", "")
    
    # Base query
    query = VehicleFuelLog.query.filter_by(vehicle_id=id)
    
    if from_date:
        query = query.filter(VehicleFuelLog.fueling_date >= from_date)
    if to_date:
        query = query.filter(VehicleFuelLog.fueling_date <= to_date)
    if fuel_type:
        query = query.filter(VehicleFuelLog.fuel_type == fuel_type)
    if fuel_station:
        query = query.filter(VehicleFuelLog.fuel_station.ilike(f"%{fuel_station}%"))
    
    # Order by date desc
    fuel_logs = query.order_by(VehicleFuelLog.fueling_date.desc()).all()
    
    # Calculate summary statistics
    total_liters = sum(float(log.fuel_liters or 0) for log in fuel_logs)
    total_cost = sum(float(log.total_cost or 0) for log in fuel_logs)
    
    # Efficiency stats
    efficiencies = [float(log.efficiency_km_per_liter or 0) for log in fuel_logs if log.efficiency_km_per_liter]
    avg_efficiency = sum(efficiencies) / len(efficiencies) if efficiencies else 0
    best_efficiency = max(efficiencies) if efficiencies else 0
    worst_efficiency = min(efficiencies) if efficiencies else 0
    
    # Cost per km stats
    costs_per_km = [float(log.cost_per_km or 0) for log in fuel_logs if log.cost_per_km]
    avg_cost_per_km = sum(costs_per_km) / len(costs_per_km) if costs_per_km else 0
    
    # This month's stats
    today = date.today()
    month_start = today.replace(day=1)
    month_logs = [log for log in fuel_logs if log.fueling_date >= month_start]
    month_liters = sum(float(log.fuel_liters or 0) for log in month_logs)
    month_cost = sum(float(log.total_cost or 0) for log in month_logs)
    
    # Unique fuel types and stations for filters
    fuel_types = db.session.query(VehicleFuelLog.fuel_type).filter_by(vehicle_id=id).distinct().all()
    fuel_types = [ft[0] for ft in fuel_types if ft[0]]
    
    fuel_stations = db.session.query(VehicleFuelLog.fuel_station).filter_by(vehicle_id=id).distinct().all()
    fuel_stations = [fs[0] for fs in fuel_stations if fs[0]]
    
    context = {
        "vehicle": vehicle,
        "fuel_logs": fuel_logs,
        "total_liters": total_liters,
        "total_cost": total_cost,
        "avg_efficiency": avg_efficiency,
        "best_efficiency": best_efficiency,
        "worst_efficiency": worst_efficiency,
        "avg_cost_per_km": avg_cost_per_km,
        "month_liters": month_liters,
        "month_cost": month_cost,
        "fuel_types": fuel_types,
        "fuel_stations": fuel_stations,
        "filters": {
            "from_date": from_date,
            "to_date": to_date,
            "fuel_type": fuel_type,
            "fuel_station": fuel_station,
        },
    }
    return render_template("vehicles/fuel_log.html", **context)


@app.route("/vehicles/<int:id>/fuel-log/create", methods=["GET", "POST"])
@permission_required("vehicles", "create")
def create_vehicle_fuel_log(id):
    """Create a new fuel log entry."""
    vehicle = Vehicle.query.get_or_404(id)
    drivers = scoped_query(Driver).filter_by(vehicle_id=id).all()
    
    if request.method == "POST":
        fuel_log = VehicleFuelLog(
            tenant_id=get_current_tenant_id(),
            vehicle_id=id,
            fueling_date=request.form.get("fueling_date"),
            odometer_reading=int(request.form.get("odometer_reading", 0)),
            fuel_liters=float(request.form.get("fuel_liters", 0)),
            fuel_price_per_liter=float(request.form.get("fuel_price_per_liter", 0)) if request.form.get("fuel_price_per_liter") else None,
            total_cost=float(request.form.get("total_cost", 0)),
            fuel_station=request.form.get("fuel_station"),
            fuel_type=request.form.get("fuel_type"),
            driver_id=request.form.get("driver_id") or None,
            payment_method=request.form.get("payment_method"),
            receipt_number=request.form.get("receipt_number"),
            notes=request.form.get("notes"),
            created_by=session.get("user_id"),
        )
        
        # Handle receipt upload
        if "receipt_file" in request.files:
            file = request.files["receipt_file"]
            if file and file.filename:
                fuel_log.receipt_path = save_uploaded_file(file, "fuel_receipts", f"fuel_{id}")
        
        # Calculate efficiency
        fuel_log = calculate_fuel_efficiency(fuel_log)
        
        db.session.add(fuel_log)
        db.session.commit()
        
        flash("Fuel log entry created successfully.", "success")
        return redirect(url_for("vehicle_fuel_log", id=id))
    
    today = date.today().strftime("%Y-%m-%d")
    fuel_types = ["Diesel", "Petrol", "CNG", "Electric"]
    payment_methods = ["Cash", "Card", "Fleet Card", "UPI", "Credit"]
    
    return render_template(
        "vehicles/fuel_form.html",
        vehicle=vehicle,
        drivers=drivers,
        fuel_types=fuel_types,
        payment_methods=payment_methods,
        today=today,
        edit=False,
    )


@app.route("/vehicles/<int:vehicle_id>/fuel-log/<int:log_id>/edit", methods=["GET", "POST"])
@permission_required("vehicles", "edit")
def edit_vehicle_fuel_log(vehicle_id, log_id):
    """Edit a fuel log entry."""
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    fuel_log = VehicleFuelLog.query.filter_by(id=log_id, vehicle_id=vehicle_id).first_or_404()
    drivers = scoped_query(Driver).filter_by(vehicle_id=vehicle_id).all()
    
    if request.method == "POST":
        fuel_log.fueling_date = request.form.get("fueling_date")
        fuel_log.odometer_reading = int(request.form.get("odometer_reading", 0))
        fuel_log.fuel_liters = float(request.form.get("fuel_liters", 0))
        fuel_log.fuel_price_per_liter = float(request.form.get("fuel_price_per_liter", 0)) if request.form.get("fuel_price_per_liter") else None
        fuel_log.total_cost = float(request.form.get("total_cost", 0))
        fuel_log.fuel_station = request.form.get("fuel_station")
        fuel_log.fuel_type = request.form.get("fuel_type")
        fuel_log.driver_id = request.form.get("driver_id") or None
        fuel_log.payment_method = request.form.get("payment_method")
        fuel_log.receipt_number = request.form.get("receipt_number")
        fuel_log.notes = request.form.get("notes")
        
        # Handle receipt upload
        if "receipt_file" in request.files:
            file = request.files["receipt_file"]
            if file and file.filename:
                fuel_log.receipt_path = save_uploaded_file(file, "fuel_receipts", f"fuel_{vehicle_id}")
        
        # Recalculate efficiency
        fuel_log = calculate_fuel_efficiency(fuel_log)
        
        db.session.commit()
        flash("Fuel log entry updated successfully.", "success")
        return redirect(url_for("vehicle_fuel_log", id=vehicle_id))
    
    fuel_types = ["Diesel", "Petrol", "CNG", "Electric"]
    payment_methods = ["Cash", "Card", "Fleet Card", "UPI", "Credit"]
    
    return render_template(
        "vehicles/fuel_form.html",
        vehicle=vehicle,
        fuel_log=fuel_log,
        drivers=drivers,
        fuel_types=fuel_types,
        payment_methods=payment_methods,
        edit=True,
    )


@app.route("/vehicles/<int:vehicle_id>/fuel-log/<int:log_id>/delete", methods=["POST"])
@permission_required("vehicles", "delete")
def delete_vehicle_fuel_log(vehicle_id, log_id):
    """Delete a fuel log entry."""
    fuel_log = VehicleFuelLog.query.filter_by(id=log_id, vehicle_id=vehicle_id).first_or_404()
    
    try:
        db.session.delete(fuel_log)
        db.session.commit()
        flash("Fuel log entry deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting fuel log: {str(e)}", "error")
    
    return redirect(url_for("vehicle_fuel_log", id=vehicle_id))


def calculate_fuel_efficiency(fuel_log):
    """Calculate fuel efficiency based on previous fuel log entry."""
    # Get previous fuel log
    previous = VehicleFuelLog.query.filter(
        VehicleFuelLog.vehicle_id == fuel_log.vehicle_id,
        VehicleFuelLog.fueling_date < fuel_log.fueling_date,
        VehicleFuelLog.id != fuel_log.id
    ).order_by(VehicleFuelLog.fueling_date.desc()).first()
    
    if previous and fuel_log.odometer_reading and previous.odometer_reading:
        distance = fuel_log.odometer_reading - previous.odometer_reading
        fuel_log.distance_since_last = distance
        
        if fuel_log.fuel_liters and float(fuel_log.fuel_liters) > 0:
            from decimal import Decimal
            fuel_log.efficiency_km_per_liter = Decimal(str(distance)) / fuel_log.fuel_liters
            if distance > 0 and fuel_log.total_cost:
                fuel_log.cost_per_km = fuel_log.total_cost / Decimal(str(distance))
    
    return fuel_log


# =============================================================================
# DOCUMENT EXPIRY ALERTS
# =============================================================================

def get_document_status(expiry_date):
    """Calculate document expiry status and message."""
    if not expiry_date:
        return 'unknown', 'No date set', 0
    
    today = date.today()
    days_remaining = (expiry_date - today).days
    
    if days_remaining < 0:
        return 'expired', 'Expired {} days ago'.format(abs(days_remaining)), days_remaining
    elif days_remaining <= 7:
        return 'critical', '{} days left'.format(days_remaining), days_remaining
    elif days_remaining <= 14:
        return 'warning', '{} days left'.format(days_remaining), days_remaining
    elif days_remaining <= 30:
        return 'expiring', '{} days left'.format(days_remaining), days_remaining
    else:
        return 'valid', '{} days left'.format(days_remaining), days_remaining


def get_vehicle_document_alerts(vehicle):
    """Get all document alerts for a vehicle."""
    alerts = []
    
    docs = [
        ('Insurance', vehicle.insurance_expiry),
        ('Fitness Certificate', vehicle.fitness_expiry),
        ('Permit 1-Year', vehicle.permit_1_year_expiry),
        ('Permit 5-Year', vehicle.permit_5_year_expiry),
        ('Road Tax', vehicle.road_tax_expiry),
        ('PUC', vehicle.puc_expiry),
    ]
    
    for doc_name, expiry in docs:
        status, message, days = get_document_status(expiry)
        if status in ['expired', 'critical', 'warning', 'expiring']:
            alerts.append({
                'document': doc_name,
                'status': status,
                'message': message,
                'expiry_date': expiry,
                'days': days
            })
    
    return alerts


@app.route("/vehicle-documents/alerts")
@permission_required("vehicles", "view")
def vehicle_documents_alerts():
    """Master dashboard for document expiry alerts across fleet."""
    # Get filter parameters
    days_filter = request.args.get("days", "30")
    doc_type = request.args.get("doc_type", "")
    
    try:
        days_limit = int(days_filter)
    except ValueError:
        days_limit = 30
    
    vehicles = scoped_query(Vehicle).filter_by(status="Active").all()
    
    alert_data = []
    for vehicle in vehicles:
        alerts = get_vehicle_document_alerts(vehicle)
        
        # Filter by days if specified
        if days_limit:
            alerts = [a for a in alerts if a['days'] <= days_limit]
        
        if alerts:
            alert_data.append({
                'vehicle': vehicle,
                'alerts': alerts,
                'critical_count': len([a for a in alerts if a['status'] == 'critical']),
                'expired_count': len([a for a in alerts if a['status'] == 'expired'])
            })
    
    # Sort by severity (expired first, then critical)
    alert_data.sort(key=lambda x: (x['expired_count'], x['critical_count']), reverse=True)
    
    # Statistics
    total_expired = sum(len([a for a in d['alerts'] if a['status'] == 'expired']) for d in alert_data)
    total_critical = sum(len([a for a in d['alerts'] if a['status'] == 'critical']) for d in alert_data)
    total_warning = sum(len([a for a in d['alerts'] if a['status'] in ['warning', 'expiring']]) for d in alert_data)
    
    return render_template(
        "vehicles/document_alerts.html",
        alert_data=alert_data,
        total_expired=total_expired,
        total_critical=total_critical,
        total_warning=total_warning,
        days_filter=days_filter,
        doc_type=doc_type,
        vehicles_count=len(alert_data)
    )


@app.route("/vehicles/<int:id>/documents/update-expiry", methods=["POST"])
@permission_required("vehicles", "edit")
def update_document_expiry(id):
    """Update document expiry dates from vehicle dashboard."""
    vehicle = Vehicle.query.get_or_404(id)
    
    # Update expiry dates from form
    if request.form.get("insurance_expiry"):
        vehicle.insurance_expiry = datetime.strptime(request.form.get("insurance_expiry"), "%Y-%m-%d").date()
    if request.form.get("fitness_expiry"):
        vehicle.fitness_expiry = datetime.strptime(request.form.get("fitness_expiry"), "%Y-%m-%d").date()
    if request.form.get("permit_1_year_expiry"):
        vehicle.permit_1_year_expiry = datetime.strptime(request.form.get("permit_1_year_expiry"), "%Y-%m-%d").date()
    if request.form.get("permit_5_year_expiry"):
        vehicle.permit_5_year_expiry = datetime.strptime(request.form.get("permit_5_year_expiry"), "%Y-%m-%d").date()
    if request.form.get("road_tax_expiry"):
        vehicle.road_tax_expiry = datetime.strptime(request.form.get("road_tax_expiry"), "%Y-%m-%d").date()
    if request.form.get("puc_expiry"):
        vehicle.puc_expiry = datetime.strptime(request.form.get("puc_expiry"), "%Y-%m-%d").date()
    
    db.session.commit()
    flash("Document expiry dates updated successfully.", "success")
    
    return redirect(url_for("vehicle_dashboard", id=id))


# =============================================================================
# VEHICLE UTILIZATION METRICS
# =============================================================================

def calculate_vehicle_utilization(vehicle_id, start_date, end_date):
    """Calculate utilization metrics for a vehicle over a date range."""
    total_days = (end_date - start_date).days + 1
    
    # Get all days with at least one trip
    from sqlalchemy import func
    active_dates = db.session.query(
        func.date(DispatchTrip.trip_date).label('trip_date')
    ).filter(
        DispatchTrip.vehicle_id == vehicle_id,
        DispatchTrip.trip_date.between(start_date, end_date)
    ).distinct().all()
    
    active_days = len(active_dates)
    idle_days = total_days - active_days
    utilization_pct = (active_days / total_days * 100) if total_days > 0 else 0
    
    # Get trip count
    total_trips = DispatchTrip.query.filter(
        DispatchTrip.vehicle_id == vehicle_id,
        DispatchTrip.trip_date.between(start_date, end_date)
    ).count()
    
    # Calculate revenue from related TransportBills
    total_revenue = db.session.query(func.sum(TransportBill.rate)).join(
        DispatchTrip, DispatchTrip.bilty_id == TransportBill.id
    ).filter(
        DispatchTrip.vehicle_id == vehicle_id,
        DispatchTrip.trip_date.between(start_date, end_date)
    ).scalar() or 0
    
    return {
        'total_days': total_days,
        'active_days': active_days,
        'idle_days': idle_days,
        'utilization_pct': round(utilization_pct, 2),
        'total_trips': total_trips,
        'avg_trips_per_active_day': round(total_trips / active_days, 2) if active_days > 0 else 0,
        'total_revenue': float(total_revenue),
        'revenue_per_active_day': round(float(total_revenue) / active_days, 2) if active_days > 0 else 0,
        'revenue_per_day': round(float(total_revenue) / total_days, 2) if total_days > 0 else 0
    }


def get_fleet_average_utilization(start_date, end_date):
    """Calculate fleet-wide average utilization."""
    from sqlalchemy import func
    vehicles = Vehicle.query.filter_by(status='Active').all()
    utilization_values = []
    
    for v in vehicles:
        util = calculate_vehicle_utilization(v.id, start_date, end_date)
        utilization_values.append(util['utilization_pct'])
    
    return sum(utilization_values) / len(utilization_values) if utilization_values else 0


def get_monthly_utilization(vehicle_id, months=6):
    """Get monthly utilization breakdown for charts."""
    results = []
    today = date.today()
    
    for i in range(months - 1, -1, -1):
        month_date = today - timedelta(days=30*i)
        start = month_date.replace(day=1)
        # Calculate end of month
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)
        
        # Don't go beyond today
        if end > today:
            end = today
        
        util = calculate_vehicle_utilization(vehicle_id, start, end)
        results.append({
            'month': start.strftime('%b %Y'),
            'month_key': start.strftime('%Y-%m'),
            'utilization_pct': util['utilization_pct'],
            'active_days': util['active_days'],
            'idle_days': util['idle_days'],
            'total_trips': util['total_trips'],
            'total_revenue': util['total_revenue']
        })
    
    return results


@app.route("/vehicles/<int:id>/utilization")
@permission_required("vehicles", "view")
def vehicle_utilization(id):
    """Display vehicle utilization metrics and analytics."""
    vehicle = Vehicle.query.get_or_404(id)
    
    # Period selection
    period = request.args.get("period", "3m")
    today = date.today()
    
    if period == "1m":
        start_date = (today - timedelta(days=30)).replace(day=1)
        period_label = "Last Month"
    elif period == "3m":
        start_date = today - timedelta(days=90)
        period_label = "Last 3 Months"
    elif period == "6m":
        start_date = today - timedelta(days=180)
        period_label = "Last 6 Months"
    elif period == "1y":
        start_date = today - timedelta(days=365)
        period_label = "Last Year"
    else:
        start_date = today - timedelta(days=90)
        period_label = "Last 3 Months"
    
    end_date = today
    
    # Calculate utilization
    utilization = calculate_vehicle_utilization(vehicle.id, start_date, end_date)
    
    # Fleet average for comparison
    fleet_avg = get_fleet_average_utilization(start_date, end_date)
    
    # Monthly breakdown for charts
    monthly_data = get_monthly_utilization(vehicle.id, 6)
    
    # Underutilization alerts
    alerts = []
    if utilization['utilization_pct'] < 50:
        alerts.append({
            'type': 'warning',
            'message': 'Vehicle utilization is below 50%. Consider reassignment or disposal.'
        })
    
    # Check for consecutive idle days (simplified - in real implementation would query trip dates)
    if utilization['idle_days'] > 7 and utilization['utilization_pct'] < 30:
        alerts.append({
            'type': 'info',
            'message': 'Vehicle has been idle for extended periods. Review deployment strategy.'
        })
    
    return render_template(
        "vehicles/utilization.html",
        vehicle=vehicle,
        utilization=utilization,
        fleet_avg=round(fleet_avg, 2),
        monthly_data=monthly_data,
        period=period,
        period_label=period_label,
        alerts=alerts
    )


@app.route("/fleet/utilization-report")
@permission_required("vehicles", "view")
def fleet_utilization_report():
    """Fleet-wide utilization comparison report."""
    period = request.args.get("period", "3m")
    today = date.today()
    
    if period == "1m":
        start_date = (today - timedelta(days=30)).replace(day=1)
    elif period == "6m":
        start_date = today - timedelta(days=180)
    elif period == "1y":
        start_date = today - timedelta(days=365)
    else:
        start_date = today - timedelta(days=90)
    
    end_date = today
    
    vehicles = Vehicle.query.filter_by(status='Active').all()
    
    fleet_data = []
    for vehicle in vehicles:
        util = calculate_vehicle_utilization(vehicle.id, start_date, end_date)
        fleet_data.append({
            'vehicle': vehicle,
            'utilization': util
        })
    
    # Sort by utilization (highest first)
    fleet_data.sort(key=lambda x: x['utilization']['utilization_pct'], reverse=True)
    
    # Fleet average
    fleet_avg = get_fleet_average_utilization(start_date, end_date)
    
    return render_template(
        "vehicles/fleet_utilization.html",
        fleet_data=fleet_data,
        fleet_avg=round(fleet_avg, 2),
        period=period,
        start_date=start_date,
        end_date=end_date
    )


# =============================================================================
# VEHICLE ASSIGNMENT HISTORY
# =============================================================================

@app.route("/vehicles/<int:id>/assignments")
@permission_required("vehicles", "view")
def vehicle_assignments(id):
    """Display vehicle driver assignment history."""
    vehicle = Vehicle.query.get_or_404(id)
    
    # Get all assignments ordered by date (newest first)
    assignments = VehicleDriverAssignment.query.filter_by(
        vehicle_id=id
    ).order_by(VehicleDriverAssignment.assigned_at.desc()).all()
    
    # Get current assignment
    current_assignment = VehicleDriverAssignment.query.filter_by(
        vehicle_id=id,
        is_active=True
    ).first()
    
    # Calculate stats for each driver who operated this vehicle
    driver_stats = []
    for assignment in assignments:
        if assignment.driver:
            # Get trips by this driver in this vehicle during assignment period
            trip_query = DispatchTrip.query.filter(
                DispatchTrip.vehicle_id == id,
                DispatchTrip.driver_id == assignment.driver_id
            )
            
            if assignment.assignment_date:
                trip_query = trip_query.filter(DispatchTrip.trip_date >= assignment.assignment_date)
            if assignment.end_date:
                trip_query = trip_query.filter(DispatchTrip.trip_date <= assignment.end_date)
            
            trip_count = trip_query.count()
            
            # Calculate revenue
            from sqlalchemy import func
            revenue = db.session.query(func.sum(TransportBill.rate)).join(
                DispatchTrip, DispatchTrip.bilty_id == TransportBill.id
            ).filter(
                DispatchTrip.vehicle_id == id,
                DispatchTrip.driver_id == assignment.driver_id
            ).scalar() or 0
            
            driver_stats.append({
                'assignment': assignment,
                'trip_count': trip_count,
                'revenue': float(revenue)
            })
    
    # Available drivers for reassignment
    available_drivers = scoped_query(Driver).filter_by(status="Active").all()
    
    return render_template(
        "vehicles/assignments.html",
        vehicle=vehicle,
        assignments=assignments,
        current_assignment=current_assignment,
        driver_stats=driver_stats,
        available_drivers=available_drivers
    )


@app.route("/vehicles/<int:id>/assignments/change", methods=["POST"])
@permission_required("vehicles", "edit")
def change_vehicle_driver(id):
    """Change driver assignment for a vehicle."""
    vehicle = Vehicle.query.get_or_404(id)
    
    new_driver_id = request.form.get("driver_id")
    assignment_notes = request.form.get("assignment_notes", "")
    end_reason = request.form.get("end_reason", "Driver reassigned")
    
    if not new_driver_id:
        flash("Please select a driver.", "error")
        return redirect(url_for("vehicle_assignments", id=id))
    
    # End current assignment if exists
    current = VehicleDriverAssignment.query.filter_by(
        vehicle_id=id,
        is_active=True
    ).first()
    
    if current:
        current.is_active = False
        current.ended_at = datetime.utcnow()
        current.end_date = date.today()
        current.end_reason = end_reason
        current.ended_by = session.get("user_id")
    
    # Update vehicle's current driver
    vehicle.driver_id = new_driver_id
    
    # Create new assignment record
    new_assignment = VehicleDriverAssignment(
        tenant_id=get_current_tenant_id(),
        vehicle_id=id,
        driver_id=new_driver_id,
        assigned_by=session.get("user_id"),
        assignment_notes=assignment_notes,
        is_active=True
    )
    
    db.session.add(new_assignment)
    db.session.commit()
    
    flash("Driver assignment updated successfully.", "success")
    return redirect(url_for("vehicle_assignments", id=id))


@app.route("/vehicles/<int:id>/assignments/remove", methods=["POST"])
@permission_required("vehicles", "edit")
def remove_vehicle_driver(id):
    """Remove driver assignment (unassign vehicle)."""
    vehicle = Vehicle.query.get_or_404(id)
    
    # End current assignment
    current = VehicleDriverAssignment.query.filter_by(
        vehicle_id=id,
        is_active=True
    ).first()
    
    if current:
        current.is_active = False
        current.ended_at = datetime.utcnow()
        current.end_date = date.today()
        current.end_reason = request.form.get("end_reason", "Vehicle unassigned")
        current.ended_by = session.get("user_id")
    
    # Remove driver from vehicle
    vehicle.driver_id = None
    
    db.session.commit()
    flash("Driver unassigned successfully.", "success")
    return redirect(url_for("vehicle_assignments", id=id))


@app.route("/drivers")
@permission_required("drivers", "view")
def drivers():
    driver_list = scoped_query(Driver).order_by(Driver.first_name).all()
    return render_template(
        "drivers/list.html",
        drivers=driver_list,
        show_tenant_column=is_superadmin(),
    )


@app.route("/drivers/create", methods=["GET", "POST"])
@permission_required("drivers", "create")
def create_driver():
    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            return render_template("drivers/form.html", **get_form_tenant_context())

        driver = Driver(
            tenant_id=tenant_id,
            driver_code=generate_driver_code(),
            # Personal Information
            first_name=request.form.get("first_name", "").strip() or None,
            last_name=request.form.get("last_name", "").strip() or None,
            father_name=request.form.get("father_name", "").strip() or None,
            date_of_birth=parse_date(request.form.get("date_of_birth")),
            gender=request.form.get("gender", "").strip() or None,
            blood_group=request.form.get("blood_group", "").strip() or None,
            nationality=request.form.get("nationality", "Indian").strip() or "Indian",
            marital_status=request.form.get("marital_status", "").strip() or None,
            # Contact Information
            mobile_number=request.form.get("mobile_number", "").strip() or None,
            alternate_mobile=request.form.get("alternate_mobile", "").strip() or None,
            email=request.form.get("email", "").strip() or None,
            emergency_contact_name=request.form.get("emergency_contact_name", "").strip() or None,
            emergency_contact_number=request.form.get("emergency_contact_number", "").strip() or None,
            emergency_contact_relation=request.form.get("emergency_contact_relation", "").strip() or None,
            # Address Information
            address_line1=request.form.get("address_line1", "").strip() or None,
            address_line2=request.form.get("address_line2", "").strip() or None,
            city=request.form.get("city", "").strip() or None,
            state=request.form.get("state", "").strip() or None,
            pincode=request.form.get("pincode", "").strip() or None,
            country=request.form.get("country", "India").strip() or "India",
            # Employment Details
            employee_code=request.form.get("employee_code", "").strip() or None,
            date_of_joining=parse_date(request.form.get("date_of_joining")),
            designation=request.form.get("designation", "Driver").strip() or "Driver",
            department=request.form.get("department", "").strip() or None,
            employment_type=request.form.get("employment_type", "Full-time").strip() or "Full-time",
            status=request.form.get("status", "Active").strip() or "Active",
            monthly_salary=parse_float(request.form.get("monthly_salary")),
            daily_wage=parse_float(request.form.get("daily_wage")),
            # Experience & Qualifications
            total_experience_years=parse_int(request.form.get("total_experience_years")),
            license_type=request.form.get("license_type", "").strip() or None,
            license_number=request.form.get("license_number", "").strip() or None,
            license_issue_date=parse_date(request.form.get("license_issue_date")),
            license_expiry_date=parse_date(request.form.get("license_expiry_date")),
            license_issuing_authority=request.form.get("license_issuing_authority", "").strip() or None,
            license_state=request.form.get("license_state", "").strip() or None,
            # Bank Details
            bank_name=request.form.get("bank_name", "").strip() or None,
            bank_branch=request.form.get("bank_branch", "").strip() or None,
            account_holder_name=request.form.get("account_holder_name", "").strip() or None,
            account_number=request.form.get("account_number", "").strip() or None,
            ifsc_code=request.form.get("ifsc_code", "").strip() or None,
            account_type=request.form.get("account_type", "Savings").strip() or "Savings",
            upi_id=request.form.get("upi_id", "").strip() or None,
            # References
            reference1_name=request.form.get("reference1_name", "").strip() or None,
            reference1_contact=request.form.get("reference1_contact", "").strip() or None,
            reference1_address=request.form.get("reference1_address", "").strip() or None,
            reference2_name=request.form.get("reference2_name", "").strip() or None,
            reference2_contact=request.form.get("reference2_contact", "").strip() or None,
            reference2_address=request.form.get("reference2_address", "").strip() or None,
            # Remarks
            remarks=request.form.get("remarks", "").strip() or None,
        )
        db.session.add(driver)
        db.session.flush()  # Get the driver ID before commit
        
        # Handle file uploads
        driver_folder = str(driver.id)
        
        # Photo upload
        if 'photo' in request.files and request.files['photo'].filename:
            photo_path = save_uploaded_file(request.files['photo'], driver_folder, 'photo')
            if photo_path:
                driver.photo_path = photo_path
        
        # Aadhaar upload
        driver.aadhaar_number = request.form.get("aadhaar_number", "").strip() or None
        if 'aadhaar_attachment' in request.files and request.files['aadhaar_attachment'].filename:
            aadhaar_path = save_uploaded_file(request.files['aadhaar_attachment'], driver_folder, 'aadhaar')
            if aadhaar_path:
                driver.aadhaar_attachment_path = aadhaar_path
        elif request.form.get("aadhaar_attachment_path"):
            driver.aadhaar_attachment_path = request.form.get("aadhaar_attachment_path")
        
        # PAN upload
        driver.pan_number = request.form.get("pan_number", "").strip() or None
        if 'pan_attachment' in request.files and request.files['pan_attachment'].filename:
            pan_path = save_uploaded_file(request.files['pan_attachment'], driver_folder, 'pan')
            if pan_path:
                driver.pan_attachment_path = pan_path
        elif request.form.get("pan_attachment_path"):
            driver.pan_attachment_path = request.form.get("pan_attachment_path")
        
        # Driving License upload
        if 'driving_license_attachment' in request.files and request.files['driving_license_attachment'].filename:
            license_path = save_uploaded_file(request.files['driving_license_attachment'], driver_folder, 'license')
            if license_path:
                driver.driving_license_attachment_path = license_path
        elif request.form.get("driving_license_attachment_path"):
            driver.driving_license_attachment_path = request.form.get("driving_license_attachment_path")
        
        # Bank Document upload
        if 'bank_attachment' in request.files and request.files['bank_attachment'].filename:
            bank_path = save_uploaded_file(request.files['bank_attachment'], driver_folder, 'bank')
            if bank_path:
                driver.bank_attachment_path = bank_path
        elif request.form.get("bank_attachment_path"):
            driver.bank_attachment_path = request.form.get("bank_attachment_path")
        
        # Police Verification upload
        if 'police_verification' in request.files and request.files['police_verification'].filename:
            police_path = save_uploaded_file(request.files['police_verification'], driver_folder, 'police')
            if police_path:
                driver.police_verification_path = police_path
        elif request.form.get("police_verification_path"):
            driver.police_verification_path = request.form.get("police_verification_path")
        
        # Medical Certificate upload
        if 'medical_certificate' in request.files and request.files['medical_certificate'].filename:
            medical_path = save_uploaded_file(request.files['medical_certificate'], driver_folder, 'medical')
            if medical_path:
                driver.medical_certificate_path = medical_path
        elif request.form.get("medical_certificate_path"):
            driver.medical_certificate_path = request.form.get("medical_certificate_path")
        
        db.session.commit()
        # Update subscription usage
        update_subscription_usage(tenant_id)
        flash(f"Driver {driver.driver_code} created successfully.", "success")
        return redirect(url_for("drivers"))

    return render_template("drivers/form.html", **get_form_tenant_context())
@app.route("/drivers/edit/<int:id>", methods=["GET", "POST"])
@permission_required("drivers", "edit")
def edit_driver(id):
    driver = get_scoped_record(Driver, id)

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(driver, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(driver)
            context.update({"driver": driver, "edit": True})
            return render_template("drivers/form.html", **context)

        driver.tenant_id = tenant_id
        # Personal Information
        driver.first_name = request.form.get("first_name", "").strip() or None
        driver.last_name = request.form.get("last_name", "").strip() or None
        driver.father_name = request.form.get("father_name", "").strip() or None
        driver.date_of_birth = parse_date(request.form.get("date_of_birth"))
        driver.gender = request.form.get("gender", "").strip() or None
        driver.blood_group = request.form.get("blood_group", "").strip() or None
        driver.nationality = request.form.get("nationality", "Indian").strip() or "Indian"
        driver.marital_status = request.form.get("marital_status", "").strip() or None
        # Contact Information
        driver.mobile_number = request.form.get("mobile_number", "").strip() or None
        driver.alternate_mobile = request.form.get("alternate_mobile", "").strip() or None
        driver.email = request.form.get("email", "").strip() or None
        driver.emergency_contact_name = request.form.get("emergency_contact_name", "").strip() or None
        driver.emergency_contact_number = request.form.get("emergency_contact_number", "").strip() or None
        driver.emergency_contact_relation = request.form.get("emergency_contact_relation", "").strip() or None
        # Address Information
        driver.address_line1 = request.form.get("address_line1", "").strip() or None
        driver.address_line2 = request.form.get("address_line2", "").strip() or None
        driver.city = request.form.get("city", "").strip() or None
        driver.state = request.form.get("state", "").strip() or None
        driver.pincode = request.form.get("pincode", "").strip() or None
        driver.country = request.form.get("country", "India").strip() or "India"
        # Employment Details
        driver.employee_code = request.form.get("employee_code", "").strip() or None
        driver.date_of_joining = parse_date(request.form.get("date_of_joining"))
        driver.designation = request.form.get("designation", "Driver").strip() or "Driver"
        driver.department = request.form.get("department", "").strip() or None
        driver.employment_type = request.form.get("employment_type", "Full-time").strip() or "Full-time"
        driver.status = request.form.get("status", "Active").strip() or "Active"
        driver.monthly_salary = parse_float(request.form.get("monthly_salary"))
        driver.daily_wage = parse_float(request.form.get("daily_wage"))
        # Experience & Qualifications
        driver.total_experience_years = parse_int(request.form.get("total_experience_years"))
        driver.license_type = request.form.get("license_type", "").strip() or None
        driver.license_number = request.form.get("license_number", "").strip() or None
        driver.license_issue_date = parse_date(request.form.get("license_issue_date"))
        driver.license_expiry_date = parse_date(request.form.get("license_expiry_date"))
        driver.license_issuing_authority = request.form.get("license_issuing_authority", "").strip() or None
        driver.license_state = request.form.get("license_state", "").strip() or None
        # Bank Details
        driver.bank_name = request.form.get("bank_name", "").strip() or None
        driver.bank_branch = request.form.get("bank_branch", "").strip() or None
        driver.account_holder_name = request.form.get("account_holder_name", "").strip() or None
        driver.account_number = request.form.get("account_number", "").strip() or None
        driver.ifsc_code = request.form.get("ifsc_code", "").strip() or None
        driver.account_type = request.form.get("account_type", "Savings").strip() or "Savings"
        driver.upi_id = request.form.get("upi_id", "").strip() or None
        # Documents & Attachments - Handle file uploads
        driver_folder = str(driver.id)
        
        # Photo upload
        if 'photo' in request.files and request.files['photo'].filename:
            photo_path = save_uploaded_file(request.files['photo'], driver_folder, 'photo')
            if photo_path:
                driver.photo_path = photo_path
        elif request.form.get("photo_path"):
            driver.photo_path = request.form.get("photo_path")
        
        # Aadhaar
        driver.aadhaar_number = request.form.get("aadhaar_number", "").strip() or None
        if 'aadhaar_attachment' in request.files and request.files['aadhaar_attachment'].filename:
            aadhaar_path = save_uploaded_file(request.files['aadhaar_attachment'], driver_folder, 'aadhaar')
            if aadhaar_path:
                driver.aadhaar_attachment_path = aadhaar_path
        elif request.form.get("aadhaar_attachment_path"):
            driver.aadhaar_attachment_path = request.form.get("aadhaar_attachment_path")
        
        # PAN
        driver.pan_number = request.form.get("pan_number", "").strip() or None
        if 'pan_attachment' in request.files and request.files['pan_attachment'].filename:
            pan_path = save_uploaded_file(request.files['pan_attachment'], driver_folder, 'pan')
            if pan_path:
                driver.pan_attachment_path = pan_path
        elif request.form.get("pan_attachment_path"):
            driver.pan_attachment_path = request.form.get("pan_attachment_path")
        
        # Driving License
        if 'driving_license_attachment' in request.files and request.files['driving_license_attachment'].filename:
            license_path = save_uploaded_file(request.files['driving_license_attachment'], driver_folder, 'license')
            if license_path:
                driver.driving_license_attachment_path = license_path
        elif request.form.get("driving_license_attachment_path"):
            driver.driving_license_attachment_path = request.form.get("driving_license_attachment_path")
        
        # Bank
        if 'bank_attachment' in request.files and request.files['bank_attachment'].filename:
            bank_path = save_uploaded_file(request.files['bank_attachment'], driver_folder, 'bank')
            if bank_path:
                driver.bank_attachment_path = bank_path
        elif request.form.get("bank_attachment_path"):
            driver.bank_attachment_path = request.form.get("bank_attachment_path")
        
        # Police Verification
        if 'police_verification' in request.files and request.files['police_verification'].filename:
            police_path = save_uploaded_file(request.files['police_verification'], driver_folder, 'police')
            if police_path:
                driver.police_verification_path = police_path
        elif request.form.get("police_verification_path"):
            driver.police_verification_path = request.form.get("police_verification_path")
        
        # Medical
        if 'medical_certificate' in request.files and request.files['medical_certificate'].filename:
            medical_path = save_uploaded_file(request.files['medical_certificate'], driver_folder, 'medical')
            if medical_path:
                driver.medical_certificate_path = medical_path
        elif request.form.get("medical_certificate_path"):
            driver.medical_certificate_path = request.form.get("medical_certificate_path")
        
        # References
        driver.reference1_name = request.form.get("reference1_name", "").strip() or None
        driver.reference1_contact = request.form.get("reference1_contact", "").strip() or None
        driver.reference1_address = request.form.get("reference1_address", "").strip() or None
        driver.reference2_name = request.form.get("reference2_name", "").strip() or None
        driver.reference2_contact = request.form.get("reference2_contact", "").strip() or None
        driver.reference2_address = request.form.get("reference2_address", "").strip() or None
        # Remarks
        driver.remarks = request.form.get("remarks", "").strip() or None

        db.session.commit()
        # Update subscription usage
        update_subscription_usage(tenant_id)
        flash(f"Driver {driver.driver_code} updated successfully.", "success")
        return redirect(url_for("drivers"))

    context = get_form_tenant_context(driver)
    context.update({"driver": driver, "edit": True})
    return render_template("drivers/form.html", **context)


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve uploaded files"""
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route("/drivers/delete/<int:id>", methods=["POST"])
@permission_required("drivers", "delete")
def delete_driver(id):
    driver = get_scoped_record(Driver, id)
    try:
        db.session.delete(driver)
        db.session.commit()
        flash("Driver deleted successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Cannot delete driver.", "error")
    return redirect(url_for("drivers"))


# ==================== EXPORT ROUTES ====================

@app.route("/drivers/export")
@permission_required("drivers", "view")
def export_drivers():
    """Export drivers to Excel with all business fields and dropdown validation"""
    drivers = scoped_query(Driver).all()
    
    # All business fields (40 fields) - organized by category
    headers = [
        # Personal Information
        'First Name', 'Last Name', 'Father Name', 'Date of Birth', 'Gender', 'Blood Group',
        'Nationality', 'Marital Status',
        # Contact Information
        'Mobile Number', 'Alternate Mobile', 'Email', 'Emergency Contact Name',
        'Emergency Contact Number', 'Emergency Contact Relation',
        # Address Information
        'Address Line 1', 'Address Line 2', 'City', 'State', 'Pincode', 'Country',
        # Employment Details
        'Employee Code', 'Date of Joining', 'Designation', 'Department', 'Employment Type',
        'Status', 'Monthly Salary', 'Daily Wage',
        # License & Experience
        'License Type', 'License Number', 'License Issue Date', 'License Expiry Date',
        'License Issuing Authority', 'License State', 'Total Experience Years',
        # Bank Details
        'Bank Name', 'Bank Branch', 'Account Holder Name', 'Account Number', 'IFSC Code',
        'Account Type', 'UPI ID',
        # Documents
        'Aadhaar Number', 'PAN Number',
        # References
        'Reference 1 Name', 'Reference 1 Contact', 'Reference 2 Name', 'Reference 2 Contact',
        # Remarks
        'Remarks'
    ]
    
    # Dropdown configurations
    dropdowns = {
        'Gender': ['Male', 'Female', 'Other'],
        'Blood Group': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
        'Nationality': ['Indian', 'Other'],
        'Marital Status': ['Single', 'Married', 'Divorced', 'Widowed'],
        'Designation': ['Driver', 'Senior Driver', 'Trainer', 'Supervisor'],
        'Department': ['Operations', 'Transport', 'Logistics'],
        'Employment Type': ['Full-time', 'Part-time', 'Contract', 'Temporary'],
        'Status': ['Active', 'Inactive', 'Suspended', 'Terminated', 'On Leave'],
        'License Type': ['LMV', 'HMV', 'HGMV', 'MCWG', 'MCWOG'],
        'Account Type': ['Savings', 'Current'],
        'Country': ['India', 'Other']
    }
    
    data = []
    for driver in drivers:
        data.append([
            # Personal
            driver.first_name or '',
            driver.last_name or '',
            driver.father_name or '',
            driver.date_of_birth.strftime('%Y-%m-%d') if driver.date_of_birth else '',
            driver.gender or '',
            driver.blood_group or '',
            driver.nationality or '',
            driver.marital_status or '',
            # Contact
            driver.mobile_number or '',
            driver.alternate_mobile or '',
            driver.email or '',
            driver.emergency_contact_name or '',
            driver.emergency_contact_number or '',
            driver.emergency_contact_relation or '',
            # Address
            driver.address_line1 or '',
            driver.address_line2 or '',
            driver.city or '',
            driver.state or '',
            driver.pincode or '',
            driver.country or '',
            # Employment
            driver.employee_code or '',
            driver.date_of_joining.strftime('%Y-%m-%d') if driver.date_of_joining else '',
            driver.designation or '',
            driver.department or '',
            driver.employment_type or '',
            driver.status or '',
            float(driver.monthly_salary) if driver.monthly_salary else 0,
            float(driver.daily_wage) if driver.daily_wage else 0,
            # License
            driver.license_type or '',
            driver.license_number or '',
            driver.license_issue_date.strftime('%Y-%m-%d') if driver.license_issue_date else '',
            driver.license_expiry_date.strftime('%Y-%m-%d') if driver.license_expiry_date else '',
            driver.license_issuing_authority or '',
            driver.license_state or '',
            driver.total_experience_years or '',
            # Bank
            driver.bank_name or '',
            driver.bank_branch or '',
            driver.account_holder_name or '',
            driver.account_number or '',
            driver.ifsc_code or '',
            driver.account_type or '',
            driver.upi_id or '',
            # Documents
            driver.aadhaar_number or '',
            driver.pan_number or '',
            # References
            driver.reference1_name or '',
            driver.reference1_contact or '',
            driver.reference2_name or '',
            driver.reference2_contact or '',
            # Remarks
            driver.remarks or ''
        ])
    
    return export_to_excel(data, headers, 'drivers.xlsx', dropdowns)


@app.route("/vehicles/export")
@permission_required("vehicles", "view")
def export_vehicles():
    """Export vehicles to Excel with all business fields and dropdown validation"""
    vehicles = scoped_query(Vehicle).all()
    
    # All business fields (25+ fields)
    headers = [
        # Basic Information
        'Registration Number', 'Vehicle Type', 'Make', 'Model', 'Year', 'Color', 'Fuel Type', 'Truck Size', 'Status',
        # Technical Details
        'Engine Number', 'Chassis Number', 'Seating Capacity', 'Load Capacity (KG)',
        # Owner Information
        'Owner Name', 'Owner Contact', 'Purchase Date',
        # Expiry Dates
        'Insurance Expiry', 'Fitness Expiry', 'Permit 1 Year Expiry', 'Permit 5 Year Expiry',
        'Road Tax Expiry', 'PUC Expiry',
        # Notes
        'Notes'
    ]

    # Dropdown configurations
    dropdowns = {
        'Vehicle Type': ['Truck', 'Trailer', 'Container', 'Tanker', 'Pickup', 'Van', 'Bus', 'Car', 'Bike', 'Tempo', 'LCV'],
        'Fuel Type': ['Diesel', 'Petrol', 'CNG', 'Electric', 'Hybrid'],
        'Truck Size': ['8 Ft', '10 Ft', '14 Ft', '17 Ft', '19 Ft', '20 Ft', '22 Ft', '24 Ft', '28 Ft', '32 Ft', '36 Ft', '40 Ft', 'Container 20 Ft', 'Container 40 Ft'],
        'Status': ['Active', 'Inactive', 'Sold', 'Scrapped', 'Under Maintenance'],
        'Color': ['White', 'Black', 'Blue', 'Red', 'Silver', 'Grey', 'Yellow', 'Green', 'Brown', 'Orange', 'Other']
    }
    
    data = []
    for vehicle in vehicles:
        data.append([
            # Basic
            vehicle.registration_number,
            vehicle.vehicle_type or '',
            vehicle.make or '',
            vehicle.model or '',
            vehicle.year or '',
            vehicle.color or '',
            vehicle.fuel_type or '',
            vehicle.truck_size or '',
            vehicle.status or '',
            # Technical
            vehicle.engine_number or '',
            vehicle.chassis_number or '',
            vehicle.seating_capacity or '',
            vehicle.load_capacity_kg or '',
            # Owner
            vehicle.owner_name or '',
            vehicle.owner_contact or '',
            vehicle.purchase_date.strftime('%Y-%m-%d') if vehicle.purchase_date else '',
            # Expiry Dates
            vehicle.insurance_expiry.strftime('%Y-%m-%d') if vehicle.insurance_expiry else '',
            vehicle.fitness_expiry.strftime('%Y-%m-%d') if vehicle.fitness_expiry else '',
            vehicle.permit_1_year_expiry.strftime('%Y-%m-%d') if vehicle.permit_1_year_expiry else '',
            vehicle.permit_5_year_expiry.strftime('%Y-%m-%d') if vehicle.permit_5_year_expiry else '',
            vehicle.road_tax_expiry.strftime('%Y-%m-%d') if vehicle.road_tax_expiry else '',
            vehicle.puc_expiry.strftime('%Y-%m-%d') if vehicle.puc_expiry else '',
            # Notes
            vehicle.notes or ''
        ])
    
    return export_to_excel(data, headers, 'vehicles.xlsx', dropdowns)


@app.route("/vendors/export")
@permission_required("vendors", "view")
def export_vendors():
    """Export vendors to Excel with all business fields and dropdown validation"""
    vendors = scoped_query(Vendor).all()
    
    # All business fields (45+ fields) - organized by category
    headers = [
        # Basic Information
        'Vendor Name', 'Vendor Type', 'Status', 'Business Nature', 'Website', 'Referral Source',
        # Contact Information
        'Contact Person', 'Designation', 'Phone Primary', 'Phone Secondary', 'Mobile', 'Email', 'Alternate Email',
        # Registered Address
        'Reg Address Line 1', 'Reg Address Line 2', 'Reg City', 'Reg State', 'Reg Pincode', 'Reg Country',
        # Office Address
        'Office Address Line 1', 'Office Address Line 2', 'Office City', 'Office State', 'Office Pincode', 'Office Country',
        # GST & Tax
        'GSTIN', 'GST Registration Date', 'GST State Code', 'PAN Number', 'TAN Number', 'Tax Regime', 'Is Composition Dealer',
        # Bank Details
        'Bank Name', 'Bank Branch', 'Account Number', 'Account Type', 'IFSC Code', 'MICR Code', 'UPI ID',
        # Primary Contact Person
        'Primary Contact Name', 'Primary Contact Designation', 'Primary Contact Mobile', 'Primary Contact Email',
        # Trade References
        'Trade Reference 1 Name', 'Trade Reference 1 Contact',
        'Trade Reference 2 Name', 'Trade Reference 2 Contact',
        # Financial
        'Credit Limit', 'Credit Period Days', 'Opening Balance', 'Currency', 'Payment Terms',
        # Compliance
        'KYC Status', 'Compliance Rating',
        # Supply
        'Supply Type', 'Lead Time Days',
        # Remarks
        'Remarks'
    ]
    
    # Dropdown configurations
    dropdowns = {
        'Vendor Type': ['supplier', 'transporter', 'broker', 'agent', 'service_provider', 'contractor', 'consultant'],
        'Status': ['active', 'inactive', 'blacklisted', 'on_hold', 'suspended'],
        'Reg Country': ['India', 'Other'],
        'Office Country': ['India', 'Other'],
        'Tax Regime': ['regular', 'composition', 'unregistered'],
        'Account Type': ['savings', 'current', 'cash_credit', 'overdraft'],
        'Currency': ['INR', 'USD', 'EUR', 'GBP'],
        'KYC Status': ['pending', 'verified', 'rejected', 'under_review'],
        'Compliance Rating': ['unrated', 'low', 'medium', 'high'],
        'Supply Type': ['goods', 'services', 'both'],
        'Is Composition Dealer': ['Yes', 'No']
    }
    
    data = []
    for vendor in vendors:
        data.append([
            # Basic
            vendor.vendor_name or '',
            vendor.vendor_type or '',
            vendor.status or '',
            vendor.business_nature or '',
            vendor.website or '',
            vendor.referral_source or '',
            # Contact
            vendor.contact_person or '',
            vendor.designation or '',
            vendor.phone_primary or '',
            vendor.phone_secondary or '',
            vendor.mobile or '',
            vendor.email or '',
            vendor.alternate_email or '',
            # Registered Address
            vendor.reg_address_line1 or '',
            vendor.reg_address_line2 or '',
            vendor.reg_city or '',
            vendor.reg_state or '',
            vendor.reg_pincode or '',
            vendor.reg_country or '',
            # Office Address
            vendor.office_address_line1 or '',
            vendor.office_address_line2 or '',
            vendor.office_city or '',
            vendor.office_state or '',
            vendor.office_pincode or '',
            vendor.office_country or '',
            # GST & Tax
            vendor.gstin or '',
            vendor.gst_registration_date.strftime('%Y-%m-%d') if vendor.gst_registration_date else '',
            vendor.gst_state_code or '',
            vendor.pan_number or '',
            vendor.tan_number or '',
            vendor.tax_regime or '',
            'Yes' if vendor.is_composition_dealer else 'No',
            # Bank
            vendor.bank_name or '',
            vendor.bank_branch or '',
            vendor.account_number or '',
            vendor.account_type or '',
            vendor.ifsc_code or '',
            vendor.micr_code or '',
            vendor.upi_id or '',
            # Primary Contact
            vendor.primary_contact_name or '',
            vendor.primary_contact_designation or '',
            vendor.primary_contact_mobile or '',
            vendor.primary_contact_email or '',
            # Trade References
            vendor.trade_reference_1_name or '',
            vendor.trade_reference_1_contact or '',
            vendor.trade_reference_2_name or '',
            vendor.trade_reference_2_contact or '',
            # Financial
            float(vendor.credit_limit) if vendor.credit_limit else 0,
            vendor.credit_period_days or 0,
            float(vendor.opening_balance) if vendor.opening_balance else 0,
            vendor.currency or 'INR',
            vendor.payment_terms or '',
            # Compliance
            vendor.kyc_status or '',
            vendor.compliance_rating or '',
            # Supply
            vendor.supply_type or '',
            vendor.lead_time_days or 0,
            # Remarks
            vendor.remarks or ''
        ])
    
    return export_to_excel(data, headers, 'vendors.xlsx', dropdowns)


@app.route("/expenses/export")
@permission_required("expenses", "view")
def export_expenses():
    """Export expenses to Excel with all business fields and dropdown validation"""
    expenses = scoped_query(Expense).all()
    
    # All business fields (12 fields)
    headers = [
        'Expense Date', 'Category', 'Description', 'Amount', 'Vehicle Number', 'Location',
        'Payment Method', 'Vendor Name', 'Vendor Contact', 'Bill Number', 'Status', 'Notes'
    ]
    
    # Dropdown configurations
    dropdowns = {
        'Category': ['Fuel', 'Maintenance', 'Insurance', 'Tolls', 'Driver Salary', 'Spare Parts', 'Tyres', 
                     'Lubricants', 'Car Wash', 'Parking', 'Fine/Penalty', 'Other'],
        'Payment Method': ['Cash', 'Card', 'UPI', 'Bank Transfer', 'Cheque', 'Wallet'],
        'Status': ['Pending', 'Approved', 'Rejected', 'Reimbursed']
    }
    
    data = []
    for expense in expenses:
        data.append([
            expense.expense_date.strftime('%Y-%m-%d') if expense.expense_date else '',
            expense.category or '',
            expense.description or '',
            float(expense.amount) if expense.amount else 0,
            expense.vehicle.registration_number if expense.vehicle else '',
            expense.location.location if expense.location else '',
            expense.payment_method or '',
            expense.vendor_name or '',
            expense.vendor_contact or '',
            expense.bill_number or '',
            expense.status or '',
            expense.notes or ''
        ])
    
    return export_to_excel(data, headers, 'expenses.xlsx', dropdowns)


@app.route("/loans/export")
@permission_required("loans", "view")
def export_loans():
    """Export loans to Excel with all business fields and dropdown validation"""
    loans = scoped_query(Loan).all()
    
    # All business fields (25 fields)
    headers = [
        # Basic
        'Loan Type', 'Vehicle Number', 'Principal Amount', 'Interest Rate', 'Tenure Months',
        # EMI Details
        'EMI Amount', 'Total Payable', 'Total Interest', 'Down Payment', 'Amount Paid', 'Balance Amount',
        'EMIs Paid', 'EMIs Remaining',
        # Lender
        'Lender Name', 'Lender Type', 'Lender Contact', 'Lender Address',
        # Agent
        'Agent Name', 'Agent Contact',
        # Dates
        'Loan Date', 'Disbursement Date', 'First EMI Date', 'Last EMI Date', 'Next EMI Due Date',
        # Account
        'Loan Account Number', 'Status', 'Purpose', 'Collateral', 'Insurance Details', 'Remarks'
    ]
    
    # Dropdown configurations
    dropdowns = {
        'Loan Type': ['Truck Loan', 'Body Loan', 'Working Capital', 'Personal Loan', 'Equipment Loan', 'Other'],
        'Lender Type': ['Bank', 'NBFC', 'Private', 'Individual', 'Co-operative'],
        'Status': ['Active', 'Closed', 'Defaulted', 'Pre-closed', 'Pending']
    }
    
    data = []
    for loan in loans:
        data.append([
            # Basic
            loan.loan_type or '',
            loan.vehicle.registration_number if loan.vehicle else '',
            float(loan.principal_amount) if loan.principal_amount else 0,
            float(loan.interest_rate) if loan.interest_rate else 0,
            loan.tenure_months or 0,
            # EMI
            float(loan.emi_amount) if loan.emi_amount else 0,
            float(loan.total_payable) if loan.total_payable else 0,
            float(loan.total_interest) if loan.total_interest else 0,
            float(loan.down_payment) if loan.down_payment else 0,
            float(loan.amount_paid) if loan.amount_paid else 0,
            float(loan.balance_amount) if loan.balance_amount else 0,
            loan.emis_paid or 0,
            loan.emis_remaining or 0,
            # Lender
            loan.lender_name or '',
            loan.lender_type or '',
            loan.lender_contact or '',
            loan.lender_address or '',
            # Agent
            loan.agent_name or '',
            loan.agent_contact or '',
            # Dates
            loan.loan_date.strftime('%Y-%m-%d') if loan.loan_date else '',
            loan.disbursement_date.strftime('%Y-%m-%d') if loan.disbursement_date else '',
            loan.first_emi_date.strftime('%Y-%m-%d') if loan.first_emi_date else '',
            loan.last_emi_date.strftime('%Y-%m-%d') if loan.last_emi_date else '',
            loan.next_emi_due_date.strftime('%Y-%m-%d') if loan.next_emi_due_date else '',
            # Account
            loan.loan_account_number or '',
            loan.status or '',
            loan.purpose or '',
            loan.collateral or '',
            loan.insurance_details or '',
            loan.remarks or ''
        ])
    
    return export_to_excel(data, headers, 'loans.xlsx', dropdowns)


@app.route("/locations/export")
@permission_required("locations", "view")
def export_locations():
    """Export rate list to Excel with all fields and dropdown validation"""
    locations = scoped_query(Location).all()
    
    # All business fields (6 core fields)
    headers = ['Location Name', 'City', 'State', 'Pincode', 'Rate', 'Distance (KM)', 'Remarks']
    
    # Indian states dropdown
    dropdowns = {
        'State': ['Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chhattisgarh', 'Goa', 'Gujarat', 
                  'Haryana', 'Himachal Pradesh', 'Jharkhand', 'Karnataka', 'Kerala', 'Madhya Pradesh', 
                  'Maharashtra', 'Manipur', 'Meghalaya', 'Mizoram', 'Nagaland', 'Odisha', 'Punjab', 
                  'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana', 'Tripura', 'Uttar Pradesh', 
                  'Uttarakhand', 'West Bengal', 'Delhi', 'Jammu and Kashmir', 'Ladakh', 'Other']
    }
    
    data = []
    for location in locations:
        data.append([
            location.location,
            location.city or '',
            location.state or '',
            location.pincode or '',
            float(location.rate) if location.rate else 0,
            location.distance_km or 0,
            location.remarks or ''
        ])
    
    return export_to_excel(data, headers, 'rate_list.xlsx', dropdowns)


@app.route("/transport-bills/export")
@permission_required("transport_bills", "view")
def export_transport_bills():
    """Export transport bills to Excel with comprehensive business fields and dropdown validation"""
    bills = scoped_query(TransportBill).all()
    
    # Comprehensive business fields (~80 fields) - key fields only for practicality
    headers = [
        # GR Basic Details
        'GR Number', 'GR Date', 'GR Type', 'Booking Mode', 'Status', 'Reference Number',
        # Party Details - Consignor
        'Consignor Name', 'Consignor GSTIN', 'Consignor Contact', 'Consignor Mobile', 'Consignor Email',
        # Party Details - Consignee
        'Consignee Name', 'Consignee GSTIN', 'Consignee Contact', 'Consignee Mobile', 'Consignee Email',
        'Bill To Party', 'Ship To Party',
        # Pickup Details
        'Pickup Point', 'Pickup Date', 'Pickup City', 'Pickup State', 'Loaded By', 'Loading Type',
        # Delivery Details
        'Delivery Location', 'Expected Delivery Date', 'Actual Delivery Date', 
        'Delivery City', 'Delivery State', 'Unloading Point',
        # Vehicle Details
        'Vehicle Number', 'Vehicle Type', 'Vehicle Capacity', 'Vehicle Ownership',
        'Driver Name', 'Driver Mobile', 'Driver License', 'Helper Name',
        # Route
        'Origin', 'Destination', 'Via Route',
        # Material Details
        'Article Description', 'Package Type', 'Number of Packages', 'Packing Mode',
        'Actual Weight', 'Charged Weight', 'Volume CBM', 'Material Value', 'Dimensions',
        'Goods Category', 'HSN Code', 'Declared Value', 'Handling Instructions',
        # Freight Charges
        'Basic Freight', 'Loading Charge', 'Unloading Charge', 'Door Pickup Charge', 'Door Delivery Charge',
        'Hamali Charge', 'Toll Charge', 'Fuel Surcharge', 'Insurance Charge', 'Other Charges',
        'Discount', 'Net Freight',
        # Tax Details
        'GST Applicable', 'IGST Rate', 'CGST Rate', 'SGST Rate', 'Taxable Amount', 'GST Amount',
        'Eway Bill Number', 'Invoice Number',
        # Payment Details
        'Freight Payment Mode', 'Advance Amount', 'Balance Amount', 'Paid Amount', 'Payment Status',
        # Transit Details
        'Dispatch DateTime', 'Tracking Number', 'Current Shipment Status', 'Expected Transit Time',
        # Delivery Confirmation
        'Delivered DateTime', 'Received By Name', 'POD Number', 'Delivery Remarks',
        # Risk/Compliance
        'Hazardous Material', 'Temperature Controlled', 'GPS Link', 'Permit Details',
        # Legacy
        'Challan Number', 'Party Information', 'Rate'
    ]
    
    # Dropdown configurations for enumerated fields
    dropdowns = {
        'Booking Mode': ['road', 'rail', 'air', 'sea'],
        'Status': ['draft', 'confirmed', 'in_transit', 'delivered', 'cancelled'],
        'GR Type': ['regular', 'urgent', 'express', 'premium'],
        'Vehicle Ownership': ['owned', 'hired', 'attached'],
        'Loading Type': ['manual', 'mechanical', 'forklift'],
        'Packing Mode': ['loose', 'packed', 'palletized', 'containerized'],
        'Goods Category': ['general', 'fragile', 'perishable', 'hazardous', 'valuable', 'automotive', 'electronics', 'furniture'],
        'Freight Payment Mode': ['to_pay', 'paid', 'to_be_billed'],
        'Payment Status': ['pending', 'partial', 'paid'],
        'Current Shipment Status': ['booking_confirmed', 'vehicle_assigned', 'in_transit', 'at_hub', 'out_for_delivery', 'delivered', 'cancelled'],
        'GST Applicable': ['Yes', 'No'],
        'Hazardous Material': ['Yes', 'No'],
        'Temperature Controlled': ['Yes', 'No'],
        'Package Type': ['box', 'carton', 'bag', 'drum', 'bundle', 'roll', 'crate', 'pallet', 'container']
    }
    
    data = []
    for bill in bills:
        data.append([
            # GR Basic
            bill.gr_number or '',
            bill.gr_date.strftime('%Y-%m-%d') if bill.gr_date else '',
            bill.gr_type or '',
            bill.booking_mode or '',
            bill.status or '',
            bill.reference_number or '',
            # Consignor
            bill.consignor_name or '',
            bill.consignor_gstin or '',
            bill.consignor_contact_person or '',
            bill.consignor_mobile or '',
            bill.consignor_email or '',
            # Consignee
            bill.consignee_name or '',
            bill.consignee_gstin or '',
            bill.consignee_contact_person or '',
            bill.consignee_mobile or '',
            bill.consignee_email or '',
            bill.bill_to_party or '',
            bill.ship_to_party or '',
            # Pickup
            bill.pickup_point or '',
            bill.pickup_date.strftime('%Y-%m-%d') if bill.pickup_date else '',
            bill.pickup_city or '',
            bill.pickup_state or '',
            bill.loaded_by or '',
            bill.loading_type or '',
            # Delivery
            bill.delivery_location or '',
            bill.expected_delivery_date.strftime('%Y-%m-%d') if bill.expected_delivery_date else '',
            bill.actual_delivery_date.strftime('%Y-%m-%d') if bill.actual_delivery_date else '',
            bill.delivery_city or '',
            bill.delivery_state or '',
            bill.unloading_point or '',
            # Vehicle
            bill.vehicle_number or '',
            bill.vehicle_type or '',
            bill.vehicle_capacity or '',
            bill.vehicle_ownership or '',
            bill.driver_name or '',
            bill.driver_mobile or '',
            bill.driver_license or '',
            bill.helper_name or '',
            # Route
            bill.origin or '',
            bill.destination or '',
            bill.via_route or '',
            # Material
            bill.article_description or '',
            bill.package_type or '',
            bill.number_of_packages or 0,
            bill.packing_mode or '',
            float(bill.actual_weight) if bill.actual_weight else 0,
            float(bill.charged_weight) if bill.charged_weight else 0,
            float(bill.volume_cbm) if bill.volume_cbm else 0,
            float(bill.material_value) if bill.material_value else 0,
            bill.dimensions or '',
            bill.goods_category or '',
            bill.hsn_code or '',
            float(bill.declared_value) if bill.declared_value else 0,
            bill.handling_instructions or '',
            # Freight
            float(bill.basic_freight) if bill.basic_freight else 0,
            float(bill.loading_charge) if bill.loading_charge else 0,
            float(bill.unloading_charge) if bill.unloading_charge else 0,
            float(bill.door_pickup_charge) if bill.door_pickup_charge else 0,
            float(bill.door_delivery_charge) if bill.door_delivery_charge else 0,
            float(bill.hamali_charge) if bill.hamali_charge else 0,
            float(bill.toll_charge) if bill.toll_charge else 0,
            float(bill.fuel_surcharge) if bill.fuel_surcharge else 0,
            float(bill.insurance_charge) if bill.insurance_charge else 0,
            float(bill.other_charges) if bill.other_charges else 0,
            float(bill.discount) if bill.discount else 0,
            float(bill.net_freight) if bill.net_freight else 0,
            # Tax
            'Yes' if bill.gst_applicable else 'No',
            float(bill.igst_rate) if bill.igst_rate else 0,
            float(bill.cgst_rate) if bill.cgst_rate else 0,
            float(bill.sgst_rate) if bill.sgst_rate else 0,
            float(bill.taxable_amount) if bill.taxable_amount else 0,
            float(bill.gst_amount) if bill.gst_amount else 0,
            bill.eway_bill_number or '',
            bill.invoice_number or '',
            # Payment
            bill.freight_payment_mode or '',
            float(bill.advance_amount) if bill.advance_amount else 0,
            float(bill.balance_amount) if bill.balance_amount else 0,
            float(bill.paid_amount) if bill.paid_amount else 0,
            bill.payment_status or '',
            # Transit
            bill.dispatch_datetime.strftime('%Y-%m-%d %H:%M') if bill.dispatch_datetime else '',
            bill.tracking_number or '',
            bill.current_shipment_status or '',
            bill.expected_transit_time or '',
            # Delivery
            bill.delivered_datetime.strftime('%Y-%m-%d %H:%M') if bill.delivered_datetime else '',
            bill.received_by_name or '',
            bill.pod_number or '',
            bill.delivery_remarks or '',
            # Risk
            'Yes' if bill.hazardous_material else 'No',
            'Yes' if bill.temperature_controlled else 'No',
            bill.gps_link or '',
            bill.permit_details or '',
            # Legacy
            bill.challan_number or '',
            bill.party_information or '',
            bill.rate or 0
        ])
    
    return export_to_excel(data, headers, 'bilty_generation.xlsx', dropdowns)


# ==================== IMPORT ROUTES ====================

@app.route("/drivers/import", methods=["POST"])
@permission_required("drivers", "create")
def import_drivers():
    """Import drivers from Excel with all business fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('drivers'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('drivers'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('drivers'))
    
    def parse_date_flexible(value):
        """Parse date from various formats"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    def parse_float_safe(value):
        """Parse float safely handling various formats"""
        if not value:
            return None
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return None
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                # Skip if no first name (required field)
                first_name = safe_str(row.get('First Name'))
                if not first_name:
                    errors.append(f'Row {idx}: First Name is required')
                    continue
                
                driver = Driver(
                    tenant_id=tenant_id,
                    driver_code=generate_driver_code(),
                    # Personal Information
                    first_name=first_name,
                    last_name=safe_str(row.get('Last Name')) or None,
                    father_name=safe_str(row.get('Father Name')) or None,
                    date_of_birth=parse_date_flexible(row.get('Date of Birth')),
                    gender=safe_str(row.get('Gender')) or None,
                    blood_group=safe_str(row.get('Blood Group')) or None,
                    nationality=safe_str(row.get('Nationality')) or None,
                    marital_status=safe_str(row.get('Marital Status')) or None,
                    # Contact Information
                    mobile_number=safe_str(row.get('Mobile Number')) or None,
                    alternate_mobile=safe_str(row.get('Alternate Mobile')) or None,
                    email=safe_str(row.get('Email')) or None,
                    emergency_contact_name=safe_str(row.get('Emergency Contact Name')) or None,
                    emergency_contact_number=safe_str(row.get('Emergency Contact Number')) or None,
                    emergency_contact_relation=safe_str(row.get('Emergency Contact Relation')) or None,
                    # Address Information
                    address_line1=safe_str(row.get('Address Line 1')) or None,
                    address_line2=safe_str(row.get('Address Line 2')) or None,
                    city=safe_str(row.get('City')) or None,
                    state=safe_str(row.get('State')) or None,
                    pincode=safe_str(row.get('Pincode')) or None,
                    country=safe_str(row.get('Country')) or None,
                    # Employment Details
                    employee_code=safe_str(row.get('Employee Code')) or None,
                    date_of_joining=parse_date_flexible(row.get('Date of Joining')),
                    designation=safe_str(row.get('Designation')) or None,
                    department=safe_str(row.get('Department')) or None,
                    employment_type=safe_str(row.get('Employment Type')) or None,
                    status=safe_str(row.get('Status')) or 'Active',
                    monthly_salary=parse_float_safe(row.get('Monthly Salary')),
                    daily_wage=parse_float_safe(row.get('Daily Wage')),
                    # License & Experience
                    license_type=safe_str(row.get('License Type')) or None,
                    license_number=safe_str(row.get('License Number')) or None,
                    license_issue_date=parse_date_flexible(row.get('License Issue Date')),
                    license_expiry_date=parse_date_flexible(row.get('License Expiry Date')),
                    license_issuing_authority=safe_str(row.get('License Issuing Authority')) or None,
                    license_state=safe_str(row.get('License State')) or None,
                    total_experience_years=parse_int(row.get('Total Experience Years')),
                    # Bank Details
                    bank_name=safe_str(row.get('Bank Name')) or None,
                    bank_branch=safe_str(row.get('Bank Branch')) or None,
                    account_holder_name=safe_str(row.get('Account Holder Name')) or None,
                    account_number=safe_str(row.get('Account Number')) or None,
                    ifsc_code=safe_str(row.get('IFSC Code')) or None,
                    account_type=safe_str(row.get('Account Type')) or None,
                    upi_id=safe_str(row.get('UPI ID')) or None,
                    # Documents
                    aadhaar_number=safe_str(row.get('Aadhaar Number')) or None,
                    pan_number=safe_str(row.get('PAN Number')) or None,
                    # References
                    reference1_name=safe_str(row.get('Reference 1 Name')) or None,
                    reference1_contact=safe_str(row.get('Reference 1 Contact')) or None,
                    reference2_name=safe_str(row.get('Reference 2 Name')) or None,
                    reference2_contact=safe_str(row.get('Reference 2 Contact')) or None,
                    # Remarks
                    remarks=safe_str(row.get('Remarks')) or None
                )
                db.session.add(driver)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} drivers with {len(errors)} errors', 'warning')
            for error in errors[:5]:  # Show first 5 errors
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} drivers', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing drivers: {str(e)}', 'error')
    
    return redirect(url_for('drivers'))


@app.route("/vehicles/import", methods=["POST"])
@permission_required("vehicles", "create")
def import_vehicles():
    """Import vehicles from Excel with all business fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('vehicles'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('vehicles'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('vehicles'))
    
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                reg_num = safe_str(row.get('Registration Number'))
                if not reg_num:
                    errors.append(f'Row {idx}: Registration Number is required')
                    continue
                    
                existing = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=reg_num).first()
                if existing:
                    errors.append(f'Row {idx}: Vehicle {reg_num} already exists')
                    continue
                
                vehicle = Vehicle(
                    tenant_id=tenant_id,
                    registration_number=reg_num,
                    # Basic
                    vehicle_type=safe_str(row.get('Vehicle Type')),
                    make=safe_str(row.get('Make')),
                    model=safe_str(row.get('Model')),
                    year=parse_int(row.get('Year')),
                    color=safe_str(row.get('Color')),
                    fuel_type=safe_str(row.get('Fuel Type')),
                    status=safe_str(row.get('Status')) or 'Active',
                    # Technical
                    engine_number=safe_str(row.get('Engine Number')),
                    chassis_number=safe_str(row.get('Chassis Number')),
                    seating_capacity=parse_int(row.get('Seating Capacity')),
                    load_capacity_kg=parse_int(row.get('Load Capacity (KG)')),
                    truck_size=safe_str(row.get('Truck Size')),
                    # Owner
                    owner_name=safe_str(row.get('Owner Name')),
                    owner_contact=safe_str(row.get('Owner Contact')),
                    purchase_date=parse_date_flexible(row.get('Purchase Date')),
                    # Expiry Dates
                    insurance_expiry=parse_date_flexible(row.get('Insurance Expiry')),
                    fitness_expiry=parse_date_flexible(row.get('Fitness Expiry')),
                    permit_1_year_expiry=parse_date_flexible(row.get('Permit 1 Year Expiry')),
                    permit_5_year_expiry=parse_date_flexible(row.get('Permit 5 Year Expiry')),
                    road_tax_expiry=parse_date_flexible(row.get('Road Tax Expiry')),
                    puc_expiry=parse_date_flexible(row.get('PUC Expiry')),
                    # Notes
                    notes=safe_str(row.get('Notes'))
                )
                db.session.add(vehicle)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} vehicles with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} vehicles', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing vehicles: {str(e)}', 'error')
    
    return redirect(url_for('vehicles'))


@app.route("/vendors/import", methods=["POST"])
@permission_required("vendors", "create")
def import_vendors():
    """Import vendors from Excel with all business fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('vendors'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('vendors'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('vendors'))
    
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    def parse_float_safe(value):
        if not value:
            return 0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0
    
    def parse_bool_safe(value):
        if not value:
            return False
        return str(value).strip().lower() in ('yes', 'true', '1', 'y')
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                vendor_name = safe_str(row.get('Vendor Name'))
                if not vendor_name:
                    errors.append(f'Row {idx}: Vendor Name is required')
                    continue
                
                # Check for duplicate GSTIN - WARN but allow (same GST can have multiple addresses)
                gstin = safe_str(row.get('GSTIN'))
                gstin_warning = None
                if gstin:
                    existing_gstin = Vendor.query.filter_by(tenant_id=tenant_id, gstin=gstin).first()
                    if existing_gstin:
                        gstin_warning = f'GSTIN "{gstin}" already used by vendor {existing_gstin.vendor_code} ({existing_gstin.vendor_name}) - Same GST, different address allowed'
                
                vendor = Vendor(
                    tenant_id=tenant_id,
                    vendor_code=generate_vendor_code(),
                    # Basic
                    vendor_name=vendor_name,
                    vendor_type=safe_str(row.get('Vendor Type')) or 'supplier',
                    status=safe_str(row.get('Status')) or 'active',
                    business_nature=safe_str(row.get('Business Nature')) or None,
                    website=safe_str(row.get('Website')) or None,
                    referral_source=safe_str(row.get('Referral Source')) or None,
                    # Contact
                    contact_person=safe_str(row.get('Contact Person')) or None,
                    designation=safe_str(row.get('Designation')) or None,
                    phone_primary=safe_str(row.get('Phone Primary')) or None,
                    phone_secondary=safe_str(row.get('Phone Secondary')) or None,
                    mobile=safe_str(row.get('Mobile')) or None,
                    email=safe_str(row.get('Email')) or None,
                    alternate_email=safe_str(row.get('Alternate Email')) or None,
                    # Registered Address
                    reg_address_line1=safe_str(row.get('Reg Address Line 1')) or None,
                    reg_address_line2=safe_str(row.get('Reg Address Line 2')) or None,
                    reg_city=safe_str(row.get('Reg City')) or None,
                    reg_state=safe_str(row.get('Reg State')) or None,
                    reg_pincode=safe_str(row.get('Reg Pincode')) or None,
                    reg_country=safe_str(row.get('Reg Country')) or 'India',
                    # Office Address
                    office_address_line1=safe_str(row.get('Office Address Line 1')) or None,
                    office_address_line2=safe_str(row.get('Office Address Line 2')) or None,
                    office_city=safe_str(row.get('Office City')) or None,
                    office_state=safe_str(row.get('Office State')) or None,
                    office_pincode=safe_str(row.get('Office Pincode')) or None,
                    office_country=safe_str(row.get('Office Country')) or 'India',
                    # GST & Tax
                    gstin=safe_str(row.get('GSTIN')) or None,
                    gst_registration_date=parse_date_flexible(row.get('GST Registration Date')),
                    gst_state_code=safe_str(row.get('GST State Code')) or None,
                    pan_number=safe_str(row.get('PAN Number')) or None,
                    tan_number=safe_str(row.get('TAN Number')) or None,
                    tax_regime=safe_str(row.get('Tax Regime')) or 'regular',
                    is_composition_dealer=parse_bool_safe(row.get('Is Composition Dealer')),
                    # Bank
                    bank_name=safe_str(row.get('Bank Name')) or None,
                    bank_branch=safe_str(row.get('Bank Branch')) or None,
                    account_number=safe_str(row.get('Account Number')) or None,
                    account_type=safe_str(row.get('Account Type')) or None,
                    ifsc_code=safe_str(row.get('IFSC Code')) or None,
                    micr_code=safe_str(row.get('MICR Code')) or None,
                    upi_id=safe_str(row.get('UPI ID')) or None,
                    # Primary Contact
                    primary_contact_name=safe_str(row.get('Primary Contact Name')) or None,
                    primary_contact_designation=safe_str(row.get('Primary Contact Designation')) or None,
                    primary_contact_mobile=safe_str(row.get('Primary Contact Mobile')) or None,
                    primary_contact_email=safe_str(row.get('Primary Contact Email')) or None,
                    # Trade References
                    trade_reference_1_name=safe_str(row.get('Trade Reference 1 Name')) or None,
                    trade_reference_1_contact=safe_str(row.get('Trade Reference 1 Contact')) or None,
                    trade_reference_2_name=safe_str(row.get('Trade Reference 2 Name')) or None,
                    trade_reference_2_contact=safe_str(row.get('Trade Reference 2 Contact')) or None,
                    # Financial
                    credit_limit=parse_float_safe(row.get('Credit Limit')),
                    credit_period_days=parse_int(row.get('Credit Period Days')),
                    opening_balance=parse_float_safe(row.get('Opening Balance')),
                    currency=safe_str(row.get('Currency')) or 'INR',
                    payment_terms=safe_str(row.get('Payment Terms')) or None,
                    # Compliance
                    kyc_status=safe_str(row.get('KYC Status')) or 'pending',
                    compliance_rating=safe_str(row.get('Compliance Rating')) or 'unrated',
                    # Supply
                    supply_type=safe_str(row.get('Supply Type')) or None,
                    lead_time_days=parse_int(row.get('Lead Time Days')),
                    # Remarks
                    remarks=safe_str(row.get('Remarks')) or None
                )
                db.session.add(vendor)
                imported_count += 1
                
                # Add GSTIN warning if applicable
                if gstin_warning:
                    warnings_list.append(f'Row {idx}: {gstin_warning}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        
        # Show import results
        if warnings_list and not errors:
            flash(f'Successfully imported {imported_count} vendors with {len(warnings_list)} warnings', 'warning')
            for warning in warnings_list[:5]:
                flash(f'Warning: {warning}', 'warning')
        elif errors:
            flash(f'Imported {imported_count} vendors with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
            if warnings_list:
                for warning in warnings_list[:3]:
                    flash(f'Warning: {warning}', 'warning')
        else:
            flash(f'Successfully imported {imported_count} vendors', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing vendors: {str(e)}', 'error')
    
    return redirect(url_for('vendors'))


@app.route("/expenses/import", methods=["POST"])
@permission_required("expenses", "create")
def import_expenses():
    """Import expenses from Excel with all business fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('expenses'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('expenses'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('expenses'))
    
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    def parse_float_safe(value):
        if not value:
            return 0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                # Find related records by name/number
                vehicle_number = row.get('Vehicle Number', '').strip()
                vehicle = None
                if vehicle_number:
                    vehicle = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=vehicle_number).first()
                
                location_name = row.get('Location', '').strip()
                location = None
                if location_name:
                    location = Location.query.filter_by(tenant_id=tenant_id, location=location_name).first()
                
                expense = Expense(
                    tenant_id=tenant_id,
                    name=generate_expense_name(),
                    expense_date=parse_date_flexible(row.get('Expense Date')),
                    category=row.get('Category', '').strip() or None,
                    description=row.get('Description', '').strip() or None,
                    amount=parse_float_safe(row.get('Amount')),
                    vehicle_id=vehicle.id if vehicle else None,
                    location_id=location.id if location else None,
                    payment_method=row.get('Payment Method', '').strip() or None,
                    vendor_name=row.get('Vendor Name', '').strip() or None,
                    vendor_contact=row.get('Vendor Contact', '').strip() or None,
                    bill_number=row.get('Bill Number', '').strip() or None,
                    status=row.get('Status', 'Pending').strip() or 'Pending',
                    notes=row.get('Notes', '').strip() or None
                )
                db.session.add(expense)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} expenses with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} expenses', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing expenses: {str(e)}', 'error')
    
    return redirect(url_for('expenses'))


@app.route("/loans/import", methods=["POST"])
@permission_required("loans", "create")
def import_loans():
    """Import loans from Excel with all business fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('loans'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('loans'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('loans'))
    
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    def parse_float_safe(value):
        if not value:
            return 0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                # Find vehicle by number
                vehicle_number = row.get('Vehicle Number', '').strip()
                vehicle = None
                if vehicle_number:
                    vehicle = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=vehicle_number).first()
                
                loan = Loan(
                    tenant_id=tenant_id,
                    name=generate_loan_name(),
                    # Basic
                    loan_type=row.get('Loan Type', '').strip() or None,
                    vehicle_id=vehicle.id if vehicle else None,
                    principal_amount=parse_float_safe(row.get('Principal Amount')),
                    interest_rate=parse_float_safe(row.get('Interest Rate')),
                    tenure_months=parse_int(row.get('Tenure Months')),
                    # EMI
                    emi_amount=parse_float_safe(row.get('EMI Amount')),
                    total_payable=parse_float_safe(row.get('Total Payable')),
                    total_interest=parse_float_safe(row.get('Total Interest')),
                    down_payment=parse_float_safe(row.get('Down Payment')),
                    amount_paid=parse_float_safe(row.get('Amount Paid')),
                    balance_amount=parse_float_safe(row.get('Balance Amount')),
                    emis_paid=parse_int(row.get('EMIs Paid')),
                    emis_remaining=parse_int(row.get('EMIs Remaining')),
                    # Lender
                    lender_name=row.get('Lender Name', '').strip() or None,
                    lender_type=row.get('Lender Type', '').strip() or None,
                    lender_contact=row.get('Lender Contact', '').strip() or None,
                    lender_address=row.get('Lender Address', '').strip() or None,
                    # Agent
                    agent_name=row.get('Agent Name', '').strip() or None,
                    agent_contact=row.get('Agent Contact', '').strip() or None,
                    # Dates
                    loan_date=parse_date_flexible(row.get('Loan Date')),
                    disbursement_date=parse_date_flexible(row.get('Disbursement Date')),
                    first_emi_date=parse_date_flexible(row.get('First EMI Date')),
                    last_emi_date=parse_date_flexible(row.get('Last EMI Date')),
                    next_emi_due_date=parse_date_flexible(row.get('Next EMI Due Date')),
                    # Account
                    loan_account_number=row.get('Loan Account Number', '').strip() or None,
                    status=safe_str(row.get('Status')) or 'Active',
                    purpose=row.get('Purpose', '').strip() or None,
                    collateral=row.get('Collateral', '').strip() or None,
                    insurance_details=row.get('Insurance Details', '').strip() or None,
                    remarks=safe_str(row.get('Remarks')) or None
                )
                db.session.add(loan)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} loans with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} loans', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing loans: {str(e)}', 'error')
    
    return redirect(url_for('loans'))


@app.route("/locations/import", methods=["POST"])
@permission_required("locations", "create")
def import_locations():
    """Import rate list from Excel with all fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('locations'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('locations'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('locations'))
    
    def parse_float_safe(value):
        if not value:
            return 0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                loc_name = row.get('Location Name', '').strip()
                if not loc_name:
                    errors.append(f'Row {idx}: Location Name is required')
                    continue
                    
                existing = Location.query.filter_by(tenant_id=tenant_id, location=loc_name).first()
                if existing:
                    errors.append(f'Row {idx}: Location {loc_name} already exists')
                    continue
                
                location = Location(
                    tenant_id=tenant_id,
                    location=loc_name,
                    city=safe_str(row.get('City')) or None,
                    state=safe_str(row.get('State')) or None,
                    pincode=safe_str(row.get('Pincode')) or None,
                    rate=parse_float_safe(row.get('Rate')),
                    distance_km=parse_float_safe(row.get('Distance (KM)')),
                    remarks=safe_str(row.get('Remarks')) or None
                )
                db.session.add(location)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} rate entries with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} rate entries', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing rate list: {str(e)}', 'error')
    
    return redirect(url_for('locations'))


@app.route("/transport-bills/import", methods=["POST"])
@permission_required("transport_bills", "create")
def import_transport_bills():
    """Import transport bills from Excel with comprehensive fields"""
    if 'file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('transport_bills'))
    
    file = request.files['file']
    if not file.filename:
        flash('No file selected', 'error')
        return redirect(url_for('transport_bills'))
    
    result = parse_excel_to_json(file)
    
    if not result['success']:
        flash(f'Error parsing file: {result["error"]}', 'error')
        return redirect(url_for('transport_bills'))
    
    def parse_date_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except:
                continue
        return None
    
    def parse_datetime_flexible(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        for fmt in ['%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M', '%Y-%m-%d', '%d/%m/%Y']:
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except:
                continue
        return None
    
    def parse_float_safe(value):
        if not value:
            return 0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0
    
    def parse_bool_safe(value):
        if not value:
            return False
        return str(value).strip().lower() in ('yes', 'true', '1', 'y')
    
    try:
        tenant_id = get_current_tenant_id()
        imported_count = 0
        errors = []
        
        for idx, row in enumerate(result['data'], 1):
            try:
                bill = TransportBill(
                    tenant_id=tenant_id,
                    name=generate_bill_name(),
                    # GR Basic
                    gr_number=row.get('GR Number', '').strip() or None,
                    gr_date=parse_date_flexible(row.get('GR Date')),
                    gr_type=row.get('GR Type', '').strip() or None,
                    booking_mode=row.get('Booking Mode', '').strip() or 'road',
                    status=row.get('Status', 'draft').strip() or 'draft',
                    reference_number=row.get('Reference Number', '').strip() or None,
                    # Consignor
                    consignor_name=row.get('Consignor Name', '').strip() or None,
                    consignor_gstin=row.get('Consignor GSTIN', '').strip() or None,
                    consignor_contact_person=row.get('Consignor Contact', '').strip() or None,
                    consignor_mobile=row.get('Consignor Mobile', '').strip() or None,
                    consignor_email=row.get('Consignor Email', '').strip() or None,
                    # Consignee
                    consignee_name=row.get('Consignee Name', '').strip() or None,
                    consignee_gstin=row.get('Consignee GSTIN', '').strip() or None,
                    consignee_contact_person=row.get('Consignee Contact', '').strip() or None,
                    consignee_mobile=row.get('Consignee Mobile', '').strip() or None,
                    consignee_email=row.get('Consignee Email', '').strip() or None,
                    bill_to_party=row.get('Bill To Party', '').strip() or None,
                    ship_to_party=row.get('Ship To Party', '').strip() or None,
                    # Pickup
                    pickup_point=row.get('Pickup Point', '').strip() or None,
                    pickup_date=parse_date_flexible(row.get('Pickup Date')),
                    pickup_city=row.get('Pickup City', '').strip() or None,
                    pickup_state=row.get('Pickup State', '').strip() or None,
                    loaded_by=row.get('Loaded By', '').strip() or None,
                    loading_type=row.get('Loading Type', '').strip() or None,
                    # Delivery
                    delivery_location=row.get('Delivery Location', '').strip() or None,
                    expected_delivery_date=parse_date_flexible(row.get('Expected Delivery Date')),
                    actual_delivery_date=parse_date_flexible(row.get('Actual Delivery Date')),
                    delivery_city=row.get('Delivery City', '').strip() or None,
                    delivery_state=row.get('Delivery State', '').strip() or None,
                    unloading_point=row.get('Unloading Point', '').strip() or None,
                    # Vehicle
                    vehicle_number=row.get('Vehicle Number', '').strip() or None,
                    vehicle_type=row.get('Vehicle Type', '').strip() or None,
                    vehicle_capacity=row.get('Vehicle Capacity', '').strip() or None,
                    vehicle_ownership=row.get('Vehicle Ownership', '').strip() or 'owned',
                    driver_name=row.get('Driver Name', '').strip() or None,
                    driver_mobile=row.get('Driver Mobile', '').strip() or None,
                    driver_license=row.get('Driver License', '').strip() or None,
                    helper_name=row.get('Helper Name', '').strip() or None,
                    # Route
                    origin=row.get('Origin', '').strip() or None,
                    destination=row.get('Destination', '').strip() or None,
                    via_route=row.get('Via Route', '').strip() or None,
                    # Material
                    article_description=row.get('Article Description', '').strip() or None,
                    package_type=row.get('Package Type', '').strip() or None,
                    number_of_packages=parse_int(row.get('Number of Packages')),
                    packing_mode=row.get('Packing Mode', '').strip() or None,
                    actual_weight=parse_float_safe(row.get('Actual Weight')),
                    charged_weight=parse_float_safe(row.get('Charged Weight')),
                    volume_cbm=parse_float_safe(row.get('Volume CBM')),
                    material_value=parse_float_safe(row.get('Material Value')),
                    dimensions=row.get('Dimensions', '').strip() or None,
                    goods_category=row.get('Goods Category', '').strip() or None,
                    hsn_code=row.get('HSN Code', '').strip() or None,
                    declared_value=parse_float_safe(row.get('Declared Value')),
                    handling_instructions=row.get('Handling Instructions', '').strip() or None,
                    # Freight
                    basic_freight=parse_float_safe(row.get('Basic Freight')),
                    loading_charge=parse_float_safe(row.get('Loading Charge')),
                    unloading_charge=parse_float_safe(row.get('Unloading Charge')),
                    door_pickup_charge=parse_float_safe(row.get('Door Pickup Charge')),
                    door_delivery_charge=parse_float_safe(row.get('Door Delivery Charge')),
                    hamali_charge=parse_float_safe(row.get('Hamali Charge')),
                    toll_charge=parse_float_safe(row.get('Toll Charge')),
                    fuel_surcharge=parse_float_safe(row.get('Fuel Surcharge')),
                    insurance_charge=parse_float_safe(row.get('Insurance Charge')),
                    other_charges=parse_float_safe(row.get('Other Charges')),
                    discount=parse_float_safe(row.get('Discount')),
                    net_freight=parse_float_safe(row.get('Net Freight')),
                    # Tax
                    gst_applicable=parse_bool_safe(row.get('GST Applicable')),
                    igst_rate=parse_float_safe(row.get('IGST Rate')),
                    cgst_rate=parse_float_safe(row.get('CGST Rate')),
                    sgst_rate=parse_float_safe(row.get('SGST Rate')),
                    taxable_amount=parse_float_safe(row.get('Taxable Amount')),
                    gst_amount=parse_float_safe(row.get('GST Amount')),
                    eway_bill_number=row.get('Eway Bill Number', '').strip() or None,
                    invoice_number=row.get('Invoice Number', '').strip() or None,
                    # Payment
                    freight_payment_mode=row.get('Freight Payment Mode', '').strip() or 'to_pay',
                    advance_amount=parse_float_safe(row.get('Advance Amount')),
                    balance_amount=parse_float_safe(row.get('Balance Amount')),
                    paid_amount=parse_float_safe(row.get('Paid Amount')),
                    payment_status=row.get('Payment Status', '').strip() or 'pending',
                    # Transit
                    dispatch_datetime=parse_datetime_flexible(row.get('Dispatch DateTime')),
                    tracking_number=row.get('Tracking Number', '').strip() or None,
                    current_shipment_status=row.get('Current Shipment Status', '').strip() or None,
                    expected_transit_time=row.get('Expected Transit Time', '').strip() or None,
                    # Delivery
                    delivered_datetime=parse_datetime_flexible(row.get('Delivered DateTime')),
                    received_by_name=row.get('Received By Name', '').strip() or None,
                    pod_number=row.get('POD Number', '').strip() or None,
                    delivery_remarks=row.get('Delivery Remarks', '').strip() or None,
                    # Risk
                    hazardous_material=parse_bool_safe(row.get('Hazardous Material')),
                    temperature_controlled=parse_bool_safe(row.get('Temperature Controlled')),
                    gps_link=row.get('GPS Link', '').strip() or None,
                    permit_details=row.get('Permit Details', '').strip() or None,
                    # Legacy
                    challan_number=row.get('Challan Number', '').strip() or None,
                    party_information=row.get('Party Information', '').strip() or None,
                    rate=parse_int(row.get('Rate'))
                )
                db.session.add(bill)
                imported_count += 1
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
                continue
        
        db.session.commit()
        if errors:
            flash(f'Imported {imported_count} bills with {len(errors)} errors', 'warning')
            for error in errors[:5]:
                flash(error, 'error')
        else:
            flash(f'Successfully imported {imported_count} bills', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing bills: {str(e)}', 'error')
    
    return redirect(url_for('transport_bills'))


@app.route("/dispatch")
@permission_required("dispatch", "view")
def dispatch_trips():
    trips = scoped_query(DispatchTrip).order_by(DispatchTrip.created_at.desc()).all()
    changed = False
    for trip in trips:
        changed = ensure_trip_tokens(trip) or changed
    if changed:
        db.session.commit()
    return render_template(
        "dispatch/list.html",
        trips=trips,
        show_tenant_column=is_superadmin(),
    )


@app.route("/dispatch/create", methods=["GET", "POST"])
@permission_required("dispatch", "create")
def create_dispatch_trip():
    selected_tenant_id = get_default_selected_tenant_id()
    bills = (
        scoped_query(TransportBill)
        .filter_by(tenant_id=selected_tenant_id)
        .filter(TransportBill.status.notin_(["delivered", "closed"]))
        .order_by(TransportBill.created_at.desc())
        .all()
    )
    vehicles = get_tenant_filtered_records(Vehicle, Vehicle.registration_number, selected_tenant_id)
    drivers = get_tenant_filtered_records(Driver, Driver.first_name, selected_tenant_id)

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
            bill = get_related_record(
                TransportBill, parse_int(request.form.get("bilty_id")), tenant_id, "Bilty", True
            )
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle", True
            )
            driver = get_related_record(
                Driver, parse_int(request.form.get("driver_id")), tenant_id, "Driver"
            )

            trip = DispatchTrip(
                tenant_id=tenant_id,
                trip_number=generate_trip_number(),
                trip_date=parse_date(request.form.get("trip_date"), date.today()),
                status="planned",
                bilty_id=bill.id,
                vehicle_id=vehicle.id,
                driver_id=driver.id if driver else None,
                origin=request.form.get("origin", "").strip() or None,
                destination=request.form.get("destination", "").strip() or None,
                planned_dispatch_at=parse_datetime_local(request.form.get("planned_dispatch_at")),
                expected_delivery_at=parse_datetime_local(request.form.get("expected_delivery_at")),
                dispatch_notes=request.form.get("dispatch_notes", "").strip() or None,
                created_by=g.current_user.id if g.current_user else None,
                customer_tracking_token=generate_access_token(),
                driver_access_token=generate_access_token(),
            )
            bill.status = "planned"
            bill.vehicle_id = vehicle.id
            if driver:
                bill.driver_name = f"{driver.first_name} {driver.last_name or ''}".strip()
                bill.driver_mobile = driver.mobile_number
                bill.driver_license = driver.license_number

            db.session.add(trip)
            db.session.flush()
            record_audit(
                "trip_planned",
                trip,
                summary=f"Trip {trip.trip_number} planned for bilty {bill.name}",
                details={"bilty": bill.name, "vehicle": vehicle.registration_number},
                tenant_id=tenant_id,
            )
            db.session.commit()
            flash(f"Dispatch trip {trip.trip_number} planned successfully.", "success")
            return redirect(url_for("dispatch_trips"))
        except ValueError as exc:
            flash(str(exc), "error")
        except Exception as exc:
            db.session.rollback()
            flash(f"Error creating dispatch trip: {str(exc)}", "error")

    context = get_form_tenant_context()
    context.update(
        {
            "bills": bills,
            "vehicles": vehicles,
            "drivers": drivers,
            "today": date.today().isoformat(),
        }
    )
    return render_template("dispatch/form.html", **context)


@app.route("/dispatch/<int:id>/update", methods=["POST"])
@permission_required("dispatch", "edit")
def update_dispatch_trip(id):
    trip = get_scoped_record(DispatchTrip, id)
    action = request.form.get("action")

    status_map = {
        "dispatch": ("dispatched", "dispatched"),
        "in_transit": ("in_transit", "in_transit"),
        "delay": ("delayed", "in_transit"),
        "close": ("closed", "delivered"),
    }
    if action not in status_map:
        flash("Invalid dispatch action.", "error")
        return redirect(url_for("dispatch_trips"))

    trip_status, bill_status = status_map[action]
    now = datetime.utcnow()
    trip.status = trip_status
    trip.current_location = request.form.get("current_location", "").strip() or trip.current_location
    trip.delay_reason = request.form.get("delay_reason", "").strip() or trip.delay_reason
    trip.last_tracking_update_at = now

    if action == "dispatch" and not trip.actual_dispatch_at:
        trip.actual_dispatch_at = now
    if action == "close":
        trip.actual_delivery_at = now

    if trip.bilty:
        trip.bilty.status = bill_status

    record_audit(
        f"trip_{action}",
        trip,
        summary=f"Trip {trip.trip_number} updated to {trip.status}",
        details={
            "status": trip.status,
            "current_location": trip.current_location,
            "delay_reason": trip.delay_reason,
        },
    )
    db.session.commit()
    flash(f"Trip {trip.trip_number} updated to {trip.status.replace('_', ' ').title()}.", "success")
    return redirect(url_for("dispatch_trips"))


@app.route("/track/<token>")
def public_track_trip(token):
    trip = DispatchTrip.query.filter_by(customer_tracking_token=token).first_or_404()
    return render_template("public/track.html", trip=trip)


@app.route("/driver/trip/<token>", methods=["GET", "POST"])
def driver_trip_update(token):
    trip = DispatchTrip.query.filter_by(driver_access_token=token).first_or_404()
    if request.method == "POST":
        action = request.form.get("action")
        now = datetime.utcnow()

        if action == "tracking":
            trip.status = "in_transit"
            trip.current_location = request.form.get("current_location", "").strip() or trip.current_location
            trip.last_tracking_update_at = now
            if trip.bilty:
                trip.bilty.status = "in_transit"
            record_audit(
                "driver_tracking_update",
                trip,
                summary=f"Driver updated location for {trip.trip_number}",
                details={"current_location": trip.current_location},
            )
            flash("Location updated.", "success")
        elif action == "pod":
            pod_file = request.files.get("pod_file")
            pod_attachment_path = None
            if pod_file and pod_file.filename:
                pod_attachment_path = save_uploaded_file(
                    pod_file,
                    subfolder="pod",
                    prefix=f"driver_pod_{trip.trip_number}_{now.strftime('%Y%m%d%H%M%S')}",
                )
            trip.status = "delivered"
            trip.actual_delivery_at = now
            trip.pod_number = request.form.get("pod_number", "").strip() or trip.pod_number
            trip.received_by_name = request.form.get("received_by_name", "").strip() or trip.received_by_name
            trip.delivered_by = request.form.get("delivered_by", "").strip() or trip.delivered_by
            trip.delivery_remarks = request.form.get("delivery_remarks", "").strip() or trip.delivery_remarks
            trip.pod_attachment_path = pod_attachment_path or trip.pod_attachment_path
            trip.last_tracking_update_at = now
            if trip.bilty:
                trip.bilty.status = "delivered"
                extended_data = {}
                if trip.bilty.extended_data:
                    try:
                        extended_data = json.loads(trip.bilty.extended_data)
                    except Exception:
                        extended_data = {}
                extended_data["delivery_confirmation"] = {
                    "delivered_by": trip.delivered_by,
                    "received_by_name": trip.received_by_name,
                    "pod_number": trip.pod_number,
                    "delivered_datetime": now.isoformat(),
                    "remarks": trip.delivery_remarks,
                    "pod_attachment_path": trip.pod_attachment_path,
                    "source": "driver_mobile",
                }
                trip.bilty.extended_data = json.dumps(extended_data)
            record_audit(
                "driver_pod_uploaded",
                trip,
                summary=f"Driver uploaded POD for {trip.trip_number}",
                details={"pod_number": trip.pod_number, "pod_attachment_path": trip.pod_attachment_path},
            )
            flash("POD submitted and delivery marked complete.", "success")
        else:
            flash("Invalid update.", "error")

        db.session.commit()
        return redirect(url_for("driver_trip_update", token=token))

    return render_template("public/driver_trip.html", trip=trip)


@app.route("/transport-bills")
@permission_required("transport_bills", "view")
def transport_bills():
    bills = scoped_query(TransportBill).order_by(TransportBill.created_at.desc()).all()
    return render_template(
        "transport_bills/list.html",
        bills=bills,
        show_tenant_column=is_superadmin(),
    )


@app.route("/transport-bills/create", methods=["GET", "POST"])
@permission_required("transport_bills", "create")
def create_transport_bill():
    selected_tenant_id = get_default_selected_tenant_id()
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    locations = get_tenant_filtered_records(Location, Location.location, selected_tenant_id)
    delivery_types = get_tenant_filtered_records(
        DeliveryType, DeliveryType.delivery_type, selected_tenant_id
    )
    vendors = get_tenant_filtered_records(
        Vendor, Vendor.vendor_name, selected_tenant_id
    )
    drivers = get_tenant_filtered_records(
        Driver, Driver.first_name, selected_tenant_id
    )
    ratelists = scoped_query(RateList).filter_by(
        tenant_id=selected_tenant_id,
        is_active=True
    ).order_by(RateList.effective_date.desc(), RateList.name).all()

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle", True
            )
            location = get_related_record(
                Location, parse_int(request.form.get("location_id")), tenant_id, "Location"
            )
            delivery_type = get_related_record(
                DeliveryType,
                parse_int(request.form.get("delivery_type_id")),
                tenant_id,
                "Delivery Type",
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context()
            context.update(
                {
                    "vehicles": vehicles,
                    "locations": locations,
                    "delivery_types": delivery_types,
                    "vendors": vendors,
                    "today": date.today().isoformat(),
                }
            )
            return render_template("transport_bills/form.html", **context)

        bill = TransportBill(
            tenant_id=tenant_id,
            name=generate_bill_name(),
            vehicle_id=vehicle.id,
            # GR Basic Details
            gr_number=request.form.get("gr_number", "").strip() or None,
            gr_date=parse_date(request.form.get("gr_date")),
            branch_booking_location=request.form.get("branch_booking_location", "").strip() or None,
            gr_type=request.form.get("gr_type", "").strip() or None,
            booking_mode=request.form.get("booking_mode", "road").strip() or "road",
            status=request.form.get("status", "draft").strip() or "draft",
            reference_number=request.form.get("reference_number", "").strip() or None,
            # Party Details
            consignor_name=request.form.get("consignor_name", "").strip() or None,
            consignor_address=request.form.get("consignor_address", "").strip() or None,
            consignor_gstin=request.form.get("consignor_gstin", "").strip() or None,
            consignor_contact_person=request.form.get("consignor_contact_person", "").strip() or None,
            consignor_mobile=request.form.get("consignor_mobile", "").strip() or None,
            consignor_email=request.form.get("consignor_email", "").strip() or None,
            consignee_name=request.form.get("consignee_name", "").strip() or None,
            consignee_address=request.form.get("consignee_address", "").strip() or None,
            consignee_gstin=request.form.get("consignee_gstin", "").strip() or None,
            consignee_contact_person=request.form.get("consignee_contact_person", "").strip() or None,
            consignee_mobile=request.form.get("consignee_mobile", "").strip() or None,
            consignee_email=request.form.get("consignee_email", "").strip() or None,
            bill_to_party=request.form.get("bill_to_party", "").strip() or None,
            ship_to_party=request.form.get("ship_to_party", "").strip() or None,
            # Pickup Details
            pickup_point=request.form.get("pickup_point", "").strip() or None,
            pickup_date=parse_date(request.form.get("pickup_date")),
            pickup_branch=request.form.get("pickup_branch", "").strip() or None,
            loaded_by=request.form.get("loaded_by", "").strip() or None,
            loading_type=request.form.get("loading_type", "").strip() or None,
            pickup_remarks=request.form.get("pickup_remarks", "").strip() or None,
            pickup_city=request.form.get("pickup_city", "").strip() or None,
            pickup_state=request.form.get("pickup_state", "").strip() or None,
            pickup_pincode=request.form.get("pickup_pincode", "").strip() or None,
            # Delivery Details
            delivery_location=request.form.get("delivery_location", "").strip() or None,
            expected_delivery_date=parse_date(request.form.get("expected_delivery_date")),
            delivery_type_id=delivery_type.id if delivery_type else None,
            delivery_city=request.form.get("delivery_city", "").strip() or None,
            delivery_state=request.form.get("delivery_state", "").strip() or None,
            delivery_pincode=request.form.get("delivery_pincode", "").strip() or None,
            unloading_point=request.form.get("unloading_point", "").strip() or None,
            delivery_remarks=request.form.get("delivery_remarks", "").strip() or None,
            # Vehicle Details
            vehicle_number=request.form.get("vehicle_number", "").strip() or None,
            vehicle_type=request.form.get("vehicle_type", "").strip() or None,
            vehicle_capacity=request.form.get("vehicle_capacity", "").strip() or None,
            vehicle_ownership=request.form.get("vehicle_ownership", "owned").strip() or "owned",
            driver_name=request.form.get("driver_name", "").strip() or None,
            driver_mobile=request.form.get("driver_mobile", "").strip() or None,
            driver_license=request.form.get("driver_license", "").strip() or None,
            helper_name=request.form.get("helper_name", "").strip() or None,
            trip_number=request.form.get("trip_number", "").strip() or None,
            route_name=request.form.get("route_name", "").strip() or None,
            # Material Details
            article_description=request.form.get("article_description", "").strip() or None,
            package_type=request.form.get("package_type", "").strip() or None,
            number_of_packages=parse_int(request.form.get("number_of_packages")) or 0,
            packing_mode=request.form.get("packing_mode", "").strip() or None,
            actual_weight=parse_float(request.form.get("actual_weight")),
            charged_weight=parse_float(request.form.get("charged_weight")),
            volume_cbm=parse_float(request.form.get("volume_cbm")),
            material_value=parse_float(request.form.get("material_value")),
            dimensions=request.form.get("dimensions", "").strip() or None,
            goods_category=request.form.get("goods_category", "").strip() or None,
            nature_of_goods=request.form.get("nature_of_goods", "").strip() or None,
            hsn_code=request.form.get("hsn_code", "").strip() or None,
            declared_value=parse_float(request.form.get("declared_value")),
            insurance_required=parse_bool(request.form.get("insurance_required")),
            insurance_value=parse_float(request.form.get("insurance_value")),
            handling_instructions=request.form.get("handling_instructions", "").strip() or None,
            goods_remarks=request.form.get("goods_remarks", "").strip() or None,
            # Freight Charges
            basic_freight=parse_float(request.form.get("basic_freight")),
            loading_charge=parse_float(request.form.get("loading_charge")),
            unloading_charge=parse_float(request.form.get("unloading_charge")),
            door_pickup_charge=parse_float(request.form.get("door_pickup_charge")),
            door_delivery_charge=parse_float(request.form.get("door_delivery_charge")),
            hamali_charge=parse_float(request.form.get("hamali_charge")),
            detention_charge=parse_float(request.form.get("detention_charge")),
            waiting_charge=parse_float(request.form.get("waiting_charge")),
            halting_charge=parse_float(request.form.get("halting_charge")),
            toll_charge=parse_float(request.form.get("toll_charge")),
            border_charge=parse_float(request.form.get("border_charge")),
            fuel_surcharge=parse_float(request.form.get("fuel_surcharge")),
            packing_charge=parse_float(request.form.get("packing_charge")),
            weighment_charge=parse_float(request.form.get("weighment_charge")),
            permit_charge=parse_float(request.form.get("permit_charge")),
            driver_allowance=parse_float(request.form.get("driver_allowance")),
            insurance_charge=parse_float(request.form.get("insurance_charge")),
            other_charges=parse_float(request.form.get("other_charges")),
            discount=parse_float(request.form.get("discount")),
            net_freight=parse_float(request.form.get("net_freight")),
            # Tax Details
            gst_applicable=parse_bool(request.form.get("gst_applicable")),
            igst_rate=parse_float(request.form.get("igst_rate")),
            cgst_rate=parse_float(request.form.get("cgst_rate")),
            sgst_rate=parse_float(request.form.get("sgst_rate")),
            taxable_amount=parse_float(request.form.get("taxable_amount")),
            gst_amount=parse_float(request.form.get("gst_amount")),
            eway_bill_number=request.form.get("eway_bill_number", "").strip() or None,
            eway_bill_date=parse_date(request.form.get("eway_bill_date")),
            eway_bill_validity=parse_date(request.form.get("eway_bill_validity")),
            invoice_number=request.form.get("invoice_number", "").strip() or None,
            # Payment Details
            freight_payment_mode=request.form.get("freight_payment_mode", "to_pay").strip() or "to_pay",
            freight_collected_from=request.form.get("freight_collected_from", "").strip() or None,
            advance_amount=parse_float(request.form.get("advance_amount")),
            balance_amount=parse_float(request.form.get("balance_amount")),
            freight_due=parse_float(request.form.get("freight_due")),
            paid_amount=parse_float(request.form.get("paid_amount")),
            payment_status=request.form.get("payment_status", "pending").strip() or "pending",
            receipt_number=request.form.get("receipt_number", "").strip() or None,
            # Transit Details
            dispatch_from_branch=request.form.get("dispatch_from_branch", "").strip() or None,
            reached_hub=request.form.get("reached_hub", "").strip() or None,
            transshipment_required=parse_bool(request.form.get("transshipment_required")),
            transshipment_hub=request.form.get("transshipment_hub", "").strip() or None,
            intermediate_vehicle_number=request.form.get("intermediate_vehicle_number", "").strip() or None,
            expected_transit_time=request.form.get("expected_transit_time", "").strip() or None,
            actual_transit_time=request.form.get("actual_transit_time", "").strip() or None,
            delay_reason=request.form.get("delay_reason", "").strip() or None,
            tracking_number=request.form.get("tracking_number", "").strip() or None,
            # Delivery Confirmation
            delivered_by=request.form.get("delivered_by", "").strip() or None,
            received_by_name=request.form.get("received_by_name", "").strip() or None,
            pod_number=request.form.get("pod_number", "").strip() or None,
            seal_number=request.form.get("seal_number", "").strip() or None,
            damage_shortage_report=request.form.get("damage_shortage_report", "").strip() or None,
            closed_by=request.form.get("closed_by", "").strip() or None,
            # Risk/Compliance
            hazardous_material=parse_bool(request.form.get("hazardous_material")),
            temperature_controlled=parse_bool(request.form.get("temperature_controlled")),
            shortage_damage_flag=parse_bool(request.form.get("shortage_damage_flag")),
            customer_complaint_number=request.form.get("customer_complaint_number", "").strip() or None,
            claim_amount=parse_float(request.form.get("claim_amount")),
            gps_link=request.form.get("gps_link", "").strip() or None,
            permit_details=request.form.get("permit_details", "").strip() or None,
            # Attachments
            bilty_attachment_1_path=request.form.get("bilty_attachment_1_path", "").strip() or None,
            bilty_attachment_2_path=request.form.get("bilty_attachment_2_path", "").strip() or None,
            # Legacy fields
            challan_number=request.form.get("challan_number", "").strip() or None,
            party_information=request.form.get("party_information", "").strip() or None,
            location_id=location.id if location else None,
            rate=parse_int(request.form.get("rate")) or 0,
            created_by=getattr(g, "current_user", None).username if getattr(g, "current_user", None) else None,
        )
        db.session.add(bill)
        db.session.flush()
        record_audit(
            "bilty_created",
            bill,
            summary=f"Bilty {bill.name} created",
            details={"vehicle_id": bill.vehicle_id, "rate": bill.rate, "status": bill.status},
        )
        db.session.commit()
        flash(f"Bilty {bill.name} created successfully.", "success")
        return redirect(url_for("transport_bills"))

    context = get_form_tenant_context()
    context.update(
        {
            "vehicles": vehicles,
            "locations": locations,
            "delivery_types": delivery_types,
            "vendors": vendors,
            "drivers": drivers,
            "ratelists": ratelists,
            "today": date.today().isoformat(),
        }
    )
    return render_template("transport_bills/form.html", **context)


@app.route("/payment-receipts/create", methods=["GET", "POST"])
@permission_required("transport_bills", "edit")
def create_payment_receipt():
    selected_tenant_id = get_default_selected_tenant_id()
    biltys = scoped_query(TransportBill).filter_by(
        tenant_id=selected_tenant_id
    ).order_by(TransportBill.name).all()

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
            bilty = get_scoped_record(TransportBill, parse_int(request.form.get("bilty_id")))
            
            # Calculate balances
            current_balance = (bilty.freight_due or 0) - (bilty.paid_amount or 0)
            amount_received = parse_float(request.form.get("amount_received"))
            tds_amount = parse_float(request.form.get("tds_amount")) or 0
            optional_amount = parse_float(request.form.get("optional_amount")) or 0
            total_received = amount_received + tds_amount + optional_amount
            balance_after = current_balance - total_received
            
            # Generate receipt number
            receipt_count = PaymentReceipt.query.filter_by(tenant_id=tenant_id).count()
            receipt_number = f"PR-{date.today().strftime('%Y%m%d')}-{receipt_count + 1}"
            
            receipt = PaymentReceipt(
                tenant_id=tenant_id,
                bilty_id=bilty.id,
                receipt_number=receipt_number,
                receipt_date=parse_date(request.form.get("receipt_date")),
                amount_received=amount_received,
                tds_amount=tds_amount,
                optional_amount=optional_amount,
                optional_amount_reason=request.form.get("optional_amount_reason"),
                balance_before=current_balance,
                balance_after=balance_after,
                is_complete=(balance_after <= 0),
                remarks=request.form.get("remarks"),
                created_by=g.current_user.id if g.current_user else None
            )
            
            db.session.add(receipt)
            db.session.flush()
            
            # Update bilty paid amount (only the actual amount received, not TDS or optional)
            bilty.paid_amount = (bilty.paid_amount or 0) + amount_received
            
            # Update bilty payment status
            if balance_after <= 0:
                bilty.payment_status = "paid"
            elif amount_received > 0:
                bilty.payment_status = "partial"
            
            record_audit(
                "payment_receipt_created",
                receipt,
                summary=f"Payment receipt {receipt_number} created for {bilty.name}",
                details={
                    "bilty": bilty.name,
                    "amount_received": amount_received,
                    "balance_after": balance_after,
                },
            )
            db.session.commit()
            flash(f"Payment receipt {receipt_number} created successfully.", "success")
            return redirect(url_for("payment_receipts"))
            
        except ValueError as exc:
            flash(str(exc), "error")
        except Exception as exc:
            db.session.rollback()
            flash(f"Error creating payment receipt: {str(exc)}", "error")

    context = get_form_tenant_context()
    context.update({
        "biltys": biltys,
        "today": date.today().isoformat(),
    })
    return render_template("payment_receipts/form.html", **context)


@app.route("/payment-receipts", methods=["GET"])
@permission_required("transport_bills", "view")
def payment_receipts():
    selected_tenant_id = get_default_selected_tenant_id()
    receipts = scoped_query(PaymentReceipt).filter_by(
        tenant_id=selected_tenant_id
    ).order_by(PaymentReceipt.created_at.desc()).all()
    
    context = get_form_tenant_context()
    context.update({"receipts": receipts})
    return render_template("payment_receipts/list.html", **context)


@app.route("/transport-bills/edit/<int:id>", methods=["GET", "POST"])
@permission_required("transport_bills", "edit")
def edit_transport_bill(id):
    bill = get_scoped_record(TransportBill, id)
    selected_tenant_id = get_default_selected_tenant_id(bill)
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    locations = get_tenant_filtered_records(Location, Location.location, selected_tenant_id)
    delivery_types = get_tenant_filtered_records(
        DeliveryType, DeliveryType.delivery_type, selected_tenant_id
    )
    vendors = get_tenant_filtered_records(
        Vendor, Vendor.vendor_name, selected_tenant_id
    )
    drivers = get_tenant_filtered_records(
        Driver, Driver.first_name, selected_tenant_id
    )
    ratelists = scoped_query(RateList).filter_by(
        tenant_id=selected_tenant_id,
        is_active=True
    ).order_by(RateList.effective_date.desc(), RateList.name).all()

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(bill, required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle", True
            )
            location = get_related_record(
                Location, parse_int(request.form.get("location_id")), tenant_id, "Location"
            )
            delivery_type = get_related_record(
                DeliveryType,
                parse_int(request.form.get("delivery_type_id")),
                tenant_id,
                "Delivery Type",
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(bill)
            context.update(
                {
                    "bill": bill,
                    "vehicles": vehicles,
                    "locations": locations,
                    "delivery_types": delivery_types,
                    "vendors": vendors,
                    "drivers": drivers,
                    "ratelists": ratelists,
                    "edit": True,
                }
            )
            return render_template("transport_bills/form.html", **context)

        bill.tenant_id = tenant_id
        bill.vehicle_id = vehicle.id
        # GR Basic Details
        bill.gr_number = request.form.get("gr_number", "").strip() or None
        bill.gr_date = parse_date(request.form.get("gr_date"))
        bill.branch_booking_location = request.form.get("branch_booking_location", "").strip() or None
        bill.gr_type = request.form.get("gr_type", "").strip() or None
        bill.booking_mode = request.form.get("booking_mode", "road").strip() or "road"
        bill.status = request.form.get("status", "draft").strip() or "draft"
        bill.reference_number = request.form.get("reference_number", "").strip() or None
        # Party Details
        bill.consignor_name = request.form.get("consignor_name", "").strip() or None
        bill.consignor_address = request.form.get("consignor_address", "").strip() or None
        bill.consignor_gstin = request.form.get("consignor_gstin", "").strip() or None
        bill.consignor_contact_person = request.form.get("consignor_contact_person", "").strip() or None
        bill.consignor_mobile = request.form.get("consignor_mobile", "").strip() or None
        bill.consignor_email = request.form.get("consignor_email", "").strip() or None
        bill.consignee_name = request.form.get("consignee_name", "").strip() or None
        bill.consignee_address = request.form.get("consignee_address", "").strip() or None
        bill.consignee_gstin = request.form.get("consignee_gstin", "").strip() or None
        bill.consignee_contact_person = request.form.get("consignee_contact_person", "").strip() or None
        bill.consignee_mobile = request.form.get("consignee_mobile", "").strip() or None
        bill.consignee_email = request.form.get("consignee_email", "").strip() or None
        bill.bill_to_party = request.form.get("bill_to_party", "").strip() or None
        bill.ship_to_party = request.form.get("ship_to_party", "").strip() or None
        # Pickup Details
        bill.pickup_point = request.form.get("pickup_point", "").strip() or None
        bill.pickup_date = parse_date(request.form.get("pickup_date"))
        bill.pickup_branch = request.form.get("pickup_branch", "").strip() or None
        bill.loaded_by = request.form.get("loaded_by", "").strip() or None
        bill.loading_type = request.form.get("loading_type", "").strip() or None
        bill.pickup_remarks = request.form.get("pickup_remarks", "").strip() or None
        bill.pickup_city = request.form.get("pickup_city", "").strip() or None
        bill.pickup_state = request.form.get("pickup_state", "").strip() or None
        bill.pickup_pincode = request.form.get("pickup_pincode", "").strip() or None
        # Delivery Details
        bill.delivery_location = request.form.get("delivery_location", "").strip() or None
        bill.expected_delivery_date = parse_date(request.form.get("expected_delivery_date"))
        bill.delivery_type_id = delivery_type.id if delivery_type else None
        bill.delivery_city = request.form.get("delivery_city", "").strip() or None
        bill.delivery_state = request.form.get("delivery_state", "").strip() or None
        bill.delivery_pincode = request.form.get("delivery_pincode", "").strip() or None
        bill.unloading_point = request.form.get("unloading_point", "").strip() or None
        bill.delivery_remarks = request.form.get("delivery_remarks", "").strip() or None
        # Vehicle Details
        bill.vehicle_number = request.form.get("vehicle_number", "").strip() or None
        bill.vehicle_type = request.form.get("vehicle_type", "").strip() or None
        bill.vehicle_capacity = request.form.get("vehicle_capacity", "").strip() or None
        bill.vehicle_ownership = request.form.get("vehicle_ownership", "owned").strip() or "owned"
        bill.driver_name = request.form.get("driver_name", "").strip() or None
        bill.driver_mobile = request.form.get("driver_mobile", "").strip() or None
        bill.driver_license = request.form.get("driver_license", "").strip() or None
        bill.helper_name = request.form.get("helper_name", "").strip() or None
        bill.trip_number = request.form.get("trip_number", "").strip() or None
        bill.route_name = request.form.get("route_name", "").strip() or None
        # Material Details
        bill.article_description = request.form.get("article_description", "").strip() or None
        bill.package_type = request.form.get("package_type", "").strip() or None
        bill.number_of_packages = parse_int(request.form.get("number_of_packages")) or 0
        bill.packing_mode = request.form.get("packing_mode", "").strip() or None
        bill.actual_weight = parse_float(request.form.get("actual_weight"))
        bill.charged_weight = parse_float(request.form.get("charged_weight"))
        bill.volume_cbm = parse_float(request.form.get("volume_cbm"))
        bill.material_value = parse_float(request.form.get("material_value"))
        bill.dimensions = request.form.get("dimensions", "").strip() or None
        bill.goods_category = request.form.get("goods_category", "").strip() or None
        bill.nature_of_goods = request.form.get("nature_of_goods", "").strip() or None
        bill.hsn_code = request.form.get("hsn_code", "").strip() or None
        bill.declared_value = parse_float(request.form.get("declared_value"))
        bill.insurance_required = parse_bool(request.form.get("insurance_required"))
        bill.insurance_value = parse_float(request.form.get("insurance_value"))
        bill.handling_instructions = request.form.get("handling_instructions", "").strip() or None
        bill.goods_remarks = request.form.get("goods_remarks", "").strip() or None
        # Freight Charges
        bill.basic_freight = parse_float(request.form.get("basic_freight"))
        bill.loading_charge = parse_float(request.form.get("loading_charge"))
        bill.unloading_charge = parse_float(request.form.get("unloading_charge"))
        bill.door_pickup_charge = parse_float(request.form.get("door_pickup_charge"))
        bill.door_delivery_charge = parse_float(request.form.get("door_delivery_charge"))
        bill.hamali_charge = parse_float(request.form.get("hamali_charge"))
        bill.detention_charge = parse_float(request.form.get("detention_charge"))
        bill.waiting_charge = parse_float(request.form.get("waiting_charge"))
        bill.halting_charge = parse_float(request.form.get("halting_charge"))
        bill.toll_charge = parse_float(request.form.get("toll_charge"))
        bill.border_charge = parse_float(request.form.get("border_charge"))
        bill.fuel_surcharge = parse_float(request.form.get("fuel_surcharge"))
        bill.packing_charge = parse_float(request.form.get("packing_charge"))
        bill.weighment_charge = parse_float(request.form.get("weighment_charge"))
        bill.permit_charge = parse_float(request.form.get("permit_charge"))
        bill.driver_allowance = parse_float(request.form.get("driver_allowance"))
        bill.insurance_charge = parse_float(request.form.get("insurance_charge"))
        bill.other_charges = parse_float(request.form.get("other_charges"))
        bill.discount = parse_float(request.form.get("discount"))
        bill.net_freight = parse_float(request.form.get("net_freight"))
        # Tax Details
        bill.gst_applicable = parse_bool(request.form.get("gst_applicable"))
        bill.igst_rate = parse_float(request.form.get("igst_rate"))
        bill.cgst_rate = parse_float(request.form.get("cgst_rate"))
        bill.sgst_rate = parse_float(request.form.get("sgst_rate"))
        bill.taxable_amount = parse_float(request.form.get("taxable_amount"))
        bill.gst_amount = parse_float(request.form.get("gst_amount"))
        bill.eway_bill_number = request.form.get("eway_bill_number", "").strip() or None
        bill.eway_bill_date = parse_date(request.form.get("eway_bill_date"))
        bill.eway_bill_validity = parse_date(request.form.get("eway_bill_validity"))
        bill.invoice_number = request.form.get("invoice_number", "").strip() or None
        # Payment Details
        bill.freight_payment_mode = request.form.get("freight_payment_mode", "to_pay").strip() or "to_pay"
        bill.freight_collected_from = request.form.get("freight_collected_from", "").strip() or None
        bill.advance_amount = parse_float(request.form.get("advance_amount"))
        bill.balance_amount = parse_float(request.form.get("balance_amount"))
        bill.freight_due = parse_float(request.form.get("freight_due"))
        bill.paid_amount = parse_float(request.form.get("paid_amount"))
        bill.payment_status = request.form.get("payment_status", "pending").strip() or "pending"
        bill.receipt_number = request.form.get("receipt_number", "").strip() or None
        # Transit Details
        bill.dispatch_from_branch = request.form.get("dispatch_from_branch", "").strip() or None
        bill.reached_hub = request.form.get("reached_hub", "").strip() or None
        bill.transshipment_required = parse_bool(request.form.get("transshipment_required"))
        bill.transshipment_hub = request.form.get("transshipment_hub", "").strip() or None
        bill.intermediate_vehicle_number = request.form.get("intermediate_vehicle_number", "").strip() or None
        bill.expected_transit_time = request.form.get("expected_transit_time", "").strip() or None
        bill.actual_transit_time = request.form.get("actual_transit_time", "").strip() or None
        bill.delay_reason = request.form.get("delay_reason", "").strip() or None
        bill.tracking_number = request.form.get("tracking_number", "").strip() or None
        # Delivery Confirmation
        bill.delivered_by = request.form.get("delivered_by", "").strip() or None
        bill.received_by_name = request.form.get("received_by_name", "").strip() or None
        bill.pod_number = request.form.get("pod_number", "").strip() or None
        bill.seal_number = request.form.get("seal_number", "").strip() or None
        bill.damage_shortage_report = request.form.get("damage_shortage_report", "").strip() or None
        bill.closed_by = request.form.get("closed_by", "").strip() or None
        # Risk/Compliance
        bill.hazardous_material = parse_bool(request.form.get("hazardous_material"))
        bill.temperature_controlled = parse_bool(request.form.get("temperature_controlled"))
        bill.shortage_damage_flag = parse_bool(request.form.get("shortage_damage_flag"))
        bill.customer_complaint_number = request.form.get("customer_complaint_number", "").strip() or None
        bill.claim_amount = parse_float(request.form.get("claim_amount"))
        bill.gps_link = request.form.get("gps_link", "").strip() or None
        bill.permit_details = request.form.get("permit_details", "").strip() or None
        # Attachments
        bill.bilty_attachment_1_path = request.form.get("bilty_attachment_1_path", "").strip() or None
        bill.bilty_attachment_2_path = request.form.get("bilty_attachment_2_path", "").strip() or None
        # Legacy fields
        bill.challan_number = request.form.get("challan_number", "").strip() or None
        bill.party_information = request.form.get("party_information", "").strip() or None
        bill.location_id = location.id if location else None
        bill.rate = parse_int(request.form.get("rate")) or 0
        bill.modified_by = getattr(g, "current_user", None).username if getattr(g, "current_user", None) else None
        
        # Versioning: Create new version instead of modifying original
        modification_reason = request.form.get("modification_reason", "").strip()
        if not modification_reason:
            flash("Modification reason is required when editing a bilty.", "error")
            context = get_form_tenant_context(bill)
            context.update(
                {
                    "bill": bill,
                    "vehicles": vehicles,
                    "locations": locations,
                    "delivery_types": delivery_types,
                    "vendors": vendors,
                    "drivers": drivers,
                    "ratelists": ratelists,
                    "edit": True,
                }
            )
            return render_template("transport_bills/form.html", **context)
        
        # Create a new version
        import copy
        new_bill = copy.deepcopy(bill)
        new_bill.id = None
        new_bill.parent_id = bill.id
        new_bill.version_number = bill.version_number + 1
        new_bill.is_original = False
        new_bill.modification_reason = modification_reason
        new_bill.created_at = datetime.utcnow()
        new_bill.updated_at = datetime.utcnow()
        
        db.session.add(new_bill)
        db.session.commit()
        
        flash(f"Bilty {bill.name} modified successfully. New version {new_bill.version_number} created.", "success")
        return redirect(url_for("transport_bills"))

    context = get_form_tenant_context(bill)
    context.update(
        {
            "bill": bill,
            "vehicles": vehicles,
            "locations": locations,
            "delivery_types": delivery_types,
            "vendors": vendors,
            "edit": True,
            "today": date.today().isoformat(),
        }
    )
    return render_template("transport_bills/form.html", **context)


@app.route("/transport-bills/delete/<int:id>", methods=["POST"])
@permission_required("transport_bills", "delete")
def delete_transport_bill(id):
    bill = get_scoped_record(TransportBill, id)
    db.session.delete(bill)
    db.session.commit()
    flash("Bilty deleted successfully.", "success")
    return redirect(url_for("transport_bills"))


@app.route("/transport-bills/versions/<int:id>")
@permission_required("transport_bills", "view")
def bilty_versions(id):
    """View version history of a bilty"""
    bill = get_scoped_record(TransportBill, id)
    
    # Get the original bill (if this is a version, get the parent)
    original_bill = bill
    if bill.parent_id:
        original_bill = TransportBill.query.filter_by(id=bill.parent_id).first()
    
    # Get all versions including the original
    all_versions = []
    if original_bill:
        all_versions = [original_bill] + TransportBill.query.filter_by(parent_id=original_bill.id).order_by(TransportBill.version_number).all()
    
    return render_template("transport_bills/versions.html", bill=bill, versions=all_versions, original_bill=original_bill)


@app.route("/transport-bills/print/<int:id>")
@permission_required("transport_bills", "view")
def print_bilty(id):
    """Print bilty in professional format"""
    bill = get_scoped_record(TransportBill, id)
    return render_template("transport_bills/print.html", bill=bill)


@app.route("/transport-bills/invoice/<int:id>")
@permission_required("transport_bills", "view")
def print_invoice(id):
    """Printable invoice view with receipt and balance summary."""
    bill = get_scoped_record(TransportBill, id)
    receipts = PaymentReceipt.query.filter_by(bilty_id=bill.id, tenant_id=bill.tenant_id).all()
    total_paid = sum(float(receipt.amount_received or 0) for receipt in receipts)
    invoice_total = float(getattr(bill, "freight_due", None) or bill.rate or 0)
    return render_template(
        "transport_bills/invoice.html",
        bill=bill,
        receipts=receipts,
        invoice_total=invoice_total,
        total_paid=total_paid,
        balance_due=invoice_total - total_paid,
    )


@app.route("/transport-bills/delivery-confirmation", methods=["GET", "POST"])
@permission_required("transport_bills", "edit")
def delivery_confirmation():
    """Delivery confirmation page - select bilty and confirm delivery"""
    selected_bill_id = parse_int(request.args.get("bill_id"))
    # Get bills that are not yet delivered
    bills = scoped_query(TransportBill).filter(
        TransportBill.status != 'delivered'
    ).order_by(TransportBill.created_at.desc()).all()
    
    if request.method == "POST":
        try:
            bill_id = parse_int(request.form.get("bill_id"))
            if not bill_id:
                flash("Please select a bilty.", "error")
                return render_template("transport_bills/delivery_confirmation.html", bills=bills, selected_bill_id=selected_bill_id)
            
            bill = get_scoped_record(TransportBill, bill_id)
            trip = (
                DispatchTrip.query.filter_by(bilty_id=bill.id, tenant_id=bill.tenant_id)
                .order_by(DispatchTrip.created_at.desc())
                .first()
            )
            
            # Update bill with delivery confirmation details
            bill.status = 'delivered'
            
            # Store delivery details in extended_data JSON
            import json
            extended_data = {}
            if bill.extended_data:
                try:
                    extended_data = json.loads(bill.extended_data)
                except:
                    pass
            
            delivery_details = {
                'delivered_by': request.form.get("delivered_by"),
                'received_by_name': request.form.get("received_by_name"),
                'received_by_signature': request.form.get("received_by_signature"),
                'pod_number': request.form.get("pod_number"),
                'seal_number': request.form.get("seal_number"),
                'delivered_datetime': request.form.get("delivered_datetime"),
                'damage_shortage_report': request.form.get("damage_shortage_report"),
                'closed_by': request.form.get("closed_by"),
                'remarks': request.form.get("remarks")
            }
            
            extended_data['delivery_confirmation'] = delivery_details
            bill.extended_data = json.dumps(extended_data)
            
            # Handle file upload
            attachment = request.files.get("attachment")
            pod_attachment_path = None
            if attachment and attachment.filename:
                pod_attachment_path = save_uploaded_file(
                    attachment,
                    subfolder="pod",
                    prefix=f"pod_{bill.name}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
                )
                if pod_attachment_path:
                    delivery_details["pod_attachment_path"] = pod_attachment_path
                    extended_data['delivery_confirmation'] = delivery_details
                    bill.extended_data = json.dumps(extended_data)

            if trip:
                trip.status = "delivered"
                trip.actual_delivery_at = parse_datetime_local(
                    request.form.get("delivered_datetime"), datetime.utcnow()
                )
                trip.pod_number = request.form.get("pod_number")
                trip.pod_attachment_path = pod_attachment_path or trip.pod_attachment_path
                trip.received_by_name = request.form.get("received_by_name")
                trip.delivered_by = request.form.get("delivered_by")
                trip.delivery_remarks = request.form.get("remarks")
                trip.last_tracking_update_at = datetime.utcnow()
            
            record_audit(
                "delivery_confirmed",
                bill,
                summary=f"Delivery confirmed for bilty {bill.name}",
                details={
                    "pod_number": request.form.get("pod_number"),
                    "received_by_name": request.form.get("received_by_name"),
                    "pod_attachment_path": pod_attachment_path,
                },
            )
            db.session.commit()
            flash(f"Delivery confirmed for bilty {bill.name}. Status updated to delivered.", "success")
            return redirect(url_for("transport_bills"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error confirming delivery: {str(e)}", "error")
    
    return render_template("transport_bills/delivery_confirmation.html", bills=bills, selected_bill_id=selected_bill_id)


# POD Dashboard Routes
@app.route("/pod/dashboard")
@permission_required("transport_bills", "view")
def pod_dashboard():
    """POD Dashboard with KPIs and status overview"""
    tenant_id = get_tenant_id()
    
    # Get POD status counts
    status_counts = db.session.query(
        PodStatus.status_name,
        PodStatus.color_code,
        func.count(PodTracking.id).label('count')
    ).join(PodTracking).filter(
        PodTracking.tenant_id == tenant_id,
        PodTracking.timestamp == func.max_over(PodTracking.timestamp).partition_by(PodTracking.transport_bill_id)
    ).group_by(PodStatus.id).all()
    
    # Get aging data
    aging_data = db.session.execute(text("""
        SELECT 
            CASE 
                WHEN DATEDIFF(NOW(), pt.timestamp) <= 2 THEN '0-2 days'
                WHEN DATEDIFF(NOW(), pt.timestamp) <= 7 THEN '3-7 days'
                WHEN DATEDIFF(NOW(), pt.timestamp) <= 15 THEN '8-15 days'
                ELSE '15+ days'
            END as aging_bucket,
            COUNT(*) as count
        FROM pod_tracking pt
        WHERE pt.tenant_id = :tenant_id
        AND pt.timestamp = (
            SELECT MAX(timestamp) 
            FROM pod_tracking pt2 
            WHERE pt2.transport_bill_id = pt.transport_bill_id
        )
        AND pt.status_code != 'completed'
        GROUP BY aging_bucket
        ORDER BY aging_bucket
    """), {'tenant_id': tenant_id}).fetchall()
    
    # Get recent POD updates
    recent_updates = PodTracking.query.filter_by(tenant_id=tenant_id)\
        .order_by(PodTracking.timestamp.desc()).limit(10).all()
    
    return render_template('pod/dashboard.html',
                         status_counts=status_counts,
                         aging_data=aging_data,
                         recent_updates=recent_updates)


@app.route("/pod/list")
@permission_required("transport_bills", "view")
def pod_list():
    """List all PODs with filters"""
    tenant_id = get_tenant_id()
    status_filter = request.args.get('status', '')
    aging_filter = request.args.get('aging', '')
    
    # Build query for latest POD tracking per bill
    subquery = db.session.query(
        PodTracking.transport_bill_id,
        func.max(PodTracking.timestamp).label('max_timestamp')
    ).filter_by(tenant_id=tenant_id)\
     .group_by(PodTracking.transport_bill_id)\
     .subquery()
    
    query = db.session.query(
        TransportBill, PodStatus, PodTracking
    ).join(PodTracking, TransportBill.id == PodTracking.transport_bill_id)\
     .join(PodStatus, PodTracking.status_code == PodStatus.status_code)\
     .join(subquery, 
            db.and_(
                PodTracking.transport_bill_id == subquery.c.transport_bill_id,
                PodTracking.timestamp == subquery.c.max_timestamp
            ))\
     .filter(PodTracking.tenant_id == tenant_id)
    
    if status_filter:
        query = query.filter(PodTracking.status_code == status_filter)
    
    if aging_filter:
        if aging_filter == '0-2':
            query = query.filter(func.datediff(func.now(), PodTracking.timestamp) <= 2)
        elif aging_filter == '3-7':
            query = query.filter(func.datediff(func.now(), PodTracking.timestamp).between(3, 7))
        elif aging_filter == '8-15':
            query = query.filter(func.datediff(func.now(), PodTracking.timestamp).between(8, 15))
        elif aging_filter == '15+':
            query = query.filter(func.datediff(func.now(), PodTracking.timestamp) > 15)
    
    pods = query.order_by(PodTracking.timestamp.desc()).all()
    
    return render_template('pod/list.html', pods=pods, 
                         status_filter=status_filter, aging_filter=aging_filter)


@app.route("/pod/update_status/<int:bill_id>", methods=['POST'])
@permission_required("transport_bills", "edit")
def update_pod_status(bill_id):
    """Update POD status"""
    tenant_id = get_tenant_id()
    
    bill = TransportBill.query.filter_by(id=bill_id, tenant_id=tenant_id).first_or_404()
    new_status = request.form.get('status')
    remarks = request.form.get('remarks', '')
    location = request.form.get('location', '')
    
    # Create new tracking entry
    tracking = PodTracking(
        transport_bill_id=bill_id,
        status_code=new_status,
        updated_by=current_user.id,
        remarks=remarks,
        location=location,
        tenant_id=tenant_id
    )
    
    db.session.add(tracking)
    
    # Update transport bill with POD details if completed
    if new_status == 'completed':
        bill.pod_number = request.form.get('pod_number', '')
        bill.received_by_name = request.form.get('received_by_name', '')
        bill.delivered_by = request.form.get('delivered_by', '')
        bill.delivery_remarks = remarks
        bill.status = 'delivered'
    
    db.session.commit()
    
    flash('POD status updated successfully', 'success')
    return redirect(url_for('pod_list'))


@app.route("/reports")
@permission_required("reports", "view")
def reports():
    return render_template("reports/dashboard.html")


@app.route("/reports/transport-bill")
@permission_required("reports", "view")
def transport_bill_report():
    vehicles = scoped_query(Vehicle).order_by(Vehicle.registration_number).all()
    return render_template(
        "reports/index.html",
        vehicles=vehicles,
        show_tenant_column=is_superadmin(),
    )


@app.route("/reports/generate", methods=["POST"])
@permission_required("reports", "view")
def generate_report():
    export = request.form.get("export") == "csv"
    if export and not has_permission("reports", "export"):
        flash("You do not have permission to export reports.", "error")
        return redirect(url_for("reports"))

    query = scoped_query(TransportBill)
    vehicle_id = parse_int(request.form.get("vehicle_id"))
    from_date = parse_date(request.form.get("from_date"))
    to_date = parse_date(request.form.get("to_date"))

    if vehicle_id:
        vehicle = scoped_query(Vehicle).filter_by(id=vehicle_id).first()
        if not vehicle:
            flash("Selected vehicle is not available.", "error")
            return redirect(url_for("reports"))
        query = query.filter_by(vehicle_id=vehicle_id)

    if from_date:
        query = query.filter(TransportBill.date >= from_date)
    if to_date:
        query = query.filter(TransportBill.date <= to_date)

    bills = query.order_by(TransportBill.date.desc()).all()

    if export:
        output = io.StringIO()
        writer = csv.writer(output)
        headers = ["Bill No", "Date"]
        if is_superadmin():
            headers.append("Tenant")
        headers.extend(
            ["Vehicle", "Challan No", "Party", "Location", "Delivery Type", "Rate"]
        )
        writer.writerow(headers)
        for bill in bills:
            row = [bill.name, bill.date.strftime("%Y-%m-%d") if bill.date else ""]
            if is_superadmin():
                row.append(bill.tenant.name if bill.tenant else "")
            row.extend(
                [
                    bill.vehicle.registration_number if bill.vehicle else "",
                    bill.challan_number or "",
                    bill.party_information or "",
                    bill.location.location if bill.location else "",
                    bill.delivery_type.delivery_type if bill.delivery_type else "",
                    bill.rate,
                ]
            )
            writer.writerow(row)

        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=transport_report.csv"},
        )

    vehicles = scoped_query(Vehicle).order_by(Vehicle.registration_number).all()
    return render_template(
        "reports/index.html",
        bills=bills,
        vehicles=vehicles,
        filter_vehicle=str(vehicle_id) if vehicle_id else "",
        filter_from=from_date.isoformat() if from_date else "",
        filter_to=to_date.isoformat() if to_date else "",
        show_tenant_column=is_superadmin(),
    )


@app.route("/reports/revenue")
@permission_required("reports", "view")
def revenue_report():
    from sqlalchemy import extract
    import io
    import csv
    
    export = request.args.get('export') == 'csv'
    if export and not has_permission("reports", "export"):
        flash("You do not have permission to export reports.", "error")
        return redirect(url_for('revenue_report'))
    
    # Get date range from query params or default to current year
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    period_type = request.args.get('period', 'monthly')  # monthly, quarterly, yearly
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(month=1, day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    query = scoped_query(TransportBill).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    )
    
    # Group by period
    if period_type == 'monthly':
        revenue_data = db.session.query(
            extract('year', TransportBill.date).label('year'),
            extract('month', TransportBill.date).label('month'),
            db.func.sum(TransportBill.rate).label('total'),
            db.func.count(TransportBill.id).label('count')
        ).filter(
            TransportBill.date >= from_date,
            TransportBill.date <= to_date
        ).group_by(
            extract('year', TransportBill.date),
            extract('month', TransportBill.date)
        ).order_by('year', 'month').all()
    elif period_type == 'quarterly':
        revenue_data = db.session.query(
            extract('year', TransportBill.date).label('year'),
            ((extract('month', TransportBill.date) - 1) / 3 + 1).label('quarter'),
            db.func.sum(TransportBill.rate).label('total'),
            db.func.count(TransportBill.id).label('count')
        ).filter(
            TransportBill.date >= from_date,
            TransportBill.date <= to_date
        ).group_by(
            extract('year', TransportBill.date),
            ((extract('month', TransportBill.date) - 1) / 3 + 1)
        ).order_by('year', 'quarter').all()
    else:  # yearly
        revenue_data = db.session.query(
            extract('year', TransportBill.date).label('year'),
            db.func.sum(TransportBill.rate).label('total'),
            db.func.count(TransportBill.id).label('count')
        ).filter(
            TransportBill.date >= from_date,
            TransportBill.date <= to_date
        ).group_by(
            extract('year', TransportBill.date)
        ).order_by('year').all()
    
    total_revenue = sum(row.total for row in revenue_data) if revenue_data else 0
    
    if export:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Period', 'Bills Count', 'Total Revenue', 'Average per Bill'])
        for row in revenue_data:
            if period_type == 'monthly':
                period = f"{calendar.month_name[int(row.month)]} {int(row.year)}"
            elif period_type == 'quarterly':
                period = f"Q{int(row.quarter)} {int(row.year)}"
            else:
                period = str(int(row.year))
            writer.writerow([period, row.count, row.total, row.total / row.count if row.count > 0 else 0])
        writer.writerow(['TOTAL', '', total_revenue, ''])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=revenue_report.csv"},
        )
    
    import calendar
    return render_template(
        "reports/revenue.html",
        revenue_data=revenue_data,
        total_revenue=total_revenue,
        from_date=from_date,
        to_date=to_date,
        period_type=period_type,
        calendar=calendar,
    )


@app.route("/reports/payment-collection")
@permission_required("reports", "view")
def payment_collection_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get all bills in date range
    bills = scoped_query(TransportBill).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    ).all()
    
    # Get payment receipts
    receipts = scoped_query(PaymentReceipt).filter(
        PaymentReceipt.receipt_date >= from_date,
        PaymentReceipt.receipt_date <= to_date
    ).all()
    
    total_billed = sum(bill.rate for bill in bills) if bills else 0
    total_collected = sum(receipt.amount_received for receipt in receipts) if receipts else 0
    pending_amount = total_billed - total_collected
    
    # Group by party/customer
    party_data = {}
    for bill in bills:
        party = bill.party_information or 'Unknown'
        if party not in party_data:
            party_data[party] = {'billed': 0, 'collected': 0}
        party_data[party]['billed'] += bill.rate
    
    for receipt in receipts:
        if receipt.transport_bill and receipt.transport_bill.party_information:
            party = receipt.transport_bill.party_information
            if party in party_data:
                party_data[party]['collected'] += receipt.amount_received
    
    return render_template(
        "reports/payment_collection.html",
        bills=bills,
        receipts=receipts,
        total_billed=total_billed,
        total_collected=total_collected,
        pending_amount=pending_amount,
        party_data=party_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/expense")
@permission_required("reports", "view")
def expense_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    category = request.args.get('category')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    query = scoped_query(Expense).filter(
        Expense.expense_date >= from_date,
        Expense.expense_date <= to_date
    )
    
    if category:
        query = query.filter(Expense.category == category)
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Group by category
    category_data = {}
    for expense in expenses:
        cat = expense.category or 'Uncategorized'
        if cat not in category_data:
            category_data[cat] = {'count': 0, 'total': 0}
        category_data[cat]['count'] += 1
        category_data[cat]['total'] += expense.amount
    
    total_expenses = sum(exp.amount for exp in expenses) if expenses else 0
    
    # Get unique categories
    all_categories = db.session.query(Expense.category).distinct().all()
    all_categories = [c[0] for c in all_categories if c[0]]
    
    return render_template(
        "reports/expense.html",
        expenses=expenses,
        category_data=category_data,
        total_expenses=total_expenses,
        all_categories=all_categories,
        selected_category=category,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/loan")
@permission_required("reports", "view")
def loan_report():
    loans = scoped_query(Loan).order_by(Loan.loan_date.desc()).all()
    
    total_loan_amount = sum(loan.loan_amount for loan in loans) if loans else 0
    total_paid = sum(loan.amount_paid for loan in loans) if loans else 0
    total_outstanding = total_loan_amount - total_paid
    
    # Active loans
    active_loans = [loan for loan in loans if loan.amount_paid < loan.loan_amount]
    
    return render_template(
        "reports/loan.html",
        loans=loans,
        active_loans=active_loans,
        total_loan_amount=total_loan_amount,
        total_paid=total_paid,
        total_outstanding=total_outstanding,
    )


@app.route("/reports/profit-loss")
@permission_required("reports", "view")
def profit_loss_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Total revenue from transport bills
    revenue = db.session.query(
        db.func.sum(TransportBill.rate)
    ).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    ).scalar() or 0
    
    # Total expenses
    expenses = db.session.query(
        db.func.sum(Expense.amount)
    ).filter(
        Expense.expense_date >= from_date,
        Expense.expense_date <= to_date
    ).scalar() or 0
    
    # Total loan repayments (considered as expense)
    loan_payments = db.session.query(
        db.func.sum(Loan.amount_paid)
    ).filter(
        Loan.loan_date >= from_date,
        Loan.loan_date <= to_date
    ).scalar() or 0
    
    total_expenses = expenses + loan_payments
    profit_loss = revenue - total_expenses
    
    # Breakdown by expense category
    expense_breakdown = db.session.query(
        Expense.category,
        db.func.sum(Expense.amount)
    ).filter(
        Expense.expense_date >= from_date,
        Expense.expense_date <= to_date
    ).group_by(Expense.category).all()
    
    return render_template(
        "reports/profit_loss.html",
        revenue=revenue,
        expenses=expenses,
        loan_payments=loan_payments,
        total_expenses=total_expenses,
        profit_loss=profit_loss,
        expense_breakdown=expense_breakdown,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/vehicle-utilization")
@permission_required("reports", "view")
def vehicle_utilization_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    vehicles = scoped_query(Vehicle).all()
    
    vehicle_data = []
    for vehicle in vehicles:
        bills = scoped_query(TransportBill).filter(
            TransportBill.vehicle_id == vehicle.id,
            TransportBill.date >= from_date,
            TransportBill.date <= to_date
        ).all()
        
        trip_count = len(bills)
        total_revenue = sum(bill.rate for bill in bills) if bills else 0
        avg_revenue = total_revenue / trip_count if trip_count > 0 else 0
        
        vehicle_data.append({
            'vehicle': vehicle,
            'trip_count': trip_count,
            'total_revenue': total_revenue,
            'avg_revenue': avg_revenue,
        })
    
    # Sort by total revenue
    vehicle_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return render_template(
        "reports/vehicle_utilization.html",
        vehicle_data=vehicle_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/driver-performance")
@permission_required("reports", "view")
def driver_performance_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    drivers = scoped_query(Driver).all()
    
    driver_data = []
    for driver in drivers:
        # Get vehicles assigned to this driver
        vehicles = scoped_query(Vehicle).filter(Vehicle.driver_id == driver.id).all()
        vehicle_ids = [v.id for v in vehicles]
        
        # Get bills for these vehicles
        if vehicle_ids:
            bills = scoped_query(TransportBill).filter(
                TransportBill.vehicle_id.in_(vehicle_ids),
                TransportBill.date >= from_date,
                TransportBill.date <= to_date
            ).all()
        else:
            bills = []
        
        trip_count = len(bills)
        total_revenue = sum(bill.rate for bill in bills) if bills else 0
        avg_revenue = total_revenue / trip_count if trip_count > 0 else 0
        
        # Calculate on-time delivery rate (assuming delivery_date vs actual delivery)
        on_time = sum(1 for bill in bills if bill.delivery_date and bill.actual_delivery_date and bill.actual_delivery_date <= bill.delivery_date)
        on_time_rate = (on_time / trip_count * 100) if trip_count > 0 else 0
        
        driver_data.append({
            'driver': driver,
            'trip_count': trip_count,
            'total_revenue': total_revenue,
            'avg_revenue': avg_revenue,
            'on_time_rate': on_time_rate,
        })
    
    # Sort by total revenue
    driver_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return render_template(
        "reports/driver_performance.html",
        driver_data=driver_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/route-efficiency")
@permission_required("reports", "view")
def route_efficiency_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Group by location
    location_data = db.session.query(
        Location,
        db.func.count(TransportBill.id).label('trip_count'),
        db.func.sum(TransportBill.rate).label('total_revenue')
    ).join(
        TransportBill, Location.id == TransportBill.location_id
    ).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    ).group_by(Location.id).all()
    
    route_data = []
    for location, trip_count, total_revenue in location_data:
        avg_revenue = total_revenue / trip_count if trip_count > 0 else 0
        route_data.append({
            'location': location,
            'trip_count': trip_count,
            'total_revenue': total_revenue,
            'avg_revenue': avg_revenue,
        })
    
    # Sort by total revenue
    route_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return render_template(
        "reports/route_efficiency.html",
        route_data=route_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/location")
@permission_required("reports", "view")
def location_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Group by location for pickup and delivery
    pickup_data = db.session.query(
        Location,
        db.func.count(TransportBill.id).label('count')
    ).join(
        TransportBill, Location.id == TransportBill.location_id
    ).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    ).group_by(Location.id).all()
    
    return render_template(
        "reports/location.html",
        pickup_data=pickup_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/tenant-activity")
@permission_required("reports", "view")
def tenant_activity_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # If not superadmin, only show current tenant's data
    if is_superadmin():
        tenants = Tenant.query.all()
    else:
        tenants = [g.current_tenant] if g.current_tenant else []
    
    tenant_data = []
    for tenant in tenants:
        bills = TransportBill.query.filter(
            TransportBill.tenant_id == tenant.id,
            TransportBill.date >= from_date,
            TransportBill.date <= to_date
        ).all()
        
        bill_count = len(bills)
        total_revenue = sum(bill.rate for bill in bills) if bills else 0
        user_count = len(tenant.users)
        
        tenant_data.append({
            'tenant': tenant,
            'bill_count': bill_count,
            'total_revenue': total_revenue,
            'user_count': user_count,
        })
    
    # Sort by total revenue
    tenant_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return render_template(
        "reports/tenant_activity.html",
        tenant_data=tenant_data,
        from_date=from_date,
        to_date=to_date,
    )


@app.route("/reports/subscription")
@permission_required("reports", "view")
def subscription_report():
    # If not superadmin, only show current tenant's data
    if is_superadmin():
        tenants = Tenant.query.all()
    else:
        tenants = [g.current_tenant] if g.current_tenant else []
    
    today = datetime.now().date()
    
    expiring_soon = []
    expired = []
    active = []
    
    for tenant in tenants:
        if tenant.plan_expiry_date:
            days_until_expiry = (tenant.plan_expiry_date - today).days
            if days_until_expiry < 0:
                expired.append(tenant)
            elif days_until_expiry <= 30:
                expiring_soon.append(tenant)
            else:
                active.append(tenant)
        else:
            active.append(tenant)
    
    return render_template(
        "reports/subscription.html",
        expiring_soon=expiring_soon,
        expired=expired,
        active=active,
    )


@app.route("/reports/document-expiry")
@permission_required("reports", "view")
def document_expiry_report():
    from models import Driver, Vehicle
    
    today = datetime.now().date()
    warning_days = 30
    
    # Driver documents expiring
    drivers = Driver.query.all()
    driver_alerts = []
    for driver in drivers:
        if driver.license_expiry_date:
            days_until = (driver.license_expiry_date - today).days
            if days_until <= warning_days:
                driver_alerts.append({
                    'type': 'Driver License',
                    'name': f"{driver.first_name} {driver.last_name}",
                    'expiry_date': driver.license_expiry_date,
                    'days_until': days_until,
                })
    
    # Vehicle documents expiring
    vehicles = Vehicle.query.all()
    vehicle_alerts = []
    for vehicle in vehicles:
        if vehicle.insurance_expiry:
            days_until = (vehicle.insurance_expiry - today).days
            if days_until <= warning_days:
                vehicle_alerts.append({
                    'type': 'Vehicle Insurance',
                    'name': vehicle.registration_number,
                    'expiry_date': vehicle.insurance_expiry,
                    'days_until': days_until,
                })
        if vehicle.fitness_expiry:
            days_until = (vehicle.fitness_expiry - today).days
            if days_until <= warning_days:
                vehicle_alerts.append({
                    'type': 'Vehicle Fitness',
                    'name': vehicle.registration_number,
                    'expiry_date': vehicle.fitness_expiry,
                    'days_until': days_until,
                })
        if vehicle.puc_expiry:
            days_until = (vehicle.puc_expiry - today).days
            if days_until <= warning_days:
                vehicle_alerts.append({
                    'type': 'Vehicle PUC',
                    'name': vehicle.registration_number,
                    'expiry_date': vehicle.puc_expiry,
                    'days_until': days_until,
                })
    
    all_alerts = driver_alerts + vehicle_alerts
    all_alerts.sort(key=lambda x: x['days_until'])
    
    return render_template(
        "reports/document_expiry.html",
        all_alerts=all_alerts,
        warning_days=warning_days,
    )


@app.route("/reports/gst")
@permission_required("reports", "view")
def gst_report():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Group by party GSTIN (assuming party_information contains GST info or we need to add GSTIN field)
    # For now, group by party
    gst_data = db.session.query(
        TransportBill.party_information,
        db.func.count(TransportBill.id).label('bill_count'),
        db.func.sum(TransportBill.rate).label('total_amount')
    ).filter(
        TransportBill.date >= from_date,
        TransportBill.date <= to_date
    ).group_by(TransportBill.party_information).all()
    
    total_amount = sum(data.total_amount for data in gst_data) if gst_data else 0
    
    return render_template(
        "reports/gst.html",
        gst_data=gst_data,
        total_amount=total_amount,
        from_date=from_date,
        to_date=to_date,
    )


# =============================================================================
# Vendor Reports
# =============================================================================

@app.route("/reports/vendor-performance")
@permission_required("reports", "view")
def vendor_performance_report():
    """Vendor performance report showing trips, delivery rates, on-time performance"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_id:
        vendors_query = vendors_query.filter(Vendor.id == int(vendor_id))
    vendors = vendors_query.all()
    
    # Calculate performance metrics for each vendor
    vendor_performance = []
    for vendor in vendors:
        # Get trips associated with this vendor (assuming vendor is linked through some relationship)
        # For now, we'll create placeholder data structure
        trips_count = 0
        on_time_count = 0
        total_revenue = 0
        
        vendor_performance.append({
            'vendor': vendor,
            'trips_count': trips_count,
            'on_time_count': on_time_count,
            'on_time_rate': (on_time_count / trips_count * 100) if trips_count > 0 else 0,
            'total_revenue': total_revenue
        })
    
    return render_template(
        "reports/vendor_performance.html",
        vendor_performance=vendor_performance,
        vendors=vendors,
        from_date=from_date,
        to_date=to_date,
        selected_vendor_id=int(vendor_id) if vendor_id else None,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/vendor-payments")
@permission_required("reports", "view")
def vendor_payment_report():
    """Vendor payment history report"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    status = request.args.get('status')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_id:
        vendors_query = vendors_query.filter(Vendor.id == int(vendor_id))
    vendors = vendors_query.all()
    
    # Placeholder for payment data
    vendor_payments = []
    for vendor in vendors:
        vendor_payments.append({
            'vendor': vendor,
            'total_paid': 0,
            'pending_amount': 0,
            'payment_count': 0,
            'last_payment_date': None
        })
    
    return render_template(
        "reports/vendor_payments.html",
        vendor_payments=vendor_payments,
        vendors=vendors,
        from_date=from_date,
        to_date=to_date,
        selected_vendor_id=int(vendor_id) if vendor_id else None,
        selected_status=status,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/vendor-compliance")
@permission_required("reports", "view")
def vendor_compliance_report():
    """Vendor compliance report showing KYC status, document verification"""
    vendor_type = request.args.get('vendor_type')
    kyc_status = request.args.get('kyc_status')
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_type:
        vendors_query = vendors_query.filter(Vendor.vendor_type == vendor_type)
    if kyc_status:
        vendors_query = vendors_query.filter(Vendor.kyc_status == kyc_status)
    vendors = vendors_query.all()
    
    # Calculate compliance stats
    total_vendors = len(vendors)
    verified_count = sum(1 for v in vendors if v.kyc_status == 'verified')
    pending_count = sum(1 for v in vendors if v.kyc_status == 'pending')
    rejected_count = sum(1 for v in vendors if v.kyc_status == 'rejected')
    
    return render_template(
        "reports/vendor_compliance.html",
        vendors=vendors,
        total_vendors=total_vendors,
        verified_count=verified_count,
        pending_count=pending_count,
        rejected_count=rejected_count,
        selected_vendor_type=vendor_type,
        selected_kyc_status=kyc_status,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/vendor-revenue")
@permission_required("reports", "view")
def vendor_revenue_report():
    """Vendor-wise revenue report"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_id:
        vendors_query = vendors_query.filter(Vendor.id == int(vendor_id))
    vendors = vendors_query.all()
    
    # Calculate revenue for each vendor
    vendor_revenue = []
    total_revenue = 0
    for vendor in vendors:
        revenue = 0  # Placeholder - calculate from actual data
        total_revenue += revenue
        vendor_revenue.append({
            'vendor': vendor,
            'revenue': revenue,
            'profit_margin': 0  # Placeholder
        })
    
    return render_template(
        "reports/vendor_revenue.html",
        vendor_revenue=vendor_revenue,
        vendors=vendors,
        total_revenue=total_revenue,
        from_date=from_date,
        to_date=to_date,
        selected_vendor_id=int(vendor_id) if vendor_id else None,
        show_tenant_column=is_superadmin()
    )


# =============================================================================
# Subscription Reports
# =============================================================================

@app.route("/reports/subscription-revenue")
@permission_required("reports", "view")
def subscription_revenue_report():
    """Subscription revenue report by plan and tenant"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    plan_id = request.args.get('plan_id')
    tenant_id = request.args.get('tenant_id')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get plans and tenants
    plans = SubscriptionPlan.query.all()
    tenants_query = Tenant.query
    if not is_superadmin():
        tenants_query = tenants_query.filter(Tenant.id == g.current_user.tenant_id)
    tenants = tenants_query.all()
    
    # Get subscription payments
    payments_query = SubscriptionPayment.query.filter(
        SubscriptionPayment.payment_date >= from_date,
        SubscriptionPayment.payment_date <= to_date
    )
    if plan_id:
        payments_query = payments_query.join(TenantSubscription).filter(
            TenantSubscription.plan_id == int(plan_id)
        )
    if tenant_id:
        payments_query = payments_query.join(TenantSubscription).filter(
            TenantSubscription.tenant_id == int(tenant_id)
        )
    payments = payments_query.all()
    
    # Calculate revenue by plan
    revenue_by_plan = {}
    total_revenue = 0
    for payment in payments:
        plan_name = payment.subscription.plan.name if payment.subscription and payment.subscription.plan else 'Unknown'
        if plan_name not in revenue_by_plan:
            revenue_by_plan[plan_name] = 0
        revenue_by_plan[plan_name] += payment.amount
        total_revenue += payment.amount
    
    return render_template(
        "reports/subscription_revenue.html",
        revenue_by_plan=revenue_by_plan,
        total_revenue=total_revenue,
        payments=payments,
        plans=plans,
        tenants=tenants,
        from_date=from_date,
        to_date=to_date,
        selected_plan_id=int(plan_id) if plan_id else None,
        selected_tenant_id=int(tenant_id) if tenant_id else None,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/subscription-usage")
@permission_required("reports", "view")
def subscription_usage_report():
    """Subscription usage report showing vehicles/drivers/users vs limits"""
    tenant_id = request.args.get('tenant_id')
    
    # Get tenants
    tenants_query = Tenant.query
    if not is_superadmin():
        tenants_query = tenants_query.filter(Tenant.id == g.current_user.tenant_id)
    if tenant_id:
        tenants_query = tenants_query.filter(Tenant.id == int(tenant_id))
    tenants = tenants_query.all()
    
    # Calculate usage for each tenant
    usage_data = []
    for tenant in tenants:
        subscription = TenantSubscription.query.filter_by(
            tenant_id=tenant.id,
            status='active'
        ).first()
        
        if subscription and subscription.plan:
            vehicles_count = Vehicle.query.filter_by(tenant_id=tenant.id).count()
            drivers_count = Driver.query.filter_by(tenant_id=tenant.id).count()
            users_count = User.query.filter_by(tenant_id=tenant.id).count()
            
            usage_data.append({
                'tenant': tenant,
                'subscription': subscription,
                'plan': subscription.plan,
                'vehicles_used': vehicles_count,
                'vehicles_limit': subscription.plan.max_vehicles,
                'drivers_used': drivers_count,
                'drivers_limit': subscription.plan.max_drivers,
                'users_used': users_count,
                'users_limit': subscription.plan.max_users,
                'vehicles_overage': max(0, vehicles_count - subscription.plan.max_vehicles),
                'drivers_overage': max(0, drivers_count - subscription.plan.max_drivers),
                'users_overage': max(0, users_count - subscription.plan.max_users)
            })
    
    return render_template(
        "reports/subscription_usage.html",
        usage_data=usage_data,
        tenants=tenants,
        selected_tenant_id=int(tenant_id) if tenant_id else None,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/subscription-payments")
@permission_required("reports", "view")
def subscription_payment_history_report():
    """Subscription payment history report"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    tenant_id = request.args.get('tenant_id')
    status = request.args.get('status')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get tenants
    tenants_query = Tenant.query
    if not is_superadmin():
        tenants_query = tenants_query.filter(Tenant.id == g.current_user.tenant_id)
    if tenant_id:
        tenants_query = tenants_query.filter(Tenant.id == int(tenant_id))
    tenants = tenants_query.all()
    
    # Get payments
    payments_query = SubscriptionPayment.query.filter(
        SubscriptionPayment.payment_date >= from_date,
        SubscriptionPayment.payment_date <= to_date
    )
    if tenant_id:
        payments_query = payments_query.join(TenantSubscription).filter(
            TenantSubscription.tenant_id == int(tenant_id)
        )
    if status:
        payments_query = payments_query.filter(SubscriptionPayment.status == status)
    payments = payments_query.order_by(SubscriptionPayment.payment_date.desc()).all()
    
    total_amount = sum(p.amount for p in payments) if payments else 0
    
    return render_template(
        "reports/subscription_payments.html",
        payments=payments,
        total_amount=total_amount,
        tenants=tenants,
        from_date=from_date,
        to_date=to_date,
        selected_tenant_id=int(tenant_id) if tenant_id else None,
        selected_status=status,
        show_tenant_column=is_superadmin()
    )


# =============================================================================
# Vendor User Reports
# =============================================================================

@app.route("/reports/vendor-user-activity")
@permission_required("reports", "view")
def vendor_user_activity_report():
    """Vendor user login activity report"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_id:
        vendors_query = vendors_query.filter(Vendor.id == int(vendor_id))
    vendors = vendors_query.all()
    
    # Get vendor users
    vendor_users = []
    for vendor in vendors:
        for vu in vendor.vendor_users:
            vendor_users.append({
                'vendor_user': vu,
                'vendor': vendor,
                'user': vu.user,
                'last_login': vu.user.created_at,  # Placeholder - should track actual login
                'login_count': 0  # Placeholder - should track actual logins
            })
    
    return render_template(
        "reports/vendor_user_activity.html",
        vendor_users=vendor_users,
        vendors=vendors,
        from_date=from_date,
        to_date=to_date,
        selected_vendor_id=int(vendor_id) if vendor_id else None,
        show_tenant_column=is_superadmin()
    )


@app.route("/reports/vendor-field-access")
@permission_required("reports", "view")
def vendor_field_access_report():
    """Vendor user field access logs report"""
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    field_category = request.args.get('field_category')
    
    if from_date_str:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    else:
        from_date = datetime.now().date().replace(day=1)
    
    if to_date_str:
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    else:
        to_date = datetime.now().date()
    
    # Get vendors
    vendors_query = scoped_query(Vendor)
    if vendor_id:
        vendors_query = vendors_query.filter(Vendor.id == int(vendor_id))
    vendors = vendors_query.all()
    
    # Get field permissions
    field_permissions = []
    for vendor in vendors:
        for vu in vendor.vendor_users:
            for perm in vu.field_permissions:
                if not field_category or perm.field_category == field_category:
                    field_permissions.append({
                        'vendor_user': vu,
                        'vendor': vendor,
                        'user': vu.user,
                        'permission': perm,
                        'created_at': perm.created_at
                    })
    
    return render_template(
        "reports/vendor_field_access.html",
        field_permissions=field_permissions,
        vendors=vendors,
        from_date=from_date,
        to_date=to_date,
        selected_vendor_id=int(vendor_id) if vendor_id else None,
        selected_field_category=field_category,
        show_tenant_column=is_superadmin()
    )


@app.route("/expenses")
@permission_required("expenses", "view")
def expenses():
    expense_query = scoped_query(Expense)
    expense_list = expense_query.order_by(Expense.expense_date.desc()).all()
    total_expenses = expense_query.with_entities(db.func.sum(Expense.amount)).scalar() or 0
    return render_template(
        "expenses/list.html",
        expenses=expense_list,
        total_expenses=total_expenses,
        show_tenant_column=is_superadmin(),
    )


@app.route("/expenses/create", methods=["GET", "POST"])
@permission_required("expenses", "create")
def create_expense():
    selected_tenant_id = get_default_selected_tenant_id()
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    locations = get_tenant_filtered_records(Location, Location.location, selected_tenant_id)
    categories = [
        "Fuel",
        "Maintenance",
        "Insurance",
        "Tolls",
        "Driver Salary",
        "Spare Parts",
        "Tyres",
        "Lubricants",
        "Car Wash",
        "Parking",
        "Fine/Penalty",
        "Other",
    ]

    if request.method == "POST":
        category = request.form.get("category", "").strip()
        amount = parse_float(request.form.get("amount"))
        try:
            tenant_id = resolve_target_tenant_id(required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle"
            )
            location = get_related_record(
                Location, parse_int(request.form.get("location_id")), tenant_id, "Location"
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context()
            context.update(
                {
                    "vehicles": vehicles,
                    "locations": locations,
                    "categories": categories,
                    "today": date.today().isoformat(),
                }
            )
            return render_template("expenses/form.html", **context)

        if not category or amount <= 0:
            flash("Category and a valid amount are required.", "error")
            context = get_form_tenant_context()
            context.update(
                {
                    "vehicles": vehicles,
                    "locations": locations,
                    "categories": categories,
                    "today": date.today().isoformat(),
                }
            )
            return render_template("expenses/form.html", **context)

        expense = Expense(
            tenant_id=tenant_id,
            name=generate_expense_name(),
            expense_date=parse_date(request.form.get("expense_date"), date.today()),
            category=category,
            description=request.form.get("description", "").strip() or None,
            amount=amount,
            vehicle_id=vehicle.id if vehicle else None,
            location_id=location.id if location else None,
            payment_method=request.form.get("payment_method", "Cash"),
            vendor_name=request.form.get("vendor_name", "").strip() or None,
            vendor_contact=request.form.get("vendor_contact", "").strip() or None,
            bill_number=request.form.get("bill_number", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            status=request.form.get("status", "Pending"),
        )
        db.session.add(expense)
        db.session.commit()
        flash(f"Expense {expense.name} added successfully.", "success")
        return redirect(url_for("expenses"))

    context = get_form_tenant_context()
    context.update(
        {
            "vehicles": vehicles,
            "locations": locations,
            "categories": categories,
            "today": date.today().isoformat(),
        }
    )
    return render_template("expenses/form.html", **context)


@app.route("/expenses/edit/<int:id>", methods=["GET", "POST"])
@permission_required("expenses", "edit")
def edit_expense(id):
    expense = get_scoped_record(Expense, id)
    selected_tenant_id = get_default_selected_tenant_id(expense)
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    locations = get_tenant_filtered_records(Location, Location.location, selected_tenant_id)
    categories = [
        "Fuel",
        "Maintenance",
        "Insurance",
        "Tolls",
        "Driver Salary",
        "Spare Parts",
        "Tyres",
        "Lubricants",
        "Car Wash",
        "Parking",
        "Fine/Penalty",
        "Other",
    ]

    if request.method == "POST":
        category = request.form.get("category", "").strip()
        amount = parse_float(request.form.get("amount"))
        try:
            tenant_id = resolve_target_tenant_id(expense, required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle"
            )
            location = get_related_record(
                Location, parse_int(request.form.get("location_id")), tenant_id, "Location"
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(expense)
            context.update(
                {
                    "expense": expense,
                    "vehicles": vehicles,
                    "locations": locations,
                    "categories": categories,
                    "edit": True,
                }
            )
            return render_template("expenses/form.html", **context)

        if not category or amount <= 0:
            flash("Category and a valid amount are required.", "error")
            context = get_form_tenant_context(expense)
            context.update(
                {
                    "expense": expense,
                    "vehicles": vehicles,
                    "locations": locations,
                    "categories": categories,
                    "edit": True,
                }
            )
            return render_template("expenses/form.html", **context)

        expense.tenant_id = tenant_id
        expense.expense_date = parse_date(request.form.get("expense_date"), date.today())
        expense.category = category
        expense.description = request.form.get("description", "").strip() or None
        expense.amount = amount
        expense.vehicle_id = vehicle.id if vehicle else None
        expense.location_id = location.id if location else None
        expense.payment_method = request.form.get("payment_method", "Cash")
        expense.vendor_name = request.form.get("vendor_name", "").strip() or None
        expense.vendor_contact = request.form.get("vendor_contact", "").strip() or None
        expense.bill_number = request.form.get("bill_number", "").strip() or None
        expense.notes = request.form.get("notes", "").strip() or None
        expense.status = request.form.get("status", "Pending")
        db.session.commit()
        flash(f"Expense {expense.name} updated successfully.", "success")
        return redirect(url_for("expenses"))

    context = get_form_tenant_context(expense)
    context.update(
        {
            "expense": expense,
            "vehicles": vehicles,
            "locations": locations,
            "categories": categories,
            "edit": True,
            "today": date.today().isoformat(),
        }
    )
    return render_template("expenses/form.html", **context)


@app.route("/expenses/delete/<int:id>", methods=["POST"])
@permission_required("expenses", "delete")
def delete_expense(id):
    expense = get_scoped_record(Expense, id)
    name = expense.name
    db.session.delete(expense)
    db.session.commit()
    flash(f"Expense {name} deleted successfully.", "success")
    return redirect(url_for("expenses"))


@app.route("/loans")
@permission_required("loans", "view")
def loans():
    loan_query = scoped_query(Loan)
    loan_list = loan_query.order_by(Loan.created_at.desc()).all()
    total_principal = loan_query.with_entities(db.func.sum(Loan.principal_amount)).scalar() or 0
    total_balance = loan_query.with_entities(db.func.sum(Loan.balance_amount)).scalar() or 0
    return render_template(
        "loans/list.html",
        loans=loan_list,
        total_principal=total_principal,
        total_balance=total_balance,
        show_tenant_column=is_superadmin(),
    )


@app.route("/loans/create", methods=["GET", "POST"])
@permission_required("loans", "create")
def create_loan():
    selected_tenant_id = get_default_selected_tenant_id()
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    loan_types = ["Truck Loan", "Body Loan"]
    lender_types = ["Bank", "NBFC", "Private", "Individual"]

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle", True
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context()
            context.update(
                {
                    "vehicles": vehicles,
                    "loan_types": loan_types,
                    "lender_types": lender_types,
                    "today": date.today().isoformat(),
                }
            )
            return render_template("loans/form.html", **context)

        loan_type = request.form.get("loan_type", "").strip()
        if not loan_type:
            flash("Vehicle and loan type are required.", "error")
            context = get_form_tenant_context()
            context.update(
                {
                    "vehicles": vehicles,
                    "loan_types": loan_types,
                    "lender_types": lender_types,
                    "today": date.today().isoformat(),
                }
            )
            return render_template("loans/form.html", **context)

        principal = parse_float(request.form.get("principal_amount"))
        interest_rate = parse_float(request.form.get("interest_rate"))
        tenure = parse_int(request.form.get("tenure_months")) or 0
        down_payment = parse_float(request.form.get("down_payment"))
        monthly_rate = interest_rate / 12 / 100

        if monthly_rate > 0 and tenure > 0:
            emi = principal * monthly_rate * ((1 + monthly_rate) ** tenure) / (
                ((1 + monthly_rate) ** tenure) - 1
            )
            total_payable = emi * tenure
            total_interest = total_payable - principal
        else:
            emi = principal / tenure if tenure > 0 else 0
            total_payable = principal
            total_interest = 0

        balance = total_payable - down_payment

        loan = Loan(
            tenant_id=tenant_id,
            name=generate_loan_name(),
            loan_type=loan_type,
            vehicle_id=vehicle.id,
            principal_amount=principal,
            interest_rate=interest_rate,
            tenure_months=tenure,
            emi_amount=emi,
            total_payable=total_payable,
            total_interest=total_interest,
            lender_name=request.form.get("lender_name", "").strip() or None,
            lender_type=request.form.get("lender_type", "").strip() or None,
            lender_contact=request.form.get("lender_contact", "").strip() or None,
            lender_address=request.form.get("lender_address", "").strip() or None,
            agent_name=request.form.get("agent_name", "").strip() or None,
            agent_contact=request.form.get("agent_contact", "").strip() or None,
            loan_date=parse_date(request.form.get("loan_date"), date.today()),
            disbursement_date=parse_date(request.form.get("disbursement_date")),
            first_emi_date=parse_date(request.form.get("first_emi_date")),
            down_payment=down_payment,
            amount_paid=down_payment,
            balance_amount=balance,
            next_emi_due_date=parse_date(request.form.get("first_emi_date")),
            loan_account_number=request.form.get("loan_account_number", "").strip() or None,
            purpose=request.form.get("purpose", "").strip() or None,
            collateral=request.form.get("collateral", "").strip() or None,
            insurance_details=request.form.get("insurance_details", "").strip() or None,
            remarks=request.form.get("remarks", "").strip() or None,
            status=request.form.get("status", "Active"),
        )
        db.session.add(loan)
        db.session.commit()
        flash(f"Loan {loan.name} created successfully. EMI: Rs. {emi:.2f}", "success")
        return redirect(url_for("loans"))

    context = get_form_tenant_context()
    context.update(
        {
            "vehicles": vehicles,
            "loan_types": loan_types,
            "lender_types": lender_types,
            "today": date.today().isoformat(),
        }
    )
    return render_template("loans/form.html", **context)


@app.route("/loans/edit/<int:id>", methods=["GET", "POST"])
@permission_required("loans", "edit")
def edit_loan(id):
    loan = get_scoped_record(Loan, id)
    selected_tenant_id = get_default_selected_tenant_id(loan)
    vehicles = get_tenant_filtered_records(
        Vehicle, Vehicle.registration_number, selected_tenant_id
    )
    loan_types = ["Truck Loan", "Body Loan"]
    lender_types = ["Bank", "NBFC", "Private", "Individual"]

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(loan, required=True)
            vehicle = get_related_record(
                Vehicle, parse_int(request.form.get("vehicle_id")), tenant_id, "Vehicle", True
            )
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(loan)
            context.update(
                {
                    "loan": loan,
                    "vehicles": vehicles,
                    "loan_types": loan_types,
                    "lender_types": lender_types,
                    "edit": True,
                }
            )
            return render_template("loans/form.html", **context)

        principal = parse_float(request.form.get("principal_amount"))
        interest_rate = parse_float(request.form.get("interest_rate"))
        tenure = parse_int(request.form.get("tenure_months")) or 0
        down_payment = parse_float(request.form.get("down_payment"))
        monthly_rate = interest_rate / 12 / 100

        if monthly_rate > 0 and tenure > 0:
            emi = principal * monthly_rate * ((1 + monthly_rate) ** tenure) / (
                ((1 + monthly_rate) ** tenure) - 1
            )
            total_payable = emi * tenure
            total_interest = total_payable - principal
        else:
            emi = principal / tenure if tenure > 0 else 0
            total_payable = principal
            total_interest = 0

        paid_amount = loan.amount_paid or 0
        loan.tenant_id = tenant_id
        loan.loan_type = request.form.get("loan_type", "").strip()
        loan.vehicle_id = vehicle.id
        loan.principal_amount = principal
        loan.interest_rate = interest_rate
        loan.tenure_months = tenure
        loan.emi_amount = emi
        loan.total_payable = total_payable
        loan.total_interest = total_interest
        loan.down_payment = down_payment
        loan.amount_paid = max(paid_amount, down_payment)
        loan.balance_amount = total_payable - loan.amount_paid
        loan.lender_name = request.form.get("lender_name", "").strip() or None
        loan.lender_type = request.form.get("lender_type", "").strip() or None
        loan.lender_contact = request.form.get("lender_contact", "").strip() or None
        loan.lender_address = request.form.get("lender_address", "").strip() or None
        loan.agent_name = request.form.get("agent_name", "").strip() or None
        loan.agent_contact = request.form.get("agent_contact", "").strip() or None
        loan.loan_date = parse_date(request.form.get("loan_date"), loan.loan_date)
        loan.disbursement_date = parse_date(request.form.get("disbursement_date"))
        loan.first_emi_date = parse_date(request.form.get("first_emi_date"))
        loan.next_emi_due_date = parse_date(
            request.form.get("first_emi_date"), loan.next_emi_due_date
        )
        loan.loan_account_number = request.form.get("loan_account_number", "").strip() or None
        loan.purpose = request.form.get("purpose", "").strip() or None
        loan.collateral = request.form.get("collateral", "").strip() or None
        loan.insurance_details = request.form.get("insurance_details", "").strip() or None
        loan.remarks = request.form.get("remarks", "").strip() or None
        loan.status = request.form.get("status", "Active")
        db.session.commit()
        flash(f"Loan {loan.name} updated successfully.", "success")
        return redirect(url_for("loans"))

    context = get_form_tenant_context(loan)
    context.update(
        {
            "loan": loan,
            "vehicles": vehicles,
            "loan_types": loan_types,
            "lender_types": lender_types,
            "edit": True,
            "today": date.today().isoformat(),
        }
    )
    return render_template("loans/form.html", **context)


@app.route("/loans/delete/<int:id>", methods=["POST"])
@permission_required("loans", "delete")
def delete_loan(id):
    loan = get_scoped_record(Loan, id)
    name = loan.name
    db.session.delete(loan)
    db.session.commit()
    flash(f"Loan {name} deleted successfully.", "success")
    return redirect(url_for("loans"))


@app.route("/vendors")
@permission_required("vendors", "view")
def vendors():
    vendor_list = scoped_query(Vendor).order_by(Vendor.vendor_name).all()
    return render_template(
        "vendors/list.html",
        vendors=vendor_list,
        show_tenant_column=is_superadmin(),
    )


@app.route("/vendors/create", methods=["GET", "POST"])
@permission_required("vendors", "create")
def create_vendor():
    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context()
            context.update({"today": date.today().isoformat()})
            return render_template("vendors/form.html", **context)

        # Check for duplicate GSTIN - WARN but allow (same GST can have multiple addresses)
        gstin = request.form.get("gstin", "").strip() or None
        gstin_warning = None
        if gstin:
            existing_gstin = Vendor.query.filter_by(tenant_id=tenant_id, gstin=gstin).first()
            if existing_gstin:
                gstin_warning = f'GSTIN "{gstin}" already used by vendor {existing_gstin.vendor_code} ({existing_gstin.vendor_name}) - Same GST, different address allowed'

        vendor = Vendor(
            tenant_id=tenant_id,
            vendor_code=generate_vendor_code(),
            # Basic Information
            vendor_name=request.form.get("vendor_name", "").strip(),
            vendor_type=request.form.get("vendor_type", "supplier").strip(),
            status=request.form.get("status", "active").strip(),
            business_nature=request.form.get("business_nature", "").strip() or None,
            establishment_date=parse_date(request.form.get("establishment_date")),
            website=request.form.get("website", "").strip() or None,
            referral_source=request.form.get("referral_source", "").strip() or None,
            # Contact Information
            contact_person=request.form.get("contact_person", "").strip() or None,
            designation=request.form.get("designation", "").strip() or None,
            phone_primary=request.form.get("phone_primary", "").strip() or None,
            phone_secondary=request.form.get("phone_secondary", "").strip() or None,
            mobile=request.form.get("mobile", "").strip() or None,
            email=request.form.get("email", "").strip() or None,
            alternate_email=request.form.get("alternate_email", "").strip() or None,
            fax_number=request.form.get("fax_number", "").strip() or None,
            # Registered Address
            reg_address_line1=request.form.get("reg_address_line1", "").strip() or None,
            reg_address_line2=request.form.get("reg_address_line2", "").strip() or None,
            reg_city=request.form.get("reg_city", "").strip() or None,
            reg_state=request.form.get("reg_state", "").strip() or None,
            reg_pincode=request.form.get("reg_pincode", "").strip() or None,
            reg_country=request.form.get("reg_country", "India").strip() or "India",
            # Office Address
            office_address_line1=request.form.get("office_address_line1", "").strip() or None,
            office_address_line2=request.form.get("office_address_line2", "").strip() or None,
            office_city=request.form.get("office_city", "").strip() or None,
            office_state=request.form.get("office_state", "").strip() or None,
            office_pincode=request.form.get("office_pincode", "").strip() or None,
            office_country=request.form.get("office_country", "India").strip() or "India",
            same_as_registered=parse_bool(request.form.get("same_as_registered")),
            # GST & Tax Details
            gstin=request.form.get("gstin", "").strip() or None,
            gst_registration_date=parse_date(request.form.get("gst_registration_date")),
            gst_state_code=request.form.get("gst_state_code", "").strip() or None,
            pan_number=request.form.get("pan_number", "").strip() or None,
            tan_number=request.form.get("tan_number", "").strip() or None,
            tin_number=request.form.get("tin_number", "").strip() or None,
            cin_number=request.form.get("cin_number", "").strip() or None,
            msme_number=request.form.get("msme_number", "").strip() or None,
            tax_regime=request.form.get("tax_regime", "regular").strip() or "regular",
            is_composition_dealer=parse_bool(request.form.get("is_composition_dealer")),
            is_tds_applicable=parse_bool(request.form.get("is_tds_applicable")),
            tds_rate=parse_float(request.form.get("tds_rate")),
            # Bank Details
            bank_name=request.form.get("bank_name", "").strip() or None,
            bank_branch=request.form.get("bank_branch", "").strip() or None,
            account_number=request.form.get("account_number", "").strip() or None,
            account_type=request.form.get("account_type", "current").strip() or "current",
            ifsc_code=request.form.get("ifsc_code", "").strip() or None,
            micr_code=request.form.get("micr_code", "").strip() or None,
            swift_code=request.form.get("swift_code", "").strip() or None,
            bank_address=request.form.get("bank_address", "").strip() or None,
            upi_id=request.form.get("upi_id", "").strip() or None,
            # Primary Contact
            primary_contact_name=request.form.get("primary_contact_name", "").strip() or None,
            primary_contact_designation=request.form.get("primary_contact_designation", "").strip() or None,
            primary_contact_phone=request.form.get("primary_contact_phone", "").strip() or None,
            primary_contact_mobile=request.form.get("primary_contact_mobile", "").strip() or None,
            primary_contact_email=request.form.get("primary_contact_email", "").strip() or None,
            # Secondary Contact
            secondary_contact_name=request.form.get("secondary_contact_name", "").strip() or None,
            secondary_contact_designation=request.form.get("secondary_contact_designation", "").strip() or None,
            secondary_contact_phone=request.form.get("secondary_contact_phone", "").strip() or None,
            secondary_contact_mobile=request.form.get("secondary_contact_mobile", "").strip() or None,
            secondary_contact_email=request.form.get("secondary_contact_email", "").strip() or None,
            # Trade References
            trade_reference_1_name=request.form.get("trade_reference_1_name", "").strip() or None,
            trade_reference_1_contact=request.form.get("trade_reference_1_contact", "").strip() or None,
            trade_reference_1_address=request.form.get("trade_reference_1_address", "").strip() or None,
            trade_reference_2_name=request.form.get("trade_reference_2_name", "").strip() or None,
            trade_reference_2_contact=request.form.get("trade_reference_2_contact", "").strip() or None,
            trade_reference_2_address=request.form.get("trade_reference_2_address", "").strip() or None,
            # Financial Information
            credit_limit=parse_float(request.form.get("credit_limit")),
            credit_period_days=parse_int(request.form.get("credit_period_days")) or 0,
            opening_balance=parse_float(request.form.get("opening_balance")),
            balance_type=request.form.get("balance_type", "dr").strip() or "dr",
            currency=request.form.get("currency", "INR").strip() or "INR",
            payment_terms=request.form.get("payment_terms", "").strip() or None,
            # Supply Details
            supply_type=request.form.get("supply_type", "").strip() or None,
            product_categories=request.form.get("product_categories", "").strip() or None,
            lead_time_days=parse_int(request.form.get("lead_time_days")) or 0,
            min_order_value=parse_float(request.form.get("min_order_value")),
            max_order_value=parse_float(request.form.get("max_order_value")),
            delivery_mode=request.form.get("delivery_mode", "").strip() or None,
            # Compliance
            kyc_status=request.form.get("kyc_status", "pending").strip() or "pending",
            verification_date=parse_date(request.form.get("verification_date")),
            verified_by=request.form.get("verified_by", "").strip() or None,
            compliance_rating=request.form.get("compliance_rating", "unrated").strip() or "unrated",
            background_check_done=parse_bool(request.form.get("background_check_done")),
            background_check_date=parse_date(request.form.get("background_check_date")),
            # System Fields
            created_by=getattr(g, "current_user", None).username if getattr(g, "current_user", None) else None,
            remarks=request.form.get("remarks", "").strip() or None,
        )
        db.session.add(vendor)
        db.session.commit()
        
        # Show success message and GSTIN warning if applicable
        if gstin_warning:
            flash(f"Vendor {vendor.vendor_name} created successfully.", "success")
            flash(gstin_warning, "warning")
        else:
            flash(f"Vendor {vendor.vendor_name} created successfully.", "success")
        
        return redirect(url_for("vendors"))

    context = get_form_tenant_context()
    context.update({"today": date.today().isoformat()})
    return render_template("vendors/form.html", **context)


@app.route("/vendors/edit/<int:id>", methods=["GET", "POST"])
@permission_required("vendors", "edit")
def edit_vendor(id):
    vendor = get_scoped_record(Vendor, id)

    if request.method == "POST":
        try:
            tenant_id = resolve_target_tenant_id(vendor, required=True)
        except ValueError as exc:
            flash(str(exc), "error")
            context = get_form_tenant_context(vendor)
            context.update({"vendor": vendor, "edit": True})
            return render_template("vendors/form.html", **context)

        vendor.tenant_id = tenant_id
        # Basic Information
        vendor.vendor_name = request.form.get("vendor_name", "").strip()
        vendor.vendor_type = request.form.get("vendor_type", "supplier").strip()
        vendor.status = request.form.get("status", "active").strip()
        vendor.business_nature = request.form.get("business_nature", "").strip() or None
        vendor.establishment_date = parse_date(request.form.get("establishment_date"))
        vendor.website = request.form.get("website", "").strip() or None
        vendor.referral_source = request.form.get("referral_source", "").strip() or None
        # Contact Information
        vendor.contact_person = request.form.get("contact_person", "").strip() or None
        vendor.designation = request.form.get("designation", "").strip() or None
        vendor.phone_primary = request.form.get("phone_primary", "").strip() or None
        vendor.phone_secondary = request.form.get("phone_secondary", "").strip() or None
        vendor.mobile = request.form.get("mobile", "").strip() or None
        vendor.email = request.form.get("email", "").strip() or None
        vendor.alternate_email = request.form.get("alternate_email", "").strip() or None
        vendor.fax_number = request.form.get("fax_number", "").strip() or None
        # Registered Address
        vendor.reg_address_line1 = request.form.get("reg_address_line1", "").strip() or None
        vendor.reg_address_line2 = request.form.get("reg_address_line2", "").strip() or None
        vendor.reg_city = request.form.get("reg_city", "").strip() or None
        vendor.reg_state = request.form.get("reg_state", "").strip() or None
        vendor.reg_pincode = request.form.get("reg_pincode", "").strip() or None
        vendor.reg_country = request.form.get("reg_country", "India").strip() or "India"
        # Office Address
        vendor.office_address_line1 = request.form.get("office_address_line1", "").strip() or None
        vendor.office_address_line2 = request.form.get("office_address_line2", "").strip() or None
        vendor.office_city = request.form.get("office_city", "").strip() or None
        vendor.office_state = request.form.get("office_state", "").strip() or None
        vendor.office_pincode = request.form.get("office_pincode", "").strip() or None
        vendor.office_country = request.form.get("office_country", "India").strip() or "India"
        vendor.same_as_registered = parse_bool(request.form.get("same_as_registered"))
        # GST & Tax Details
        vendor.gstin = request.form.get("gstin", "").strip() or None
        vendor.gst_registration_date = parse_date(request.form.get("gst_registration_date"))
        vendor.gst_state_code = request.form.get("gst_state_code", "").strip() or None
        vendor.pan_number = request.form.get("pan_number", "").strip() or None
        vendor.tan_number = request.form.get("tan_number", "").strip() or None
        vendor.tin_number = request.form.get("tin_number", "").strip() or None
        vendor.cin_number = request.form.get("cin_number", "").strip() or None
        vendor.msme_number = request.form.get("msme_number", "").strip() or None
        vendor.tax_regime = request.form.get("tax_regime", "regular").strip() or "regular"
        vendor.is_composition_dealer = parse_bool(request.form.get("is_composition_dealer"))
        vendor.is_tds_applicable = parse_bool(request.form.get("is_tds_applicable"))
        vendor.tds_rate = parse_float(request.form.get("tds_rate"))
        # Bank Details
        vendor.bank_name = request.form.get("bank_name", "").strip() or None
        vendor.bank_branch = request.form.get("bank_branch", "").strip() or None
        vendor.account_number = request.form.get("account_number", "").strip() or None
        vendor.account_type = request.form.get("account_type", "current").strip() or "current"
        vendor.ifsc_code = request.form.get("ifsc_code", "").strip() or None
        vendor.micr_code = request.form.get("micr_code", "").strip() or None
        vendor.swift_code = request.form.get("swift_code", "").strip() or None
        vendor.bank_address = request.form.get("bank_address", "").strip() or None
        vendor.upi_id = request.form.get("upi_id", "").strip() or None
        # Primary Contact
        vendor.primary_contact_name = request.form.get("primary_contact_name", "").strip() or None
        vendor.primary_contact_designation = request.form.get("primary_contact_designation", "").strip() or None
        vendor.primary_contact_phone = request.form.get("primary_contact_phone", "").strip() or None
        vendor.primary_contact_mobile = request.form.get("primary_contact_mobile", "").strip() or None
        vendor.primary_contact_email = request.form.get("primary_contact_email", "").strip() or None
        # Secondary Contact
        vendor.secondary_contact_name = request.form.get("secondary_contact_name", "").strip() or None
        vendor.secondary_contact_designation = request.form.get("secondary_contact_designation", "").strip() or None
        vendor.secondary_contact_phone = request.form.get("secondary_contact_phone", "").strip() or None
        vendor.secondary_contact_mobile = request.form.get("secondary_contact_mobile", "").strip() or None
        vendor.secondary_contact_email = request.form.get("secondary_contact_email", "").strip() or None
        # Trade References
        vendor.trade_reference_1_name = request.form.get("trade_reference_1_name", "").strip() or None
        vendor.trade_reference_1_contact = request.form.get("trade_reference_1_contact", "").strip() or None
        vendor.trade_reference_1_address = request.form.get("trade_reference_1_address", "").strip() or None
        vendor.trade_reference_2_name = request.form.get("trade_reference_2_name", "").strip() or None
        vendor.trade_reference_2_contact = request.form.get("trade_reference_2_contact", "").strip() or None
        vendor.trade_reference_2_address = request.form.get("trade_reference_2_address", "").strip() or None
        # Financial Information
        vendor.credit_limit = parse_float(request.form.get("credit_limit"))
        vendor.credit_period_days = parse_int(request.form.get("credit_period_days")) or 0
        vendor.opening_balance = parse_float(request.form.get("opening_balance"))
        vendor.balance_type = request.form.get("balance_type", "dr").strip() or "dr"
        vendor.currency = request.form.get("currency", "INR").strip() or "INR"
        vendor.payment_terms = request.form.get("payment_terms", "").strip() or None
        # Supply Details
        vendor.supply_type = request.form.get("supply_type", "").strip() or None
        vendor.product_categories = request.form.get("product_categories", "").strip() or None
        vendor.lead_time_days = parse_int(request.form.get("lead_time_days")) or 0
        vendor.min_order_value = parse_float(request.form.get("min_order_value"))
        vendor.max_order_value = parse_float(request.form.get("max_order_value"))
        vendor.delivery_mode = request.form.get("delivery_mode", "").strip() or None
        # Compliance
        vendor.kyc_status = request.form.get("kyc_status", "pending").strip() or "pending"
        vendor.verification_date = parse_date(request.form.get("verification_date"))
        vendor.verified_by = request.form.get("verified_by", "").strip() or None
        vendor.compliance_rating = request.form.get("compliance_rating", "unrated").strip() or "unrated"
        vendor.background_check_done = parse_bool(request.form.get("background_check_done"))
        vendor.background_check_date = parse_date(request.form.get("background_check_date"))
        # System Fields
        vendor.modified_by = getattr(g, "current_user", None).username if getattr(g, "current_user", None) else None
        vendor.remarks = request.form.get("remarks", "").strip() or None

        db.session.commit()
        flash(f"Vendor {vendor.vendor_name} updated successfully.", "success")
        return redirect(url_for("vendors"))

    context = get_form_tenant_context(vendor)
    context.update({"vendor": vendor, "edit": True})
    return render_template("vendors/form.html", **context)


@app.route("/vendors/delete/<int:id>", methods=["POST"])
@permission_required("vendors", "delete")
def delete_vendor(id):
    vendor = get_scoped_record(Vendor, id)
    name = vendor.vendor_name
    db.session.delete(vendor)
    db.session.commit()
    flash(f"Vendor {name} deleted successfully.", "success")
    return redirect(url_for("vendors"))


# =============================================================================
# Vendor User Login Management Routes
# =============================================================================

@app.route("/vendors/<int:vendor_id>/create-login", methods=["GET", "POST"])
@permission_required("vendors", "edit")
def create_vendor_login(vendor_id):
    """Create a vendor login user with field-level permissions"""
    vendor = get_scoped_record(Vendor, vendor_id)
    
    # Define vendor field categories and fields
    VENDOR_FIELDS = {
        'basic': [
            {'name': 'vendor_code', 'label': 'Vendor Code', 'required': True},
            {'name': 'vendor_name', 'label': 'Vendor Name', 'required': True},
            {'name': 'vendor_type', 'label': 'Vendor Type', 'required': False},
            {'name': 'status', 'label': 'Status', 'required': False},
            {'name': 'business_nature', 'label': 'Business Nature', 'required': False},
            {'name': 'establishment_date', 'label': 'Establishment Date', 'required': False},
            {'name': 'website', 'label': 'Website', 'required': False},
        ],
        'contact': [
            {'name': 'contact_person', 'label': 'Contact Person', 'required': False},
            {'name': 'designation', 'label': 'Designation', 'required': False},
            {'name': 'phone_primary', 'label': 'Primary Phone', 'required': False},
            {'name': 'phone_secondary', 'label': 'Secondary Phone', 'required': False},
            {'name': 'mobile', 'label': 'Mobile', 'required': False},
            {'name': 'email', 'label': 'Email', 'required': False},
            {'name': 'alternate_email', 'label': 'Alternate Email', 'required': False},
            {'name': 'fax_number', 'label': 'Fax Number', 'required': False},
        ],
        'address': [
            {'name': 'reg_address_line1', 'label': 'Registered Address Line 1', 'required': False},
            {'name': 'reg_address_line2', 'label': 'Registered Address Line 2', 'required': False},
            {'name': 'reg_city', 'label': 'Registered City', 'required': False},
            {'name': 'reg_state', 'label': 'Registered State', 'required': False},
            {'name': 'reg_pincode', 'label': 'Registered Pincode', 'required': False},
            {'name': 'reg_country', 'label': 'Registered Country', 'required': False},
            {'name': 'office_address_line1', 'label': 'Office Address Line 1', 'required': False},
            {'name': 'office_address_line2', 'label': 'Office Address Line 2', 'required': False},
            {'name': 'office_city', 'label': 'Office City', 'required': False},
            {'name': 'office_state', 'label': 'Office State', 'required': False},
            {'name': 'office_pincode', 'label': 'Office Pincode', 'required': False},
            {'name': 'office_country', 'label': 'Office Country', 'required': False},
        ],
        'gst': [
            {'name': 'gstin', 'label': 'GSTIN', 'required': False},
            {'name': 'pan_number', 'label': 'PAN Number', 'required': False},
            {'name': 'tan_number', 'label': 'TAN Number', 'required': False},
            {'name': 'gst_registration_type', 'label': 'GST Registration Type', 'required': False},
            {'name': 'gst_compliance_status', 'label': 'GST Compliance Status', 'required': False},
        ],
        'financial': [
            {'name': 'credit_limit', 'label': 'Credit Limit', 'required': False},
            {'name': 'credit_period_days', 'label': 'Credit Period (Days)', 'required': False},
            {'name': 'opening_balance', 'label': 'Opening Balance', 'required': False},
            {'name': 'balance_type', 'label': 'Balance Type', 'required': False},
            {'name': 'currency', 'label': 'Currency', 'required': False},
            {'name': 'payment_terms', 'label': 'Payment Terms', 'required': False},
        ],
        'bank': [
            {'name': 'bank_name', 'label': 'Bank Name', 'required': False},
            {'name': 'bank_account_number', 'label': 'Bank Account Number', 'required': False},
            {'name': 'bank_account_type', 'label': 'Bank Account Type', 'required': False},
            {'name': 'bank_ifsc_code', 'label': 'Bank IFSC Code', 'required': False},
            {'name': 'bank_branch', 'label': 'Bank Branch', 'required': False},
        ],
    }
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        full_name = request.form.get("full_name", "").strip()
        password = request.form.get("password", "")
        is_primary = request.form.get("is_primary") == "on"
        can_view_own_data = request.form.get("can_view_own_data") == "on"
        can_edit_own_data = request.form.get("can_edit_own_data") == "on"
        
        if not all([username, email, full_name, password]):
            flash("Username, email, full name, and password are required.", "error")
            return render_template("vendors/vendor_login_form.html", 
                                  vendor=vendor, 
                                  vendor_fields=VENDOR_FIELDS)
        
        # Check if username or email already exists
        if User.query.filter_by(username=username).first():
            flash("Username already exists.", "error")
            return render_template("vendors/vendor_login_form.html", 
                                  vendor=vendor, 
                                  vendor_fields=VENDOR_FIELDS)
        
        if User.query.filter_by(email=email).first():
            flash("Email already exists.", "error")
            return render_template("vendors/vendor_login_form.html", 
                                  vendor=vendor, 
                                  vendor_fields=VENDOR_FIELDS)
        
        try:
            # Create user with vendor role
            user = User(
                username=username,
                email=email,
                full_name=full_name,
                role="vendor",
                tenant_id=vendor.tenant_id,
                is_active=True
            )
            user.set_password(password)
            db.session.add(user)
            db.session.flush()  # Get user ID
            
            # Create vendor user association
            vendor_user = VendorUser(
                vendor_id=vendor_id,
                user_id=user.id,
                tenant_id=vendor.tenant_id,
                is_primary=is_primary,
                can_view_own_data=can_view_own_data,
                can_edit_own_data=can_edit_own_data,
                created_by=g.current_user.id
            )
            db.session.add(vendor_user)
            db.session.flush()  # Get vendor_user ID
            
            # Create field permissions based on form submission
            for category, fields in VENDOR_FIELDS.items():
                for field in fields:
                    field_name = field['name']
                    can_view = request.form.get(f"view_{field_name}") == "on"
                    can_edit = request.form.get(f"edit_{field_name}") == "on"
                    is_required = request.form.get(f"required_{field_name}") == "on"
                    display_order = int(request.form.get(f"order_{field_name}", 0))
                    
                    permission = VendorFieldPermission(
                        vendor_user_id=vendor_user.id,
                        field_name=field_name,
                        field_category=category,
                        can_view=can_view,
                        can_edit=can_edit,
                        is_required=is_required,
                        display_order=display_order
                    )
                    db.session.add(permission)
            
            db.session.commit()
            flash(f"Vendor login created for {full_name}.", "success")
            return redirect(url_for("edit_vendor", id=vendor_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error creating vendor login: {str(e)}", "error")
    
    return render_template("vendors/vendor_login_form.html", 
                          vendor=vendor, 
                          vendor_fields=VENDOR_FIELDS)


@app.route("/vendor-users/<int:id>/edit", methods=["GET", "POST"])
@permission_required("vendors", "edit")
def edit_vendor_user(id):
    """Edit vendor user permissions"""
    vendor_user = VendorUser.query.get_or_404(id)
    
    # Check access - only superadmin or tenant admin of same tenant can edit
    if not is_superadmin() and vendor_user.tenant_id != g.current_user.tenant_id:
        flash("You don't have permission to edit this vendor user.", "error")
        return redirect(url_for("vendors"))
    
    # Define vendor field categories and fields
    VENDOR_FIELDS = {
        'basic': [
            {'name': 'vendor_code', 'label': 'Vendor Code', 'required': True},
            {'name': 'vendor_name', 'label': 'Vendor Name', 'required': True},
            {'name': 'vendor_type', 'label': 'Vendor Type', 'required': False},
            {'name': 'status', 'label': 'Status', 'required': False},
            {'name': 'business_nature', 'label': 'Business Nature', 'required': False},
            {'name': 'establishment_date', 'label': 'Establishment Date', 'required': False},
            {'name': 'website', 'label': 'Website', 'required': False},
        ],
        'contact': [
            {'name': 'contact_person', 'label': 'Contact Person', 'required': False},
            {'name': 'designation', 'label': 'Designation', 'required': False},
            {'name': 'phone_primary', 'label': 'Primary Phone', 'required': False},
            {'name': 'phone_secondary', 'label': 'Secondary Phone', 'required': False},
            {'name': 'mobile', 'label': 'Mobile', 'required': False},
            {'name': 'email', 'label': 'Email', 'required': False},
            {'name': 'alternate_email', 'label': 'Alternate Email', 'required': False},
            {'name': 'fax_number', 'label': 'Fax Number', 'required': False},
        ],
        'address': [
            {'name': 'reg_address_line1', 'label': 'Registered Address Line 1', 'required': False},
            {'name': 'reg_address_line2', 'label': 'Registered Address Line 2', 'required': False},
            {'name': 'reg_city', 'label': 'Registered City', 'required': False},
            {'name': 'reg_state', 'label': 'Registered State', 'required': False},
            {'name': 'reg_pincode', 'label': 'Registered Pincode', 'required': False},
            {'name': 'reg_country', 'label': 'Registered Country', 'required': False},
            {'name': 'office_address_line1', 'label': 'Office Address Line 1', 'required': False},
            {'name': 'office_address_line2', 'label': 'Office Address Line 2', 'required': False},
            {'name': 'office_city', 'label': 'Office City', 'required': False},
            {'name': 'office_state', 'label': 'Office State', 'required': False},
            {'name': 'office_pincode', 'label': 'Office Pincode', 'required': False},
            {'name': 'office_country', 'label': 'Office Country', 'required': False},
        ],
        'gst': [
            {'name': 'gstin', 'label': 'GSTIN', 'required': False},
            {'name': 'pan_number', 'label': 'PAN Number', 'required': False},
            {'name': 'tan_number', 'label': 'TAN Number', 'required': False},
            {'name': 'gst_registration_type', 'label': 'GST Registration Type', 'required': False},
            {'name': 'gst_compliance_status', 'label': 'GST Compliance Status', 'required': False},
        ],
        'financial': [
            {'name': 'credit_limit', 'label': 'Credit Limit', 'required': False},
            {'name': 'credit_period_days', 'label': 'Credit Period (Days)', 'required': False},
            {'name': 'opening_balance', 'label': 'Opening Balance', 'required': False},
            {'name': 'balance_type', 'label': 'Balance Type', 'required': False},
            {'name': 'currency', 'label': 'Currency', 'required': False},
            {'name': 'payment_terms', 'label': 'Payment Terms', 'required': False},
        ],
        'bank': [
            {'name': 'bank_name', 'label': 'Bank Name', 'required': False},
            {'name': 'bank_account_number', 'label': 'Bank Account Number', 'required': False},
            {'name': 'bank_account_type', 'label': 'Bank Account Type', 'required': False},
            {'name': 'bank_ifsc_code', 'label': 'Bank IFSC Code', 'required': False},
            {'name': 'bank_branch', 'label': 'Bank Branch', 'required': False},
        ],
    }
    
    # Get existing permissions
    existing_permissions = {
        perm.field_name: perm 
        for perm in vendor_user.field_permissions
    }
    
    if request.method == "POST":
        # Update vendor user settings
        vendor_user.is_primary = request.form.get("is_primary") == "on"
        vendor_user.can_view_own_data = request.form.get("can_view_own_data") == "on"
        vendor_user.can_edit_own_data = request.form.get("can_edit_own_data") == "on"
        
        # Update user details if provided
        new_password = request.form.get("password", "").strip()
        if new_password:
            vendor_user.user.set_password(new_password)
        
        vendor_user.user.is_active = request.form.get("is_active") == "on"
        
        # Update field permissions
        for category, fields in VENDOR_FIELDS.items():
            for field in fields:
                field_name = field['name']
                can_view = request.form.get(f"view_{field_name}") == "on"
                can_edit = request.form.get(f"edit_{field_name}") == "on"
                is_required = request.form.get(f"required_{field_name}") == "on"
                display_order = int(request.form.get(f"order_{field_name}", 0))
                
                if field_name in existing_permissions:
                    # Update existing permission
                    perm = existing_permissions[field_name]
                    perm.can_view = can_view
                    perm.can_edit = can_edit
                    perm.is_required = is_required
                    perm.display_order = display_order
                else:
                    # Create new permission
                    permission = VendorFieldPermission(
                        vendor_user_id=vendor_user.id,
                        field_name=field_name,
                        field_category=category,
                        can_view=can_view,
                        can_edit=can_edit,
                        is_required=is_required,
                        display_order=display_order
                    )
                    db.session.add(permission)
        
        db.session.commit()
        flash("Vendor user permissions updated successfully.", "success")
        return redirect(url_for("edit_vendor", id=vendor_user.vendor_id))
    
    return render_template("vendors/vendor_login_form.html", 
                          vendor=vendor_user.vendor,
                          vendor_user=vendor_user,
                          vendor_fields=VENDOR_FIELDS,
                          existing_permissions=existing_permissions,
                          edit=True)


@app.route("/vendor-users/<int:id>/delete", methods=["POST"])
@permission_required("vendors", "delete")
def delete_vendor_user(id):
    """Delete a vendor user login"""
    vendor_user = VendorUser.query.get_or_404(id)
    
    # Check access
    if not is_superadmin() and vendor_user.tenant_id != g.current_user.tenant_id:
        flash("You don't have permission to delete this vendor user.", "error")
        return redirect(url_for("vendors"))
    
    vendor_id = vendor_user.vendor_id
    user = vendor_user.user
    
    # Delete field permissions first
    VendorFieldPermission.query.filter_by(vendor_user_id=id).delete()
    
    # Delete vendor user association
    db.session.delete(vendor_user)
    
    # Delete the user
    db.session.delete(user)
    
    db.session.commit()
    flash("Vendor login deleted successfully.", "success")
    return redirect(url_for("edit_vendor", id=vendor_id))


# =============================================================================
# Vendor Address Management Routes
# =============================================================================

@app.route("/vendors/<int:vendor_id>/addresses/add", methods=["POST"])
@permission_required("vendors", "edit")
def add_vendor_address(vendor_id):
    """Add a new address to a vendor"""
    vendor = get_scoped_record(Vendor, vendor_id)
    
    try:
        # If this is the first address, make it primary
        is_primary = vendor.addresses.filter_by(is_active=True).count() == 0
        
        address = VendorAddress(
            tenant_id=get_current_tenant_id(),
            vendor_id=vendor_id,
            address_type=request.form.get("address_type", "Office"),
            address_line1=request.form.get("address_line1", "").strip(),
            address_line2=request.form.get("address_line2", "").strip() or None,
            city=request.form.get("city", "").strip(),
            state=request.form.get("state", "").strip(),
            pincode=request.form.get("pincode", "").strip(),
            country=request.form.get("country", "India").strip(),
            contact_person=request.form.get("contact_person", "").strip() or None,
            phone=request.form.get("phone", "").strip() or None,
            email=request.form.get("email", "").strip() or None,
            is_primary=is_primary,
            is_active=True
        )
        
        db.session.add(address)
        db.session.commit()
        flash("Address added successfully.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error adding address: {str(e)}", "error")
    
    return redirect(url_for("edit_vendor", id=vendor_id))


@app.route("/vendors/<int:vendor_id>/addresses/<int:address_id>/update", methods=["POST"])
@permission_required("vendors", "edit")
def update_vendor_address(vendor_id, address_id):
    """Update a vendor address"""
    vendor = get_scoped_record(Vendor, vendor_id)
    address = VendorAddress.query.filter_by(id=address_id, vendor_id=vendor_id, tenant_id=get_current_tenant_id()).first_or_404()
    
    try:
        address.address_type = request.form.get("address_type", address.address_type)
        address.address_line1 = request.form.get("address_line1", "").strip()
        address.address_line2 = request.form.get("address_line2", "").strip() or None
        address.city = request.form.get("city", "").strip()
        address.state = request.form.get("state", "").strip()
        address.pincode = request.form.get("pincode", "").strip()
        address.country = request.form.get("country", "India").strip()
        address.contact_person = request.form.get("contact_person", "").strip() or None
        address.phone = request.form.get("phone", "").strip() or None
        address.email = request.form.get("email", "").strip() or None
        
        # Handle primary address logic
        is_primary = request.form.get("is_primary") == "on"
        if is_primary and not address.is_primary:
            # Unset other primary addresses
            VendorAddress.query.filter_by(vendor_id=vendor_id, is_primary=True).update({"is_primary": False})
        address.is_primary = is_primary
        
        db.session.commit()
        flash("Address updated successfully.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating address: {str(e)}", "error")
    
    return redirect(url_for("edit_vendor", id=vendor_id))


@app.route("/vendors/<int:vendor_id>/addresses/<int:address_id>/delete", methods=["POST"])
@permission_required("vendors", "edit")
def delete_vendor_address(vendor_id, address_id):
    """Delete a vendor address"""
    vendor = get_scoped_record(Vendor, vendor_id)
    address = VendorAddress.query.filter_by(id=address_id, vendor_id=vendor_id, tenant_id=get_current_tenant_id()).first_or_404()
    
    try:
        was_primary = address.is_primary
        db.session.delete(address)
        
        # If we deleted the primary address, make another one primary
        if was_primary:
            next_address = VendorAddress.query.filter_by(vendor_id=vendor_id, is_active=True).first()
            if next_address:
                next_address.is_primary = True
        
        db.session.commit()
        flash("Address deleted successfully.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting address: {str(e)}", "error")
    
    return redirect(url_for("edit_vendor", id=vendor_id))


@app.route("/api/vendors/<int:vendor_id>/addresses")
def get_vendor_addresses_api(vendor_id):
    """API endpoint to get all addresses for a vendor (for dropdowns)"""
    try:
        vendor = get_scoped_record(Vendor, vendor_id)
        addresses = vendor.addresses.filter_by(is_active=True).all()
        
        result = []
        
        # Add VendorAddress records
        for addr in addresses:
            result.append({
                "id": addr.id,
                "address_type": addr.address_type,
                "address_line1": addr.address_line1,
                "address_line2": addr.address_line2 or "",
                "city": addr.city,
                "state": addr.state,
                "pincode": addr.pincode,
                "country": addr.country,
                "full_address": f"{addr.address_line1}, {addr.city}, {addr.state} - {addr.pincode}",
                "is_primary": addr.is_primary,
                "contact_person": addr.contact_person or "",
                "phone": addr.phone or "",
                "email": addr.email or ""
            })
        
        # If no VendorAddress records exist, add vendor's own addresses
        if not result:
            # Add Registered Address
            if vendor.reg_address_line1:
                reg_full = f"{vendor.reg_address_line1}"
                if vendor.reg_address_line2:
                    reg_full += f", {vendor.reg_address_line2}"
                reg_full += f", {vendor.reg_city}, {vendor.reg_state} - {vendor.reg_pincode}"
                
                result.append({
                    "id": "reg",
                    "address_type": "Registered",
                    "address_line1": vendor.reg_address_line1,
                    "address_line2": vendor.reg_address_line2 or "",
                    "city": vendor.reg_city,
                    "state": vendor.reg_state,
                    "pincode": vendor.reg_pincode,
                    "country": vendor.reg_country,
                    "full_address": reg_full,
                    "is_primary": True,
                    "contact_person": vendor.contact_person or "",
                    "phone": vendor.phone_primary or vendor.mobile or "",
                    "email": vendor.email or ""
                })
            
            # Add Office Address (if different from registered)
            if vendor.office_address_line1 and not vendor.same_as_registered:
                office_full = f"{vendor.office_address_line1}"
                if vendor.office_address_line2:
                    office_full += f", {vendor.office_address_line2}"
                office_full += f", {vendor.office_city}, {vendor.office_state} - {vendor.office_pincode}"
                
                result.append({
                    "id": "office",
                    "address_type": "Office",
                    "address_line1": vendor.office_address_line1,
                    "address_line2": vendor.office_address_line2 or "",
                    "city": vendor.office_city,
                    "state": vendor.office_state,
                    "pincode": vendor.office_pincode,
                    "country": vendor.office_country,
                    "full_address": office_full,
                    "is_primary": False,
                    "contact_person": vendor.contact_person or "",
                    "phone": vendor.phone_primary or vendor.mobile or "",
                    "email": vendor.email or ""
                })
        
        print(f"API: Returning {len(result)} addresses for vendor {vendor_id}")
        return jsonify(result)
    except Exception as e:
        print(f"API Error: {str(e)}")
        return jsonify({"error": str(e)}), 500


# =============================================================================
# Flask CLI Commands for Data Import
# =============================================================================

@app.cli.command("import-excel")
def import_excel_command():
    """Import data from Excel files (Vendor Details.xlsx, Vehicle and Loan Details.xlsx, ALL PARTY RATE FILE.xlsx)"""
    import os
    from datetime import datetime
    from openpyxl import load_workbook

    def safe_str(value, max_len=None):
        if value is None:
            return None
        result = str(value).strip()
        if max_len and len(result) > max_len:
            result = result[:max_len]
        return result if result else None

    def safe_float(value):
        if value is None:
            return 0.0
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace('₹', '').strip()
            return float(value)
        except:
            return 0.0

    def parse_int(value):
        if value is None:
            return None
        try:
            return int(float(value))
        except:
            return None

    def parse_date(date_value):
        if not date_value:
            return None
        if isinstance(date_value, datetime):
            return date_value.date() if hasattr(date_value, 'date') else date_value
        if isinstance(date_value, str):
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(date_value, fmt).date()
                except:
                    continue
        return None

    def get_or_create_tenant():
        tenant = Tenant.query.first()
        if not tenant:
            tenant = Tenant(name="Default Tenant", subdomain="default")
            db.session.add(tenant)
            db.session.commit()
            print(f"Created tenant: {tenant.name} (ID: {tenant.id})")
        return tenant

    def import_vendors(filepath, tenant_id):
        print(f"\n{'='*60}")
        print(f"Importing Vendors from: {filepath}")
        print(f"{'='*60}")

        wb = load_workbook(filepath)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        print(f"Total rows: {ws.max_row - 1}")

        imported = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))

                vendor_name = safe_str(data.get('VENDOR NAME') or data.get('Company Name') or data.get('Vendor Name'))
                if not vendor_name:
                    continue

                existing = Vendor.query.filter_by(tenant_id=tenant_id, vendor_name=vendor_name).first()
                if existing:
                    print(f"  Skipping (exists): {vendor_name}")
                    continue

                vendor_code = generate_vendor_code(tenant_id)

                vendor = Vendor(
                    tenant_id=tenant_id,
                    vendor_code=vendor_code,
                    vendor_name=vendor_name,
                    vendor_type=safe_str(data.get('TYPE') or data.get('Vendor Type'), 50) or 'supplier',
                    status='active',
                    contact_person=safe_str(data.get('CONTACT PERSON') or data.get('Contact Person'), 100),
                    mobile=safe_str(data.get('MOBILE') or data.get('Mobile') or data.get('PHONE'), 20),
                    email=safe_str(data.get('EMAIL') or data.get('Email'), 150),
                    reg_city=safe_str(data.get('CITY') or data.get('City'), 100),
                    reg_state=safe_str(data.get('STATE') or data.get('State'), 100),
                    gstin=safe_str(data.get('GSTIN') or data.get('GST'), 20),
                    pan=safe_str(data.get('PAN'), 20),
                )

                db.session.add(vendor)
                imported += 1
                print(f"  Imported: {vendor_name} (Code: {vendor_code})")

            except Exception as e:
                print(f"  ERROR Row {row_idx}: {e}")

        db.session.commit()
        print(f"\nVendors imported: {imported}")
        return imported

    def import_vehicles(filepath, tenant_id):
        print(f"\n{'='*60}")
        print(f"Importing Vehicles from: {filepath}")
        print(f"{'='*60}")

        wb = load_workbook(filepath)
        vehicle_sheet = None

        for sheet_name in wb.sheetnames:
            if 'vehicle' in sheet_name.lower():
                vehicle_sheet = wb[sheet_name]
                break

        if not vehicle_sheet:
            vehicle_sheet = wb[wb.sheetnames[0]]

        headers = [cell.value for cell in vehicle_sheet[1]]
        print(f"Sheet: {vehicle_sheet.title}")
        print(f"Headers: {headers}")

        imported = 0

        for row_idx, row in enumerate(vehicle_sheet.iter_rows(min_row=2, max_row=vehicle_sheet.max_row, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))

                reg_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number') or data.get('Registration Number'))
                if not reg_number:
                    continue

                existing = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=reg_number).first()
                if existing:
                    print(f"  Skipping (exists): {reg_number}")
                    continue

                vehicle = Vehicle(
                    tenant_id=tenant_id,
                    registration_number=reg_number,
                    vehicle_type=safe_str(data.get('TYPE') or data.get('Type'), 50) or 'Truck',
                    make=safe_str(data.get('MAKE') or data.get('Make'), 100),
                    model=safe_str(data.get('MODEL') or data.get('Model'), 100),
                    year=parse_int(data.get('YEAR') or data.get('Year')),
                    owner_name=safe_str(data.get('OWNER NAME') or data.get('Owner Name'), 200),
                    owner_contact=safe_str(data.get('OWNER CONTACT') or data.get('Owner Contact'), 50),
                    load_capacity=safe_str(data.get('LOAD CAPACITY') or data.get('Truck Size'), 50),
                    insurance_expiry=parse_date(data.get('INSURANCE EXPIRY') or data.get('Insurance Expiry')),
                    fitness_expiry=parse_date(data.get('FITNESS EXPIRY') or data.get('Fitness Expiry')),
                    status='active',
                )

                db.session.add(vehicle)
                imported += 1
                print(f"  Imported: {reg_number}")

            except Exception as e:
                print(f"  ERROR Row {row_idx}: {e}")

        db.session.commit()
        print(f"\nVehicles imported: {imported}")
        return imported

    def import_loans(filepath, tenant_id):
        print(f"\n{'='*60}")
        print(f"Importing Loans from: {filepath}")
        print(f"{'='*60}")

        wb = load_workbook(filepath)
        loan_sheet = None

        for sheet_name in wb.sheetnames:
            if 'loan' in sheet_name.lower():
                loan_sheet = wb[sheet_name]
                break

        if not loan_sheet and len(wb.sheetnames) > 1:
            loan_sheet = wb[wb.sheetnames[1]]

        if not loan_sheet:
            print("No loan sheet found")
            return 0

        headers = [cell.value for cell in loan_sheet[1]]
        print(f"Sheet: {loan_sheet.title}")
        print(f"Headers: {headers}")

        imported = 0

        for row_idx, row in enumerate(loan_sheet.iter_rows(min_row=2, max_row=loan_sheet.max_row, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))

                vehicle_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number'))
                lender_name = safe_str(data.get('LENDER NAME') or data.get('Lender') or data.get('Bank Name'))

                if not vehicle_number and not lender_name:
                    continue

                vehicle = None
                if vehicle_number:
                    vehicle = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=vehicle_number).first()

                principal = safe_float(data.get('PRINCIPAL AMOUNT') or data.get('Principal') or data.get('Loan Amount'))

                existing_count = Loan.query.filter_by(tenant_id=tenant_id).count()
                loan_name = f"LOAN-{existing_count + imported + 1:05d}"

                loan = Loan(
                    tenant_id=tenant_id,
                    name=loan_name,
                    vehicle_id=vehicle.id if vehicle else None,
                    loan_type=safe_str(data.get('LOAN TYPE') or data.get('Loan Type'), 50) or 'Vehicle Loan',
                    lender_name=lender_name or 'Unknown Lender',
                    principal_amount=principal,
                    interest_rate=safe_float(data.get('INTEREST RATE') or data.get('Interest Rate')),
                    emi_amount=safe_float(data.get('EMI AMOUNT') or data.get('EMI')),
                    total_loan_amount=principal,
                    remaining_balance=principal,
                    status='active',
                )

                db.session.add(loan)
                imported += 1
                print(f"  Imported: {loan_name} - {lender_name or 'N/A'} (₹{principal:,.2f})")

            except Exception as e:
                print(f"  ERROR Row {row_idx}: {e}")

        db.session.commit()
        print(f"\nLoans imported: {imported}")
        return imported

    def import_rate_list(filepath, tenant_id):
        print(f"\n{'='*60}")
        print(f"Importing Rate List from: {filepath}")
        print(f"{'='*60}")

        wb = load_workbook(filepath)

        imported = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if ws.max_row < 2:
                continue

            headers = [cell.value for cell in ws[1]]
            print(f"\nSheet: {sheet_name}")
            print(f"Headers: {headers}")

            truck_size = sheet_name.strip() if any(x in sheet_name.lower() for x in ['wheel', 'tyre', 'truck', '10', '12', '14', '16', '18', '20']) else None

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
                try:
                    data = dict(zip(headers, row))

                    location_name = safe_str(data.get('LOCATION') or data.get('Location') or data.get('DESTINATION') or data.get('To'))
                    if not location_name:
                        continue

                    rate = safe_float(data.get('RATE') or data.get('Rate') or data.get('PRICE') or data.get('Price') or data.get('FREIGHT'))
                    distance = safe_float(data.get('DISTANCE') or data.get('Distance') or data.get('KM'))

                    city = safe_str(data.get('CITY') or data.get('City'), 100)
                    state = safe_str(data.get('STATE') or data.get('State'), 100)

                    full_location = location_name
                    if truck_size and truck_size not in location_name:
                        full_location = f"{location_name} ({truck_size})"

                    existing = Location.query.filter_by(tenant_id=tenant_id, location=full_location).first()
                    if existing:
                        print(f"  Skipping (exists): {full_location}")
                        continue

                    location = Location(
                        tenant_id=tenant_id,
                        location=full_location,
                        city=city or location_name,
                        state=state,
                        distance_km=distance if distance > 0 else None,
                        rate=rate if rate > 0 else None,
                        remarks=truck_size,
                    )

                    db.session.add(location)
                    imported += 1
                    if rate > 0:
                        print(f"  Imported: {full_location} - ₹{rate:,.2f}")
                    else:
                        print(f"  Imported: {full_location}")

                except Exception as e:
                    print(f"  ERROR Sheet {sheet_name}, Row {row_idx}: {e}")

        db.session.commit()
        print(f"\nRate list imported: {imported}")
        return imported

    # Main execution
    print("="*60)
    print("TRANSPORT MANAGEMENT SYSTEM - EXCEL DATA IMPORT")
    print("="*60)

    tenant = get_or_create_tenant()
    tenant_id = tenant.id
    print(f"\nUsing Tenant: {tenant.name} (ID: {tenant_id})")

    stats = {}

    # Import Vendors
    if os.path.exists('Vendor Details.xlsx'):
        stats['vendors'] = import_vendors('Vendor Details.xlsx', tenant_id)
    else:
        print("Vendor Details.xlsx not found")
        stats['vendors'] = 0

    # Import Vehicles
    if os.path.exists('Vehicle and Loan Details.xlsx'):
        stats['vehicles'] = import_vehicles('Vehicle and Loan Details.xlsx', tenant_id)
        stats['loans'] = import_loans('Vehicle and Loan Details.xlsx', tenant_id)
    else:
        print("Vehicle and Loan Details.xlsx not found")
        stats['vehicles'] = 0
        stats['loans'] = 0

    # Import Rate List
    if os.path.exists('ALL PARTY RATE FILE.xlsx'):
        stats['rate_list'] = import_rate_list('ALL PARTY RATE FILE.xlsx', tenant_id)
    else:
        print("ALL PARTY RATE FILE.xlsx not found")
        stats['rate_list'] = 0

    # Summary
    print("\n" + "="*60)
    print("IMPORT COMPLETE - FINAL SUMMARY")
    print("="*60)
    for module, count in stats.items():
        print(f"  {module.capitalize():12}: {count} imported")

    print(f"\nCompleted at: {datetime.now()}")




# Subscription Plan Management Routes

@app.route("/subscription-plans")
@superadmin_required
def subscription_plans():
    plans = SubscriptionPlan.query.order_by(SubscriptionPlan.monthly_price).all()
    return render_template("subscriptions/plans.html", plans=plans)


@app.route("/subscription-plans/create", methods=["GET", "POST"])
@superadmin_required
def create_subscription_plan():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        monthly_price = request.form.get("monthly_price", "").strip()
        annual_price = request.form.get("annual_price", "").strip()
        max_vehicles = request.form.get("max_vehicles", "").strip()
        max_drivers = request.form.get("max_drivers", "").strip()
        max_users = request.form.get("max_users", "").strip()
        max_storage_gb = request.form.get("max_storage_gb", "").strip()
        features = request.form.get("features", "").strip()
        
        if not all([name, monthly_price, annual_price, max_vehicles, max_drivers, max_users, max_storage_gb]):
            flash("All required fields must be filled.", "error")
            return render_template("subscriptions/plan_form.html")
        
        try:
            plan = SubscriptionPlan(
                name=name,
                description=description or None,
                monthly_price=float(monthly_price),
                annual_price=float(annual_price),
                max_vehicles=int(max_vehicles),
                max_drivers=int(max_drivers),
                max_users=int(max_users),
                max_storage_gb=int(max_storage_gb),
                features=features or None,
                is_active=True
            )
            db.session.add(plan)
            db.session.commit()
            flash(f"Subscription plan '{name}' created successfully.", "success")
            return redirect(url_for("subscription_plans"))
        except ValueError as e:
            flash("Invalid numeric values provided.", "error")
        except Exception as e:
            flash(f"Error creating plan: {str(e)}", "error")
    
    return render_template("subscriptions/plan_form.html")


@app.route("/subscription-plans/<int:id>/edit", methods=["GET", "POST"])
@superadmin_required
def edit_subscription_plan(id):
    plan = SubscriptionPlan.query.get_or_404(id)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        monthly_price = request.form.get("monthly_price", "").strip()
        annual_price = request.form.get("annual_price", "").strip()
        max_vehicles = request.form.get("max_vehicles", "").strip()
        max_drivers = request.form.get("max_drivers", "").strip()
        max_users = request.form.get("max_users", "").strip()
        max_storage_gb = request.form.get("max_storage_gb", "").strip()
        features = request.form.get("features", "").strip()
        is_active = request.form.get("is_active") == "on"
        
        if not all([name, monthly_price, annual_price, max_vehicles, max_drivers, max_users, max_storage_gb]):
            flash("All required fields must be filled.", "error")
            return render_template("subscriptions/plan_form.html", plan=plan, edit=True)
        
        try:
            plan.name = name
            plan.description = description or None
            plan.monthly_price = float(monthly_price)
            plan.annual_price = float(annual_price)
            plan.max_vehicles = int(max_vehicles)
            plan.max_drivers = int(max_drivers)
            plan.max_users = int(max_users)
            plan.max_storage_gb = int(max_storage_gb)
            plan.features = features or None
            plan.is_active = is_active
            db.session.commit()
            flash(f"Subscription plan '{name}' updated successfully.", "success")
            return redirect(url_for("subscription_plans"))
        except ValueError as e:
            flash("Invalid numeric values provided.", "error")
        except Exception as e:
            flash(f"Error updating plan: {str(e)}", "error")
    
    return render_template("subscriptions/plan_form.html", plan=plan, edit=True)


@app.route("/subscription-plans/<int:id>/delete", methods=["POST"])
@superadmin_required
def delete_subscription_plan(id):
    plan = SubscriptionPlan.query.get_or_404(id)
    
    # Check if plan is in use
    in_use = TenantSubscription.query.filter_by(plan_id=id).first()
    if in_use:
        flash("Cannot delete plan that is currently assigned to tenants.", "error")
        return redirect(url_for("subscription_plans"))
    
    plan_name = plan.name
    db.session.delete(plan)
    db.session.commit()
    flash(f"Subscription plan '{plan_name}' deleted successfully.", "success")
    return redirect(url_for("subscription_plans"))


# Tenant Subscription Management Routes

@app.route("/tenant-subscriptions")
@superadmin_required
def tenant_subscriptions():
    subscriptions = TenantSubscription.query.order_by(TenantSubscription.expiry_date.desc()).all()
    return render_template("subscriptions/tenant_subscriptions.html", subscriptions=subscriptions)


@app.route("/tenants/<int:tenant_id>/subscription", methods=["GET", "POST"])
@superadmin_required
def assign_tenant_subscription(tenant_id):
    tenant = Tenant.query.get_or_404(tenant_id)
    plans = SubscriptionPlan.query.filter_by(is_active=True).all()
    
    # Get existing subscription
    existing_subscription = TenantSubscription.query.filter_by(tenant_id=tenant_id).first()
    
    if request.method == "POST":
        plan_id = request.form.get("plan_id")
        billing_cycle = request.form.get("billing_cycle", "monthly")
        start_date_str = request.form.get("start_date")
        custom_monthly_price = request.form.get("custom_monthly_price", "").strip()
        custom_annual_price = request.form.get("custom_annual_price", "").strip()
        discount_percentage = request.form.get("discount_percentage", "0").strip()
        auto_renew = request.form.get("auto_renew") == "on"
        payment_method = request.form.get("payment_method", "manual")
        payment_notes = request.form.get("payment_notes", "").strip()
        
        if not start_date_str:
            flash("Start date is required.", "error")
            return render_template("subscriptions/assign_subscription.html", tenant=tenant, plans=plans, existing_subscription=existing_subscription)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        
        # Calculate expiry date based on billing cycle
        if billing_cycle == "monthly":
            expiry_date = start_date.replace(day=1) + timedelta(days=32)
            expiry_date = expiry_date.replace(day=1) - timedelta(days=1)
        elif billing_cycle == "quarterly":
            expiry_date = start_date + timedelta(days=90)
        else:  # annual
            expiry_date = start_date.replace(year=start_date.year + 1) - timedelta(days=1)
        
        # Calculate price
        plan = SubscriptionPlan.query.get(plan_id) if plan_id else None
        if plan:
            if billing_cycle == "monthly":
                base_price = plan.monthly_price
            elif billing_cycle == "quarterly":
                base_price = plan.monthly_price * 3
            else:  # annual
                base_price = plan.annual_price
            
            # Apply discount
            discount = float(discount_percentage) if discount_percentage else 0
            final_price = base_price * (1 - discount / 100)
        else:
            final_price = 0
        
        # Update existing or create new subscription
        if existing_subscription:
            existing_subscription.plan_id = plan_id
            existing_subscription.billing_cycle = billing_cycle
            existing_subscription.start_date = start_date
            existing_subscription.expiry_date = expiry_date
            existing_subscription.custom_monthly_price = float(custom_monthly_price) if custom_monthly_price else None
            existing_subscription.custom_annual_price = float(custom_annual_price) if custom_annual_price else None
            existing_subscription.discount_percentage = float(discount_percentage) if discount_percentage else 0
            existing_subscription.auto_renew = auto_renew
            existing_subscription.payment_method = payment_method
            existing_subscription.payment_notes = payment_notes or None
            existing_subscription.status = "active"
            existing_subscription.next_payment_date = expiry_date
            db.session.commit()
            flash(f"Subscription updated for tenant '{tenant.name}'.", "success")
        else:
            subscription = TenantSubscription(
                tenant_id=tenant_id,
                plan_id=plan_id,
                billing_cycle=billing_cycle,
                start_date=start_date,
                expiry_date=expiry_date,
                custom_monthly_price=float(custom_monthly_price) if custom_monthly_price else None,
                custom_annual_price=float(custom_annual_price) if custom_annual_price else None,
                discount_percentage=float(discount_percentage) if discount_percentage else 0,
                auto_renew=auto_renew,
                payment_method=payment_method,
                payment_notes=payment_notes or None,
                status="active",
                next_payment_date=expiry_date
            )
            db.session.add(subscription)
            db.session.commit()
            flash(f"Subscription assigned to tenant '{tenant.name}'.", "success")
        
        return redirect(url_for("tenant_subscriptions"))
    
    return render_template("subscriptions/assign_subscription.html", tenant=tenant, plans=plans, existing_subscription=existing_subscription)


@app.route("/subscriptions/<int:id>/suspend", methods=["POST"])
@superadmin_required
def suspend_subscription(id):
    subscription = TenantSubscription.query.get_or_404(id)
    subscription.status = "suspended"
    db.session.commit()
    flash(f"Subscription for '{subscription.tenant.name}' has been suspended.", "success")
    return redirect(url_for("tenant_subscriptions"))


@app.route("/subscriptions/<int:id>/activate", methods=["POST"])
@superadmin_required
def activate_subscription(id):
    subscription = TenantSubscription.query.get_or_404(id)
    subscription.status = "active"
    db.session.commit()
    flash(f"Subscription for '{subscription.tenant.name}' has been activated.", "success")
    return redirect(url_for("tenant_subscriptions"))


@app.route("/subscriptions/<int:id>/payment", methods=["GET", "POST"])
@superadmin_required
def record_subscription_payment(id):
    subscription = TenantSubscription.query.get_or_404(id)
    
    if request.method == "POST":
        amount = request.form.get("amount", "").strip()
        payment_date_str = request.form.get("payment_date")
        payment_method = request.form.get("payment_method", "manual")
        transaction_id = request.form.get("transaction_id", "").strip()
        notes = request.form.get("notes", "").strip()
        
        if not all([amount, payment_date_str]):
            flash("Amount and payment date are required.", "error")
            return render_template("subscriptions/payment_form.html", subscription=subscription)
        
        try:
            payment = SubscriptionPayment(
                subscription_id=id,
                amount=float(amount),
                payment_date=datetime.strptime(payment_date_str, '%Y-%m-%d').date(),
                payment_method=payment_method,
                transaction_id=transaction_id or None,
                notes=notes or None
            )
            db.session.add(payment)
            
            # Update subscription next payment date
            if subscription.billing_cycle == "monthly":
                subscription.next_payment_date = subscription.expiry_date + timedelta(days=32)
                subscription.next_payment_date = subscription.next_payment_date.replace(day=1) - timedelta(days=1)
            elif subscription.billing_cycle == "quarterly":
                subscription.next_payment_date = subscription.expiry_date + timedelta(days=90)
            else:  # annual
                subscription.next_payment_date = subscription.expiry_date.replace(year=subscription.expiry_date.year + 1)
            
            db.session.commit()
            flash("Payment recorded successfully.", "success")
            return redirect(url_for("tenant_subscriptions"))
        except ValueError as e:
            flash("Invalid values provided.", "error")
        except Exception as e:
            flash(f"Error recording payment: {str(e)}", "error")
    
    return render_template("subscriptions/payment_form.html", subscription=subscription)


# Subscription Usage Tracking Functions

def update_subscription_usage(tenant_id):
    """Update usage counts for a tenant's subscription"""
    subscription = TenantSubscription.query.filter_by(tenant_id=tenant_id).first()
    if not subscription:
        return
    
    # Count current usage
    vehicle_count = Vehicle.query.filter_by(tenant_id=tenant_id).count()
    driver_count = Driver.query.filter_by(tenant_id=tenant_id).count()
    user_count = User.query.filter_by(tenant_id=tenant_id).count()
    
    # Update subscription
    subscription.current_vehicles = vehicle_count
    subscription.current_drivers = driver_count
    subscription.current_users = user_count
    db.session.commit()


def check_subscription_limits(tenant_id):
    """Check if tenant is within subscription limits"""
    subscription = TenantSubscription.query.filter_by(tenant_id=tenant_id).first()
    if not subscription or not subscription.plan:
        return True  # No subscription, allow everything
    
    vehicle_count = Vehicle.query.filter_by(tenant_id=tenant_id).count()
    driver_count = Driver.query.filter_by(tenant_id=tenant_id).count()


# =============================================================================
# Customer Credit Management Routes
# =============================================================================

@app.route("/customers/<int:id>/credit")
@permission_required("vendors", "view")
def customer_credit_management(id):
    """Customer credit management dashboard"""
    from datetime import date, timedelta
    
    customer = get_scoped_record(Vendor, id)
    
    # Get or create credit info
    credit_info = CustomerCredit.query.filter_by(vendor_id=id).first()
    if not credit_info:
        credit_info = CustomerCredit(
            vendor_id=id,
            tenant_id=get_current_tenant_id(),
            credit_limit=0,
            credit_period_days=30,
            payment_terms='Net 30'
        )
        db.session.add(credit_info)
        db.session.commit()
    
    # Get recent transactions
    transactions = CustomerTransaction.query.filter_by(vendor_id=id).order_by(
        CustomerTransaction.transaction_date.desc()
    ).limit(10).all()
    
    # Calculate aging
    today = date.today()
    aging_30 = 0
    aging_60 = 0
    aging_90 = 0
    aging_90_plus = 0
    
    for transaction in transactions:
        if transaction.transaction_type == 'invoice':
            days_overdue = (today - transaction.due_date).days if transaction.due_date else 0
            if days_overdue <= 30:
                aging_30 += transaction.amount
            elif days_overdue <= 60:
                aging_60 += transaction.amount
            elif days_overdue <= 90:
                aging_90 += transaction.amount
            else:
                aging_90_plus += transaction.amount
    
    return render_template(
        "customers/credit_management.html",
        customer=customer,
        credit_info=credit_info,
        transactions=transactions,
        aging_30=aging_30,
        aging_60=aging_60,
        aging_90=aging_90,
        aging_90_plus=aging_90_plus
    )


@app.route("/customers/<int:id>/credit/update", methods=["POST"])
@permission_required("vendors", "edit")
def update_customer_credit(id):
    """Update customer credit information"""
    customer = get_scoped_record(Vendor, id)
    
    credit_info = CustomerCredit.query.filter_by(vendor_id=id).first()
    if not credit_info:
        credit_info = CustomerCredit(
            vendor_id=id,
            tenant_id=get_current_tenant_id()
        )
        db.session.add(credit_info)
    
    # Update credit info
    credit_info.credit_limit = float(request.form.get('credit_limit', 0))
    credit_info.credit_period_days = int(request.form.get('credit_period_days', 30))
    credit_info.payment_terms = request.form.get('payment_terms', 'Net 30')
    credit_info.is_credit_hold = 'credit_hold' in request.form
    credit_info.hold_reason = request.form.get('hold_reason', '')
    credit_info.credit_status = request.form.get('credit_status', 'active')
    
    db.session.commit()
    
    flash('Customer credit information updated successfully!', 'success')
    return redirect(url_for('customer_credit_management', id=id))


@app.route("/customers/aging-report")
@permission_required("vendors", "view")
def customer_aging_report():
    """Master aging report for all customers"""
    from datetime import date, timedelta
    
    customers = scoped_query(Vendor).all()
    aging_data = []
    
    for customer in customers:
        credit_info = CustomerCredit.query.filter_by(vendor_id=customer.id).first()
        
        # Calculate aging for this customer
        aging = calculate_customer_aging(customer.id)
        
        aging_data.append({
            'customer': customer,
            'credit_info': credit_info,
            'aging': aging
        })
    
    return render_template(
        "customers/aging_report.html",
        aging_data=aging_data,
        total_customers=len(customers)
    )


def calculate_customer_aging(vendor_id):
    """Calculate aging buckets for a customer"""
    from datetime import date
    
    today = date.today()
    transactions = CustomerTransaction.query.filter_by(
        vendor_id=vendor_id,
        transaction_type='invoice'
    ).all()
    
    aging = {
        'current': 0,      # 0-30 days
        'days_31_60': 0,   # 31-60 days
        'days_61_90': 0,   # 61-90 days
        'days_90_plus': 0,   # 90+ days
        'total': 0
    }
    
    for transaction in transactions:
        if not transaction.due_date:
            continue
            
        days_overdue = (today - transaction.due_date).days
        amount = float(transaction.amount)
        
        if days_overdue <= 30:
            aging['current'] += amount
        elif days_overdue <= 60:
            aging['days_31_60'] += amount
        elif days_overdue <= 90:
            aging['days_61_90'] += amount
        else:
            aging['days_90_plus'] += amount
        
        aging['total'] += amount
    
    return aging


# =============================================================================
# Customer Classification System Routes
# =============================================================================

@app.route("/customers/classification")
@permission_required("vendors", "view")
def customer_classification():
    """Customer classification management interface"""
    categories = CustomerCategory.query.filter_by(
        tenant_id=get_current_tenant_id(),
        is_active=True
    ).order_by(CustomerCategory.sort_order).all()
    
    return render_template(
        "customers/classification.html",
        categories=categories
    )


@app.route("/customers/classification/create", methods=["POST"])
@permission_required("vendors", "edit")
def create_customer_category():
    """Create new customer category"""
    category = CustomerCategory(
        tenant_id=get_current_tenant_id(),
        category_name=request.form.get('category_name'),
        category_code=request.form.get('category_code'),
        description=request.form.get('description'),
        min_credit_limit=float(request.form.get('min_credit_limit', 0)),
        max_credit_limit=float(request.form.get('max_credit_limit', 0)),
        default_payment_terms=request.form.get('default_payment_terms', 'Net 30'),
        service_priority=request.form.get('service_priority', 'medium'),
        sla_hours=int(request.form.get('sla_hours', 48)),
        sort_order=int(request.form.get('sort_order', 0))
    )
    
    db.session.add(category)
    db.session.commit()
    
    flash('Customer category created successfully!', 'success')
    return redirect(url_for('customer_classification'))


@app.route("/customers/<int:id>/classify", methods=["POST"])
@permission_required("vendors", "edit")
def classify_customer(id):
    """Classify individual customer"""
    customer = get_scoped_record(Vendor, id)
    
    # Update classification
    customer.customer_type = request.form.get('customer_type', 'regular')
    customer.customer_tier = request.form.get('customer_tier', 'bronze')
    customer.customer_segment = request.form.get('customer_segment')
    customer.customer_lifecycle_status = request.form.get('lifecycle_status', 'active')
    customer.classification_notes = request.form.get('classification_notes')
    customer.classification_date = date.today()
    customer.classified_by = session['user_id']
    
    # Auto-adjust credit limit based on category
    category = CustomerCategory.query.filter_by(
        category_code=customer.customer_tier
    ).first()
    
    if category and not customer.credit_info:
        customer.credit_info.credit_limit = category.max_credit_limit
        db.session.add(customer.credit_info)
    
    db.session.commit()
    
    flash('Customer classified successfully!', 'success')
    return redirect(url_for('edit_vendor', id=id))


def get_customer_classification_badge(customer_type, customer_tier):
    """Generate appropriate badge for customer classification"""
    type_colors = {
        'regular': 'primary',
        'premium': 'info',
        'vip': 'warning',
        'one_time': 'secondary'
    }
    
    tier_colors = {
        'bronze': 'secondary',
        'silver': 'light',
        'gold': 'warning',
        'platinum': 'success'
    }
    
    return {
        'type_badge': f'<span class="badge bg-{type_colors.get(customer_type, "secondary")}">{customer_type.title()}</span>',
        'tier_badge': f'<span class="badge bg-{tier_colors.get(customer_tier, "secondary")}">{customer_tier.title()}</span>'
    }


def auto_classify_customer(customer):
    """Automatically classify customer based on transaction history"""
    if not customer.transactions:
        return 'one_time'
    
    # Count transactions in last 12 months
    recent_transactions = len([t for t in customer.transactions 
                             if t.transaction_date >= (date.today() - timedelta(days=365))])
    
    if recent_transactions > 50:
        return 'vip'
    elif recent_transactions > 20:
        return 'premium'
    elif recent_transactions > 5:
        return 'regular'
    else:
        return 'one_time'


# =============================================================================
# Customer Communication Hub Routes
# =============================================================================

@app.route("/customers/communications")
@permission_required("vendors", "view")
def customer_communications():
    """Customer communication history interface"""
    vendor_id = request.args.get('vendor_id', 0, type=int)
    customer = get_scoped_record(Vendor, vendor_id) if vendor_id else None
    
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('vendors'))
    
    # Get communications with filters
    communications = CustomerCommunication.query.filter_by(vendor_id=customer.id).order_by(
        CustomerCommunication.communication_date.desc()
    ).all()
    
    # Get feedback
    feedback = CustomerFeedback.query.filter_by(vendor_id=customer.id).order_by(
        CustomerFeedback.feedback_date.desc()
    ).all()
    
    return render_template(
        "customers/communications.html",
        customer=customer,
        communications=communications,
        feedback=feedback
    )


@app.route("/customers/<int:id>/add-communication", methods=["POST"])
@permission_required("vendors", "edit")
def add_customer_communication(id):
    """Add new customer communication"""
    customer = get_scoped_record(Vendor, id)
    
    communication = CustomerCommunication(
        tenant_id=get_current_tenant_id(),
        vendor_id=customer.id,
        communication_type=request.form.get('communication_type'),
        subject=request.form.get('subject'),
        message=request.form.get('message'),
        direction='outbound',
        priority=request.form.get('priority', 'medium'),
        status=request.form.get('status', 'open'),
        communicated_by=session['user_id'],
        next_followup=datetime.strptime(request.form.get('next_followup'), '%Y-%m-%d') if request.form.get('next_followup') else None
    )
    
    db.session.add(communication)
    db.session.commit()
    
    flash('Communication added successfully!', 'success')
    return redirect(url_for('customer_communications', vendor_id=id))


@app.route("/customers/<int:id>/feedback")
@permission_required("vendors", "view")
def customer_feedback():
    """Customer feedback management interface"""
    vendor_id = request.args.get('vendor_id', 0, type=int)
    customer = get_scoped_record(Vendor, vendor_id) if vendor_id else None
    
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('vendors'))
    
    # Get feedback
    feedback = CustomerFeedback.query.filter_by(vendor_id=customer.id).order_by(
        CustomerFeedback.feedback_date.desc()
    ).all()
    
    return render_template(
        "customers/feedback.html",
        customer=customer,
        feedback=feedback
    )


@app.route("/customers/<int:id>/add-feedback", methods=["POST"])
@permission_required("vendors", "edit")
def add_customer_feedback(id):
    """Add new customer feedback"""
    customer = get_scoped_record(Vendor, id)
    
    feedback = CustomerFeedback(
        tenant_id=get_current_tenant_id(),
        vendor_id=customer.id,
        feedback_type=request.form.get('feedback_type'),
        rating=int(request.form.get('rating', 0)),
        feedback_text=request.form.get('feedback_text'),
        resolution_status=request.form.get('resolution_status', 'pending')
    )
    
    db.session.add(feedback)
    db.session.commit()
    
    flash('Feedback recorded successfully!', 'success')
    return redirect(url_for('customer_feedback', vendor_id=id))


@app.route("/customers/automated-reminders")
@permission_required("vendors", "view")
def customer_automated_reminders():
    """Automated reminder settings interface"""
    return render_template("customers/automated_reminders.html")


# =============================================================================
# Advanced Customer Analytics Routes
# =============================================================================

@app.route("/customers/analytics/dashboard")
@permission_required("vendors", "view")
def customer_analytics_dashboard():
    """Customer analytics dashboard with KPIs and insights"""
    customers = scoped_query(Vendor).all()
    
    # Calculate overall metrics
    total_customers = len(customers)
    total_revenue = sum(credit.current_outstanding or 0 for credit in [c.credit_info for c in customers if c.credit_info])
    high_value_customers = len([c for c in customers if c.customer_tier in ['gold', 'platinum']])
    
    return render_template(
        "customers/analytics_dashboard.html",
        customers=customers,
        total_customers=total_customers,
        total_revenue=total_revenue,
        high_value_customers=high_value_customers
    )


@app.route("/customers/<int:id>/analytics")
@permission_required("vendors", "view")
def customer_analytics_detail():
    """Detailed analytics for individual customer"""
    customer = get_scoped_record(Vendor, id)
    
    # Get period filters
    period_type = request.args.get('period', 'monthly')
    period_end = date.today()
    
    if period_type == 'monthly':
        period_start = period_end - timedelta(days=30)
    elif period_type == 'quarterly':
        period_start = period_end - timedelta(days=90)
    elif period_type == 'yearly':
        period_start = period_end - timedelta(days=365)
    else:
        period_start = period_end - timedelta(days=30)
    
    # Calculate analytics
    analytics = calculate_customer_analytics(id, period_start, period_end)
    
    # Get historical data for trends
    historical_data = CustomerAnalytics.query.filter_by(vendor_id=id).order_by(
        CustomerAnalytics.period_start.desc()
    ).limit(12).all()  # Last 12 periods
    
    return render_template(
        "customers/analytics_detail.html",
        customer=customer,
        analytics=analytics,
        historical_data=historical_data,
        period_type=period_type
    )


@app.route("/customers/analytics/performance")
@permission_required("vendors", "view")
def customer_performance_report():
    """Customer performance comparison report"""
    customers = scoped_query(Vendor).all()
    customer_data = []
    
    for customer in customers:
        # Get latest analytics
        latest_analytics = CustomerAnalytics.query.filter_by(vendor_id=customer.id).order_by(
            CustomerAnalytics.period_start.desc()
        ).first()
        
        if latest_analytics:
            customer_data.append({
                'customer': customer,
                'analytics': latest_analytics,
                'rank': 0  # Will be calculated
            })
    
    # Sort by revenue
    customer_data.sort(key=lambda x: x['analytics'].total_revenue, reverse=True)
    
    # Assign ranks
    for i, data in enumerate(customer_data, 1):
        data['rank'] = i
    
    return render_template(
        "customers/performance_report.html",
        customer_data=customer_data
    )


def calculate_customer_analytics(vendor_id, period_start, period_end):
    """Calculate comprehensive analytics for a customer"""
    from sqlalchemy import func
    
    # Get transport bills for the period
    bills = TransportBill.query.filter(
        TransportBill.party_information == str(vendor_id),  # Assuming party_information stores vendor_id
        TransportBill.date.between(period_start, period_end)
    ).all()
    
    # Get payments for the period
    payments = CustomerTransaction.query.filter(
        CustomerTransaction.vendor_id == vendor_id,
        CustomerTransaction.transaction_type == 'payment',
        CustomerTransaction.transaction_date.between(period_start, period_end)
    ).all()
    
    # Get feedback for the period
    feedback = CustomerFeedback.query.filter(
        CustomerFeedback.vendor_id == vendor_id,
        CustomerFeedback.feedback_date.between(period_start, period_end)
    ).all()
    
    # Calculate metrics
    total_revenue = sum(float(b.rate) for b in bills)
    total_bills = len(bills)
    total_deliveries = len(bills)  # Simplified: each bill = one delivery
    on_time_deliveries = total_deliveries  # Simplified for now
    total_payments = sum(float(t.amount) for t in payments)
    outstanding_balance = total_revenue - total_payments
    
    # Calculate satisfaction metrics
    total_feedback = len(feedback)
    positive_feedback = len([f for f in feedback if f.rating and f.rating >= 4])
    satisfaction_score = (sum(f.rating or 0 for f in feedback) / total_feedback) if total_feedback > 0 else 0
    positive_feedback_pct = (positive_feedback / total_feedback * 100) if total_feedback > 0 else 0
    
    # Calculate delivery performance
    on_time_rate = 100.0  # Simplified for now
    avg_order_value = total_revenue / total_bills if total_bills > 0 else 0
    
    # Calculate lifetime value (simplified)
    lifetime_value = total_revenue * 0.2  # 20% of revenue as estimated lifetime value
    
    # Calculate churn risk (simplified)
    churn_probability = 5.0  # Default low risk
    if on_time_rate < 80:
        churn_probability = 20.0  # High risk
    elif on_time_rate < 90:
        churn_probability = 10.0  # Medium risk
    
    # Calculate growth rate (month-over-month)
    previous_period_revenue = get_previous_period_revenue(vendor_id, period_start)
    growth_rate = ((total_revenue - previous_period_revenue) / previous_period_revenue * 100) if previous_period_revenue > 0 else 0
    
    return {
        'total_revenue': total_revenue,
        'total_bills': total_bills,
        'total_deliveries': total_deliveries,
        'on_time_deliveries': on_time_deliveries,
        'on_time_delivery_rate': on_time_rate,
        'avg_order_value': avg_order_value,
        'total_payments': total_payments,
        'outstanding_balance': outstanding_balance,
        'satisfaction_score': satisfaction_score,
        'total_feedback': total_feedback,
        'positive_feedback_percentage': positive_feedback_pct,
        'customer_lifetime_value': lifetime_value,
        'churn_probability': churn_probability,
        'growth_rate': growth_rate
    }


def get_previous_period_revenue(vendor_id, current_period_start):
    """Get revenue from previous period for growth calculation"""
    previous_period_end = current_period_start - timedelta(days=30)
    revenue = db.session.query(func.sum(TransportBill.rate)).filter(
        TransportBill.party_information == str(vendor_id),
        TransportBill.date.between(previous_period_end, current_period_start)
    ).scalar() or 0
    return revenue


# =============================================================================
# Enhanced Customer Portal Routes
# =============================================================================

@app.route("/customer-portal/dashboard")
def enhanced_customer_dashboard():
    """Enhanced customer dashboard with modern UI"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        flash('Please login to access your dashboard', 'error')
        return redirect(url_for('customer_login'))
    
    # Get customer data
    customer = account.vendor
    recent_bills = TransportBill.query.filter_by(
        party_information=str(customer.id)
    ).order_by(TransportBill.date.desc()).limit(10).all()
    
    # Get notifications
    notifications = CustomerNotification.query.filter_by(
        vendor_id=customer.id,
        is_read=False
    ).order_by(CustomerNotification.created_at.desc()).limit(5).all()
    
    # Get unread count
    unread_count = CustomerNotification.query.filter_by(
        vendor_id=customer.id,
        is_read=False
    ).count()
    
    return render_template(
        "customer_portal/enhanced_dashboard.html",
        account=account,
        customer=customer,
        recent_bills=recent_bills,
        notifications=notifications,
        unread_count=unread_count
    )


@app.route("/customer-portal/orders")
def customer_orders():
    """Customer order tracking interface"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return redirect(url_for('customer_login'))
    
    customer = account.vendor
    
    # Get filters
    status_filter = request.args.get('status', 'all')
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    
    # Build query
    query = TransportBill.query.filter_by(party_information=str(customer.id))
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if date_from:
        try:
            query = query.filter(TransportBill.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except:
            pass
    
    if date_to:
        try:
            query = query.filter(TransportBill.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except:
            pass
    
    orders = query.order_by(TransportBill.date.desc()).all()
    
    return render_template(
        "customer_portal/orders.html",
        account=account,
        customer=customer,
        orders=orders,
        status_filter=status_filter,
        date_from=date_from,
        date_to=date_to
    )


@app.route("/customer-portal/documents")
def customer_documents():
    """Customer document management interface"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return redirect(url_for('customer_login'))
    
    customer = account.vendor
    
    # Get documents
    documents = CustomerDocument.query.filter_by(vendor_id=customer.id).order_by(
        CustomerDocument.uploaded_at.desc()
    ).all()
    
    return render_template(
        "customer_portal/documents.html",
        account=account,
        customer=customer,
        documents=documents
    )


@app.route("/customer-portal/upload-document", methods=["POST"])
def upload_customer_document():
    """Handle customer document uploads"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return jsonify({'error': 'Unauthorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Simple file validation
    allowed_extensions = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png'}
    if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
        filename = secure_filename(file.filename)
        
        # Create document record
        document = CustomerDocument(
            tenant_id=account.tenant_id,
            vendor_id=account.vendor_id,
            document_type=request.form.get('document_type', 'other'),
            document_name=request.form.get('document_name', filename),
            file_path=filename,
            file_size=0,  # Would be set after actual file save
            mime_type=file.mimetype,
            uploaded_by=account.id
        )
        
        db.session.add(document)
        db.session.commit()
        
        return jsonify({'success': True, 'document_id': document.id})
    
    return jsonify({'error': 'File type not allowed'}), 400


@app.route("/customer-portal/profile")
def customer_profile():
    """Customer profile management interface"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return redirect(url_for('customer_login'))
    
    customer = account.vendor
    
    if request.method == 'POST':
        # Update profile information
        customer.contact_person = request.form.get('primary_contact_name')
        customer.mobile = request.form.get('primary_contact_phone')
        customer.email = request.form.get('primary_contact_email')
        
        # Update account preferences
        account.language_preference = request.form.get('language_preference', 'en')
        account.timezone_preference = request.form.get('timezone_preference', 'UTC')
        account.portal_theme = request.form.get('portal_theme', 'light')
        
        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('customer_profile'))
    
    return render_template(
        "customer_portal/profile.html",
        account=account,
        customer=customer
    )


@app.route("/customer-portal/notifications")
def customer_notifications():
    """Customer notification center"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return redirect(url_for('customer_login'))
    
    customer = account.vendor
    
    # Get notifications with pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    notifications = CustomerNotification.query.filter_by(vendor_id=customer.id).order_by(
        CustomerNotification.created_at.desc()
    ).paginate(page=page, per_page=per_page)
    
    return render_template(
        "customer_portal/notifications.html",
        account=account,
        customer=customer,
        notifications=notifications
    )


@app.route("/customer-portal/mark-notification-read/<int:notification_id>", methods=["POST"])
def mark_notification_read(notification_id):
    """Mark notification as read"""
    account = db.session.get(CustomerPortalAccount, session.get("customer_id"))
    if not account or not account.is_active:
        return jsonify({'error': 'Unauthorized'}), 401
    
    notification = CustomerNotification.query.filter_by(
        id=notification_id,
        vendor_id=account.vendor_id
    ).first()
    
    if notification:
        notification.is_read = True
        notification.read_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True})
    
    return jsonify({'error': 'Notification not found'}), 404
    user_count = User.query.filter_by(tenant_id=tenant_id).count()
    
    limits_exceeded = []
    
    if vehicle_count >= subscription.plan.max_vehicles:
        limits_exceeded.append(f"Vehicles ({vehicle_count}/{subscription.plan.max_vehicles})")
    if driver_count >= subscription.plan.max_drivers:
        limits_exceeded.append(f"Drivers ({driver_count}/{subscription.plan.max_drivers})")
    if user_count >= subscription.plan.max_users:
        limits_exceeded.append(f"Users ({user_count}/{subscription.plan.max_users})")
    
    return limits_exceeded if limits_exceeded else True

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
